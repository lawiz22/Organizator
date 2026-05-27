"""
Organizator — Organisateur intelligent de disques durs
Application NiceGUI avec workflow en 6 étapes guidées :
  1. Configuration sources/destination
  2. Scan et analyse des types de fichiers
  3. Sélection des catégories à conserver
  4. Décision sur les fichiers non catégorisés
  5. Aperçu et exécution de l'organisation
  6. Rapport et suivi des métadonnées
"""

import os
import sys
import asyncio
import gc
import time
import json
import sqlite3
import copy
import io
import base64
import shutil
import hashlib
import datetime
import threading
import concurrent.futures
from pathlib import Path

# ── Patch pywebview/NiceGUI native-mode race condition ───────────────────────
# window.events.moved fires on multiple threads simultaneously; concurrent
# _send_bytes() calls on the same multiprocessing Connection are unsupported.
# Adding a per-connection lock at the _send_bytes level silences the crash.
try:
    import multiprocessing.connection as _mp_conn
    _orig_send_bytes = _mp_conn.Connection._send_bytes

    def _locked_send_bytes(self, buf):
        if not hasattr(self, '_send_lock'):
            self._send_lock = threading.Lock()
        with self._send_lock:
            _orig_send_bytes(self, buf)

    _mp_conn.Connection._send_bytes = _locked_send_bytes
except Exception:
    pass
# ─────────────────────────────────────────────────────────────────────────────
from collections import defaultdict
from functools import lru_cache
from typing import Optional
from urllib.parse import quote, unquote
from urllib import request as urllib_request
from urllib import error as urllib_error
import subprocess
import argparse
import socket

from fastapi import HTTPException
from fastapi.responses import FileResponse

try:
    from dotenv import load_dotenv
    if os.path.exists('.env'):
        load_dotenv('.env')
except ImportError:
    pass

from nicegui import app, ui, run, background_tasks

try:
    from send2trash import send2trash
    HAS_SEND2TRASH = True
except Exception:
    HAS_SEND2TRASH = False

# ============================================================
# NOMENCLATURE DES CATÉGORIES
# ============================================================

DEFAULT_CATEGORIES = {
    "MUSIQUE-AUDIO": {
        "icon": "music_note",
        "color": "purple",
        "extensions": {
            ".mp3": "AUDIO-DIVERS",
            ".m4a": "AUDIO-DIVERS",
            ".aac": "AUDIO-DIVERS",
            ".ogg": "AUDIO-DIVERS",
            ".wma": "AUDIO-DIVERS",
            ".opus": "AUDIO-DIVERS",
            ".wav": "AUDIO-DIVERS",
            ".flac": "AUDIO-DIVERS",
            ".aiff": "AUDIO-DIVERS",
            ".aif": "AUDIO-DIVERS",
            ".alac": "AUDIO-DIVERS",
        }
    },
    "PHOTO-VIDEO": {
        "icon": "photo_library",
        "color": "teal",
        "extensions": {
            ".jpg": "Images",
            ".jpeg": "Images",
            ".png": "Images",
            ".webp": "Images",
            ".bmp": "Images",
            ".tiff": "Images",
            ".tif": "Images",
            ".raw": "Images",
            ".cr2": "Images",
            ".nef": "Images",
            ".arw": "Images",
            ".dng": "Images",
            ".heic": "Images",
            ".heif": "Images",
            ".mp4": "Video",
            ".avi": "Video",
            ".mov": "Video",
            ".mkv": "Video",
            ".webm": "Video",
            ".mts": "Video",
            ".m2ts": "Video",
            ".wmv": "Video",
            ".flv": "Video",
            ".3gp": "Video",
        }
    },
    "DOCUMENTS": {
        "icon": "description",
        "color": "blue",
        "extensions": {
            ".pdf": "PDF",
            ".doc": "DOC",
            ".docx": "DOCX",
            ".odt": "ODT",
            ".xls": "XLS",
            ".xlsx": "XLSX",
            ".ods": "ODS",
            ".ppt": "PPT",
            ".pptx": "PPTX",
            ".odp": "ODP",
            ".txt": "TXT",
            ".md": "TXT",
            ".rtf": "RTF",
            ".csv": "CSV",
            ".epub": "EPUB",
        }
    },
    "PROJETS": {
        "icon": "folder_special",
        "color": "orange",
        "extensions": {}  # Détection par signature de projet
    }
}

CATEGORIES = copy.deepcopy(DEFAULT_CATEGORIES)

# Signatures de projets : extension marqueur → (sous-catégorie, description)
DEFAULT_PROJECT_SIGNATURES = {
    # Projets audio - PRIORITÉ HAUTE (détectés en premier, fichiers audio internes exclus)
    ".als": ("Ableton-Live", "Projet Ableton Live"),
    ".alp": ("Ableton-Live", "Pack Ableton Live"),
    ".cpr": ("Cubase", "Projet Cubase"),
    ".npr": ("Nuendo", "Projet Nuendo"),
    ".psd": ("Photoshop", "Document Photoshop"),
    ".psb": ("Photoshop", "Document Photoshop (grand)"),
    ".aep": ("After-Effects", "Projet After Effects"),
    ".aet": ("After-Effects", "Template After Effects"),
    ".uproject": ("Unreal-Engine", "Projet Unreal Engine"),
}

PROJECT_SIGNATURES = dict(DEFAULT_PROJECT_SIGNATURES)

AUDIO_EXTENSIONS = {
    ".mp3", ".m4a", ".aac", ".ogg", ".wma", ".opus",
    ".wav", ".flac", ".aiff", ".aif", ".alac",
}

AUDIO_PROJECT_CATEGORIES = {"Ableton-Live", "Cubase", "Nuendo"}

RULES_FILE = "organizator_rules.json"

# Dossiers système à exclure du scan
EXCLUDED_DIRS = {
    "$recycle.bin", "system volume information", "$windows.~bt",
    "$windows.~ws", "windows", "program files", "program files (x86)",
    "programdata", "appdata", "node_modules", ".git", ".venv",
    "__pycache__", ".vs", ".idea", "thumbs.db"
}

MONTHS_FR = {
    1: "janvier", 2: "février", 3: "mars", 4: "avril",
    5: "mai", 6: "juin", 7: "juillet", 8: "août",
    9: "septembre", 10: "octobre", 11: "novembre", 12: "décembre"
}

INSTALLER_EXTENSIONS = {".msi", ".exe", ".zip", ".rar"}
ARCHIVE_IMAGE_EXTENSIONS = {
    ".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tiff", ".tif", ".heic", ".heif"
}

# ============================================================
# ÉTAT GLOBAL DE L'APPLICATION
# ============================================================

class OrganizerState:
    def __init__(self):
        self.current_step = 0           # Étape active (0-5)
        self.is_scanning = False
        self.is_organizing = False
        self.is_loading = False          # UI rebuild in progress (prevent WS timeout)
        self.cancel_flag = False

        # Config
        self.sources: list[str] = []    # Dossiers/disques sources
        self.destination: str = ""       # Répertoire destination
        self.mode: str = "copy"          # "copy" ou "move"
        self.delete_empty_sources_after_move: bool = False  # Supprimer sources vides en mode move

        # Résultats du scan
        self.scan_results: dict = {}     # {categorie: {sous_cat: [ScanEntry]}}
        self.scan_stats: dict = {}       # {categorie: {count, size}}
        self.project_roots: list = []    # Racines de projets détectés
        self.project_detection_evidence: dict = {}  # {root: {subcategory, marker_ext}}
        self.uncategorized: list = []    # Fichiers non catégorisés
        self.installer_files: list = []   # Fichiers exécutables et archives (.msi/.exe/.zip/.rar)

        # Sélections utilisateur
        self.selected_categories: dict = {}   # {(cat, sous_cat): bool}
        self.selected_files: dict = {}   # {path: bool} pour sélection individuelle
        self.uncategorized_decisions: dict = {}  # {path: "ignore"|"assign:CAT"}
        self.installer_decisions: dict = {}  # {path: "keep"|"trash"|"delete"|"assign:CAT"}

        # Métadonnées musicales éditées par l'utilisateur
        self.music_metadata_overrides: dict = {}  # {path: {"artist": ..., "album": ...}}
        self.video_output_overrides: dict = {}  # {path: {"folder": ..., "filename": ...}}
        self.video_skip_date_hierarchy: bool = False
        self.video_no_date_folder_name: str = "DiversVideo"

        # Reclassification manuelle des images IA ↔ Vraies Photos
        self.image_reclassify_overrides: dict = {}  # {path: "IA-Images" | "Images"}

        # NSFW — filtrage et destination alternative
        self.nsfw_destination: str = ""          # dossier destination pour images NSFW (si renseigné)
        self.nsfw_exclude: bool = False          # exclure les NSFW (SENSUEL+EXPLICIT) de l'organisation principale
        self.nsfw_filter_ui: str = "all"         # filtre UI: "all" | "sain" | "nsfw"

        # Validation IA / Archivage final
        self.llm_validate_aesthetic: bool = True
        self.llm_validate_nsfw: bool = True
        self.llm_validate_tags: bool = True
        self.archive_roots_text: str = ""

        # Progression
        self.progress: float = 0.0
        self.status_text: str = "Prêt"
        self.logs: list[str] = []
        self.full_log_history: list[str] = []

        # Rapport final
        self.report: dict = {}

    def add_log(self, msg: str):
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {msg}"
        print(line)
        self.logs.append(line)
        self.full_log_history.append(line)

    def reset_scan(self):
        self.scan_results.clear()
        self.scan_stats.clear()
        self.project_roots.clear()
        self.project_detection_evidence.clear()
        self.uncategorized.clear()
        self.installer_files.clear()
        self.selected_categories.clear()
        self.selected_files.clear()
        self.uncategorized_decisions.clear()
        self.installer_decisions.clear()
        self.music_metadata_overrides.clear()
        self.video_output_overrides.clear()
        self.image_reclassify_overrides.clear()
        self.nsfw_destination = ""
        self.nsfw_exclude = False
        self.nsfw_filter_ui = "all"
        self.logs.clear()
        self.full_log_history.clear()
        self.report.clear()
        self.progress = 0.0
        self.status_text = "Prêt"

    def reset_all(self):
        self.reset_scan()
        self.sources.clear()
        self.destination = ""
        self.mode = "copy"
        self.delete_empty_sources_after_move = False
        self.video_skip_date_hierarchy = False
        self.video_no_date_folder_name = "DiversVideo"


state = OrganizerState()

# Démarrage optionnel directement sur la section IA/LLM.
START_IN_LLM_MODE = False


def normalize_path_safe(path: str) -> str:
    try:
        return os.path.normcase(os.path.abspath(path or ""))
    except Exception:
        return ""


def split_archive_roots(raw: str) -> list[str]:
    values = []
    for part in (raw or "").replace(";", "\n").splitlines():
        p = part.strip().strip('"')
        if p:
            values.append(p)
    # unique, ordre conservé
    return list(dict.fromkeys(values))


def infer_default_archive_roots(destination: str, sources: list[str]) -> list[str]:
    roots = []
    if destination:
        roots.append(destination)
    for src in sources or []:
        drive = os.path.splitdrive(src or "")[0]
        if drive:
            roots.append(os.path.join(drive + os.sep, "00-Archives"))
    return list(dict.fromkeys([r for r in roots if r]))


def is_under_archive_roots(path: str, archive_roots: list[str]) -> bool:
    if not path:
        return False
    n_path = normalize_path_safe(path)
    if not n_path:
        return False
    for root in archive_roots or []:
        n_root = normalize_path_safe(root)
        if not n_root:
            continue
        if n_path == n_root or n_path.startswith(n_root + os.sep):
            return True
    return False


def is_image_path(path: str) -> bool:
    if not path:
        return False
    ext = os.path.splitext(path)[1].lower()
    return ext in ARCHIVE_IMAGE_EXTENSIONS


def _is_port_available(host: str, port: int) -> bool:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        try:
            sock.bind((host, port))
            return True
        except OSError:
            return False


def _find_available_port(host: str, start_port: int, attempts: int = 30) -> int:
    for candidate in range(start_port, start_port + attempts):
        if _is_port_available(host, candidate):
            return candidate
    raise RuntimeError(
        f"Aucun port disponible trouvé entre {start_port} et {start_port + attempts - 1} sur {host}."
    )


def normalize_extension(value: str) -> str:
    ext = (value or "").strip().lower()
    if not ext:
        return ""
    if not ext.startswith("."):
        ext = "." + ext
    return ext


def parse_extensions_csv(raw: str) -> list[str]:
    parts = [normalize_extension(x) for x in raw.replace(";", ",").split(",")]
    return [p for p in parts if p]


def load_custom_rules() -> None:
    """Charge les catégories/types projets persistants depuis le fichier local."""
    if not os.path.exists(RULES_FILE):
        return
    try:
        with open(RULES_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)

        for cat_name, cat_data in (data.get("categories") or {}).items():
            cat_name_u = str(cat_name).strip().upper()
            if not cat_name_u:
                continue
            if cat_name_u not in CATEGORIES:
                CATEGORIES[cat_name_u] = {
                    "icon": cat_data.get("icon", "category"),
                    "color": cat_data.get("color", "cyan"),
                    "extensions": {},
                }

            for ext, subcat in (cat_data.get("extensions") or {}).items():
                n_ext = normalize_extension(ext)
                if n_ext:
                    CATEGORIES[cat_name_u]["extensions"][n_ext] = str(subcat)

        for ext, payload in (data.get("project_signatures") or {}).items():
            n_ext = normalize_extension(ext)
            if not n_ext:
                continue
            if isinstance(payload, list) and len(payload) >= 2:
                PROJECT_SIGNATURES[n_ext] = (str(payload[0]), str(payload[1]))
            elif isinstance(payload, dict):
                sub = str(payload.get("subcategory", "Projet"))
                desc = str(payload.get("description", f"Projet {sub}"))
                PROJECT_SIGNATURES[n_ext] = (sub, desc)
    except Exception as e:
        state.add_log(f"⚠️ Impossible de charger {RULES_FILE}: {e}")


def save_custom_rules() -> None:
    """Sauvegarde uniquement les ajouts/modifications utilisateur."""
    try:
        categories_payload = {}
        for cat_name, cat_data in CATEGORIES.items():
            default_cat = DEFAULT_CATEGORIES.get(cat_name)
            if default_cat is None:
                categories_payload[cat_name] = {
                    "icon": cat_data.get("icon", "category"),
                    "color": cat_data.get("color", "cyan"),
                    "extensions": dict(cat_data.get("extensions", {})),
                }
                continue

            changed_ext = {}
            current_ext = cat_data.get("extensions", {})
            default_ext = default_cat.get("extensions", {})
            for ext, subcat in current_ext.items():
                if default_ext.get(ext) != subcat:
                    changed_ext[ext] = subcat
            if changed_ext:
                categories_payload[cat_name] = {
                    "icon": cat_data.get("icon", default_cat.get("icon", "category")),
                    "color": cat_data.get("color", default_cat.get("color", "cyan")),
                    "extensions": changed_ext,
                }

        project_payload = {}
        for ext, val in PROJECT_SIGNATURES.items():
            if DEFAULT_PROJECT_SIGNATURES.get(ext) != val:
                project_payload[ext] = [val[0], val[1]]

        payload = {
            "categories": categories_payload,
            "project_signatures": project_payload,
        }

        with open(RULES_FILE, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception as e:
        state.add_log(f"⚠️ Impossible d'écrire {RULES_FILE}: {e}")


def _resolve_logs_dir(destination: str, sources: list[str]) -> Optional[str]:
    target_root = destination if destination else ""
    if not target_root and sources:
        source_root = os.path.splitdrive(sources[0])[0] + os.sep
        target_root = os.path.join(source_root, "Organizator")
    if not target_root:
        return None
    return os.path.join(target_root, "Logs_ORGZ")


def export_report_json(
    report: dict,
    destination: str,
    mode: str,
    sources: list[str],
    detailed_context: Optional[dict] = None,
    full_logs: Optional[list[str]] = None,
) -> Optional[str]:
    """Exporte le rapport JSON complet dans Logs_ORGZ avec un nom horodaté."""
    try:
        target_dir = _resolve_logs_dir(destination, sources)
        if not target_dir:
            return None

        os.makedirs(target_dir, exist_ok=True)
        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = os.path.join(target_dir, f"rapport_organizator_{stamp}.json")

        payload = {
            "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
            "mode": mode,
            "destination": destination or "sur_place",
            "sources": sources,
            "report": report,
            "context": detailed_context or {},
            "logs": full_logs or [],
        }
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        return file_path
    except Exception as e:
        state.add_log(f"⚠️ Export JSON du rapport impossible: {e}")
        return None


def export_report_json_compact(report: dict, destination: str, mode: str, sources: list[str]) -> Optional[str]:
    """Exporte une version compacte du rapport JSON (résumé sans détails ligne à ligne)."""
    try:
        target_dir = _resolve_logs_dir(destination, sources)
        if not target_dir:
            return None

        os.makedirs(target_dir, exist_ok=True)
        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = os.path.join(target_dir, f"rapport_organizator_compact_{stamp}.json")

        organized = report.get("organized", [])
        compact = {
            "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
            "mode": mode,
            "destination": destination or "sur_place",
            "sources": sources,
            "summary": {
                "success": report.get("success", 0),
                "errors": report.get("errors", 0),
                "deleted": report.get("deleted", 0),
                "trashed": report.get("trashed", 0),
                "skipped": report.get("skipped", 0),
                "organized_count": len(organized),
                "deleted_files_count": len(report.get("deleted_files", [])),
                "trashed_files_count": len(report.get("trashed_files", [])),
                "deleted_empty_source_dirs_count": len(report.get("deleted_empty_source_dirs", [])),
            },
            "organized_preview": [
                {
                    "original": row.get("original", ""),
                    "destination": row.get("destination", ""),
                }
                for row in organized[:20]
            ],
            "deleted_empty_source_dirs": report.get("deleted_empty_source_dirs", []),
            "source_cleanup": report.get("source_cleanup", {}),
        }

        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(compact, f, ensure_ascii=False, indent=2)
        return file_path
    except Exception as e:
        state.add_log(f"⚠️ Export JSON compact impossible: {e}")
        return None


def export_report_txt_detailed(
    report: dict,
    destination: str,
    mode: str,
    sources: list[str],
    full_logs: list[str],
    detailed_context: Optional[dict] = None,
) -> Optional[str]:
    """Exporte un rapport TXT détaillé dans Logs_ORGZ."""
    try:
        target_dir = _resolve_logs_dir(destination, sources)
        if not target_dir:
            return None

        os.makedirs(target_dir, exist_ok=True)
        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = os.path.join(target_dir, f"rapport_organizator_detaille_{stamp}.txt")

        ctx = detailed_context or {}
        lines = [
            "ORGANIZATOR - RAPPORT DETAILLE",
            f"Genere le: {datetime.datetime.now().isoformat(timespec='seconds')}",
            f"Mode: {mode}",
            f"Destination: {destination or 'sur_place'}",
            f"Sources ({len(sources)}):",
        ]
        lines.extend([f"  - {src}" for src in sources])
        lines.append("")

        lines.append("RESUME")
        lines.append(f"  Success: {report.get('success', 0)}")
        lines.append(f"  Errors: {report.get('errors', 0)}")
        lines.append(f"  Deleted: {report.get('deleted', 0)}")
        lines.append(f"  Trashed: {report.get('trashed', 0)}")
        lines.append(f"  Skipped: {report.get('skipped', 0)}")
        lines.append("")

        lines.append("CONTEXTE")
        lines.append(f"  selected_categories_count: {ctx.get('selected_categories_count', 0)}")
        lines.append(f"  selected_files_count: {ctx.get('selected_files_count', 0)}")
        lines.append(f"  uncategorized_assign_count: {ctx.get('uncategorized_assign_count', 0)}")
        lines.append(f"  uncategorized_delete_count: {ctx.get('uncategorized_delete_count', 0)}")
        lines.append(f"  installer_assign_count: {ctx.get('installer_assign_count', 0)}")
        lines.append(f"  installer_delete_count: {ctx.get('installer_delete_count', 0)}")
        lines.append(f"  installer_trash_count: {ctx.get('installer_trash_count', 0)}")
        lines.append("")

        def append_section(title: str, values: list):
            lines.append(title)
            if not values:
                lines.append("  (aucun)")
                lines.append("")
                return
            for v in values:
                lines.append(f"  - {v}")
            lines.append("")

        append_section("FICHIERS ORGANISES", [
            f"{row.get('original', '')} -> {row.get('destination', '')} [{row.get('categorie', '')}/{row.get('sous_categorie', '')}]"
            for row in report.get("organized", [])
        ])
        append_section("FICHIERS SUPPRIMES", report.get("deleted_files", []))
        append_section("FICHIERS EN CORBEILLE", report.get("trashed_files", []))
        append_section("INSTALLERS ASSIGNES", [
            f"{row.get('original', '')} -> {row.get('destination', '')} [{row.get('categorie', '')}/{row.get('sous_categorie', '')}]"
            for row in report.get("installer_assigned", [])
        ])
        append_section("DOSSIERS SOURCE SUPPRIMES", report.get("deleted_empty_source_dirs", []))

        lines.append("JOURNAL COMPLET")
        if full_logs:
            lines.extend([f"  {log_line}" for log_line in full_logs])
        else:
            lines.append("  (vide)")
        lines.append("")

        with open(file_path, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))
        return file_path
    except Exception as e:
        state.add_log(f"⚠️ Export TXT détaillé impossible: {e}")
        return None


def cleanup_empty_source_dirs(sources: list[str]) -> dict:
    """Supprime les dossiers source devenus réellement vides après un move."""
    deleted = []
    non_empty = []
    failed = []

    for src in sources:
        try:
            root = os.path.abspath(src)
            if not os.path.isdir(root):
                continue

            drive, tail = os.path.splitdrive(root)
            is_drive_root = bool(drive) and tail in ("\\", "/", "")
            if is_drive_root:
                non_empty.append({"path": root, "reason": "drive_root_not_deleted"})
                continue

            # Nettoyage des sous-dossiers vides de bas en haut.
            for dirpath, dirnames, filenames in os.walk(root, topdown=False):
                if dirpath == root:
                    continue
                if not dirnames and not filenames:
                    try:
                        os.rmdir(dirpath)
                    except Exception:
                        pass

            try:
                if not os.listdir(root):
                    os.rmdir(root)
                    deleted.append(root)
                else:
                    non_empty.append({"path": root, "reason": "not_empty"})
            except Exception as e:
                failed.append({"path": root, "error": str(e)})
        except Exception as e:
            failed.append({"path": src, "error": str(e)})

    return {
        "deleted": deleted,
        "non_empty": non_empty,
        "failed": failed,
    }


load_custom_rules()


# ============================================================
# ENTRÉE DE SCAN (structure légère)
# ============================================================

class ScanEntry:
    __slots__ = ('path', 'name', 'ext', 'size', 'mtime', 'category', 'subcategory', 'nsfw_status', 'companions')

    def __init__(self, path: str, name: str, ext: str, size: int, mtime: float,
                 category: str, subcategory: str):
        self.path = path
        self.name = name
        self.ext = ext
        self.size = size
        self.mtime = mtime
        self.category = category
        self.subcategory = subcategory
        self.nsfw_status: str = ""          # "" | "SAIN" | "SENSUEL" | "EXPLICIT"
        self.companions: list = []          # [path_txt, path_validation_json, ...]


# ============================================================
# SCANNER DE DISQUES
# ============================================================

class DiskScanner:
    # Extensions d'images pouvant contenir des métadonnées IA
    AI_DETECTABLE_EXTS = frozenset({".png", ".jpg", ".jpeg", ".webp"})
    AI_TEXT_MARKERS = ("steps:", "negative prompt:", "model hash:", "cfg scale:", "parameters")
    AI_CHUNK_KEYS = (b"parameters", b"prompt", b"workflow", b"comment")

    @staticmethod
    def detect_ai_image_metadata(path: str) -> bool:
        """Détecte si une image contient des métadonnées de génération IA (SD, ComfyUI, Midjourney…)."""
        ext = os.path.splitext(path)[1].lower()
        try:
            if ext == ".png":
                import struct
                with open(path, "rb") as f:
                    if f.read(8) != b'\x89PNG\r\n\x1a\n':
                        return False
                    while True:
                        hdr = f.read(8)
                        if len(hdr) < 8:
                            break
                        length = struct.unpack(">I", hdr[:4])[0]
                        chunk_type = hdr[4:]
                        data = f.read(length)
                        f.read(4)  # CRC
                        if chunk_type in (b"tEXt", b"iTXt", b"zTXt"):
                            text = data.decode("latin-1", errors="replace").lower()
                            for marker in DiskScanner.AI_CHUNK_KEYS:
                                if marker in data[:128]:
                                    return True
                            for marker in DiskScanner.AI_TEXT_MARKERS:
                                if marker in text:
                                    return True
                        elif chunk_type == b"IEND":
                            break
            elif ext in (".jpg", ".jpeg", ".webp"):
                try:
                    import exifread
                    with open(path, "rb") as f:
                        tags = exifread.process_file(f, stop_tag="UserComment", details=False)
                    for tag_key in ("EXIF UserComment", "Image ImageDescription", "Image XPComment"):
                        val = str(tags.get(tag_key, "")).lower()
                        for marker in DiskScanner.AI_TEXT_MARKERS:
                            if marker in val:
                                return True
                except Exception:
                    pass
        except Exception:
            pass
        return False

    def __init__(self, log_callback, progress_callback):
        self.log = log_callback
        self.set_progress = progress_callback

    def scan(self, sources: list[str]) -> tuple[dict, list, list, list]:
        """
                Retourne:
                    results: {categorie: {sous_cat: [ScanEntry]}}
                    project_roots: [(root_path, sous_cat_projet)]
                    uncategorized: [ScanEntry]
                    installers: [ScanEntry]
        """
        results = defaultdict(lambda: defaultdict(list))
        project_roots = []
        project_detection_evidence = {}
        uncategorized = []
        installers = []
        project_root_lookup: dict[str, tuple[str, str]] = {}
        audio_project_roots = set()
        source_root_norms = {
            os.path.normcase(os.path.normpath(src))
            for src in sources
            if src
        }

        # Construire le lookup d'extensions à plat
        ext_lookup: dict[str, tuple[str, str]] = {}  # ext -> (categorie, sous_cat)
        for cat, info in CATEGORIES.items():
            if cat == "PROJETS":
                continue
            for ext, subcat in info["extensions"].items():
                ext_lookup[ext] = (cat, subcat)

        project_ext_lookup: dict[str, tuple[str, str]] = {}
        for ext, (subcat, _) in PROJECT_SIGNATURES.items():
            project_ext_lookup[ext] = ("PROJETS", subcat)

        total_files = 0
        processed = 0
        state.cancel_flag = False

        # Première passe : compter les fichiers et détecter les racines de projets
        self.log("Comptage des fichiers...")
        project_signature_priority = [
            ".als", ".alp", ".cpr", ".npr", ".psd", ".psb", ".aep", ".aet", ".uproject"
        ]

        for source in sources:
            for root, dirs, files in os.walk(source):
                dirs[:] = [d for d in dirs if d.lower() not in EXCLUDED_DIRS]
                norm_root = os.path.normcase(root)
                total_files += len(files)
                if norm_root in project_root_lookup:
                    continue

                root_extensions = {os.path.splitext(fname)[1].lower() for fname in files}
                marker_ext = None
                for ext in project_signature_priority:
                    if ext in root_extensions:
                        marker_ext = ext
                        break

                if marker_ext is None:
                    continue

                # Ne pas marquer la racine source entière comme projet:
                # sinon tous les fichiers normaux de ce dossier seraient ignorés.
                if norm_root in source_root_norms:
                    continue

                _, subcat = project_ext_lookup[marker_ext]
                project_root_lookup[norm_root] = (root, subcat)
                project_detection_evidence[root] = {
                    "subcategory": subcat,
                    "marker_ext": marker_ext,
                }
                if subcat in AUDIO_PROJECT_CATEGORIES:
                    audio_project_roots.add(norm_root)

        project_roots = list(project_root_lookup.values())

        self.log(f"Total fichiers à analyser : {total_files:,}")

        # Deuxième passe : classification
        for source in sources:
            self.log(f"Scan de : {source}")
            for root, dirs, files in os.walk(source):
                if state.cancel_flag:
                    self.log("⚠️ Scan annulé.")
                    break

                dirs[:] = [d for d in dirs if d.lower() not in EXCLUDED_DIRS]

                for fname in files:
                    if state.cancel_flag:
                        break

                    fpath = os.path.join(root, fname)
                    ext = os.path.splitext(fname)[1].lower()
                    norm_root = os.path.normcase(root)
                    in_project_root = any(
                        norm_root == project_root or norm_root.startswith(project_root + os.sep)
                        for project_root in project_root_lookup
                    )

                    try:
                        stat = os.stat(fpath)
                        fsize = stat.st_size
                        fmtime = stat.st_mtime
                    except OSError:
                        processed += 1
                        continue

                    # Vérifier si c'est un fichier projet
                    if ext in INSTALLER_EXTENSIONS:
                        if in_project_root:
                            processed += 1
                            continue
                        entry = ScanEntry(fpath, fname, ext, fsize, fmtime, "EXECUTABLES-ARCHIVES", ext.upper())
                        installers.append(entry)

                    elif ext in project_ext_lookup:
                        cat, subcat = project_ext_lookup[ext]
                        # Racine du projet = dossier parent du fichier signature
                        proj_root = root
                        # Ajouter aussi comme entrée pour les stats
                        entry = ScanEntry(fpath, fname, ext, fsize, fmtime, cat, subcat)
                        results[cat][subcat].append(entry)

                    elif ext in ext_lookup:
                        if in_project_root or (
                            ext in AUDIO_EXTENSIONS and any(
                                norm_root == project_root or norm_root.startswith(project_root + os.sep)
                                for project_root in audio_project_roots
                            )
                        ):
                            processed += 1
                            continue
                        cat, subcat = ext_lookup[ext]
                        # Détection automatique images IA
                        if cat == "PHOTO-VIDEO" and subcat == "Images" and ext in DiskScanner.AI_DETECTABLE_EXTS:
                            if DiskScanner.detect_ai_image_metadata(fpath):
                                subcat = "IA-Images"
                        entry = ScanEntry(fpath, fname, ext, fsize, fmtime, cat, subcat)
                        results[cat][subcat].append(entry)

                    else:
                        if in_project_root:
                            processed += 1
                            continue
                        entry = ScanEntry(fpath, fname, ext, fsize, fmtime, "?", ext or "(sans extension)")
                        uncategorized.append(entry)

                    processed += 1
                    if total_files > 0 and processed % 500 == 0:
                        prog = processed / total_files
                        self.set_progress(prog, f"Scan : {processed:,} / {total_files:,} fichiers")

        self.set_progress(1.0, "Scan terminé !")

        # Associer les fichiers compagnons (.txt tags, _validation.json) aux images
        DiskScanner._resolve_companions(results, uncategorized)

        return dict(results), project_roots, uncategorized, installers, project_detection_evidence

    @staticmethod
    def _resolve_companions(results: dict, uncategorized: list) -> None:
        """
        Détecte les fichiers compagnons (.txt tags, .json même stem, _validation.json NSFW) des images.
        - Attache les chemins des compagnons à entry.companions
        - Lit le tier NSFW depuis _validation.json → entry.nsfw_status
        - Retire les compagnons de toutes les catégories / non-catégorisés
        Les _validation.json sont TOUJOURS cachés du scan même sans image correspondante.
        """
        import json as _json

        # ── Étape 1 : collecter TOUS les _validation.json du scan ─────────────
        # Ils ne sont jamais des fichiers indépendants → on les retire inconditionnellement.
        validation_entries: list = []   # (entry, norm_path)

        def _is_validation_json(entry) -> bool:
            return (entry.ext.lower() == ".json" and
                    os.path.splitext(entry.name)[0].lower().endswith("_validation"))

        def _is_aesthetic_json(entry) -> bool:
            return (entry.ext.lower() == ".json" and
                    os.path.splitext(entry.name)[0].lower().endswith("_aesthetic"))

        for cat in list(results.keys()):
            for subcat in list(results[cat].keys()):
                kept, removed = [], []
                for e in results[cat][subcat]:
                    if _is_validation_json(e) or _is_aesthetic_json(e):
                        removed.append(e)
                    else:
                        kept.append(e)
                if removed:
                    results[cat][subcat] = kept
                    for e in removed:
                        validation_entries.append(e)
                if not results[cat][subcat]:
                    del results[cat][subcat]
            if not results[cat]:
                del results[cat]

        for i in range(len(uncategorized) - 1, -1, -1):
            if _is_validation_json(uncategorized[i]) or _is_aesthetic_json(uncategorized[i]):
                validation_entries.append(uncategorized[i])
                del uncategorized[i]

        # ── Étape 2 : indexer les images pour attacher les compagnons ──────────
        image_by_dir_stem: dict = {}
        for subcat in ("Images", "IA-Images"):
            for entry in results.get("PHOTO-VIDEO", {}).get(subcat, []):
                norm_dir = os.path.normcase(os.path.dirname(entry.path))
                stem = os.path.splitext(entry.name)[0].lower()
                image_by_dir_stem[(norm_dir, stem)] = entry

        # Toutes les clés sont stockées normalisées pour éviter tout écart de casse/slash
        companion_paths: set = set()   # os.path.normcase(path)

        def _check_and_attach(candidate_path: str):
            cname = os.path.basename(candidate_path)
            cstem, cext = os.path.splitext(cname)
            cext = cext.lower()
            norm_dir = os.path.normcase(os.path.dirname(candidate_path))
            norm_candidate = os.path.normcase(candidate_path)

            if cext == ".txt":
                # Essai 1 : stem exact (ex: image.png + image.txt)
                key = (norm_dir, cstem.lower())
                img_entry = image_by_dir_stem.get(key)
                # Essai 2 : suffixes connus de sidecars (_prompt, _tags, _caption, _meta)
                # ex: ComfyUI_00088_prompt.txt → image ComfyUI_00088.png
                # Gère aussi le double underscore (ComfyUI_00088__prompt.txt → stem ComfyUI_00088_)
                if img_entry is None:
                    cstem_low = cstem.lower()
                    for sidecar_suffix in ("_prompt", "_tags", "_caption", "_meta", "_desc", "_description"):
                        if cstem_low.endswith(sidecar_suffix):
                            base = cstem_low[: -len(sidecar_suffix)]
                            cand = image_by_dir_stem.get((norm_dir, base))
                            if cand is None and base.endswith("_"):
                                # double-underscore: ComfyUI_00088__prompt → ComfyUI_00088
                                cand = image_by_dir_stem.get((norm_dir, base.rstrip("_")))
                            if cand is not None:
                                img_entry = cand
                                break
                if img_entry:
                    img_entry.companions.append(candidate_path)
                    companion_paths.add(norm_candidate)

            elif cext == ".ia":
                # {stem}.ia — sidecar détection IA (JSON {is_ai, confidence, method, ...})
                key = (norm_dir, cstem.lower())
                img_entry = image_by_dir_stem.get(key)
                if img_entry:
                    img_entry.companions.append(candidate_path)
                    companion_paths.add(norm_candidate)

            elif cext == ".json":
                cstem_lower = cstem.lower()
                if cstem_lower.endswith("_validation"):
                    # {stem}_validation.json — déjà retiré du scan à l'étape 1,
                    # on l'attache à l'image si elle est dans le scan
                    base_stem = cstem_lower[:-len("_validation")]
                    key = (norm_dir, base_stem)
                    img_entry = image_by_dir_stem.get(key)
                    if img_entry:
                        img_entry.companions.append(candidate_path)
                        companion_paths.add(norm_candidate)
                        try:
                            with open(candidate_path, "r", encoding="utf-8") as f:
                                data = _json.load(f)
                            tier = data.get("result", {}).get("tier", "")
                            if tier:
                                img_entry.nsfw_status = tier.upper()
                        except Exception:
                            pass
                elif cstem_lower.endswith("_aesthetic"):
                    # {stem}_aesthetic.json — fichier compagnon de score esthétique
                    base_stem = cstem_lower[:-len("_aesthetic")]
                    key = (norm_dir, base_stem)
                    img_entry = image_by_dir_stem.get(key)
                    if img_entry:
                        img_entry.companions.append(candidate_path)
                        companion_paths.add(norm_candidate)
                else:
                    # {stem}.json  → métadonnées / tags companion
                    key = (norm_dir, cstem_lower)
                    img_entry = image_by_dir_stem.get(key)
                    if img_entry:
                        img_entry.companions.append(candidate_path)
                        companion_paths.add(norm_candidate)

        # Attacher les _validation.json collectés à l'étape 1 à leur image
        for val_entry in validation_entries:
            _check_and_attach(val_entry.path)

        if not image_by_dir_stem:
            return  # Pas d'images → .txt/.json réguliers non traités, OK

        # Scanner TOUTES les catégories pour .txt, .json et .ia réguliers
        for cat, subcats in results.items():
            for subcat, entries in subcats.items():
                for entry in list(entries):
                    if entry.ext.lower() in (".txt", ".json", ".ia"):
                        _check_and_attach(entry.path)

        for entry in list(uncategorized):
            _check_and_attach(entry.path)

        # Fallback disque : pour chaque image, vérifier si compagnons présents sur disque
        # mais non scannés (hors périmètre, règle manquante, etc.)
        for (norm_dir, stem), img_entry in image_by_dir_stem.items():
            real_dir = os.path.dirname(img_entry.path)
            img_stem = os.path.splitext(img_entry.name)[0]
            for suffix in (
                img_stem + ".txt",
                img_stem + "_prompt.txt",
                img_stem + "__prompt.txt",
                img_stem + "_tags.txt",
                img_stem + "_caption.txt",
                img_stem + ".json",
                img_stem + "_validation.json",
                img_stem + "_aesthetic.json",
                img_stem + ".ia",
            ):
                candidate = os.path.join(real_dir, suffix)
                if os.path.normcase(candidate) not in companion_paths and os.path.isfile(candidate):
                    _check_and_attach(candidate)

        if not companion_paths:
            return

        # Retirer les compagnons .txt / .json / .ia réguliers des catégories (comparaison normalisée)
        for cat in list(results.keys()):
            for subcat in list(results[cat].keys()):
                results[cat][subcat] = [
                    e for e in results[cat][subcat]
                    if os.path.normcase(e.path) not in companion_paths
                ]
                if not results[cat][subcat]:
                    del results[cat][subcat]
            if not results[cat]:
                del results[cat]
        for i in range(len(uncategorized) - 1, -1, -1):
            if os.path.normcase(uncategorized[i].path) in companion_paths:
                del uncategorized[i]


    @staticmethod
    def compute_stats(results: dict, project_roots: list, uncategorized: list, installers: list) -> dict:
        stats = {}
        for cat, subcats in results.items():
            total_count = 0
            total_size = 0
            sub_stats = {}
            for subcat, entries in subcats.items():
                c = len(entries)
                s = sum(e.size for e in entries)
                sub_stats[subcat] = {"count": c, "size": s}
                total_count += c
                total_size += s
            stats[cat] = {"count": total_count, "size": total_size, "subcats": sub_stats}

        # Stats projets (racines)
        if project_roots:
            proj_count = len(project_roots)
            proj_size = 0
            for root, _ in project_roots:
                for dirpath, _, files in os.walk(root):
                    for f in files:
                        try:
                            proj_size += os.path.getsize(os.path.join(dirpath, f))
                        except OSError:
                            pass
            stats.setdefault("PROJETS", {})["count"] = proj_count
            stats["PROJETS"]["size"] = proj_size

        if uncategorized:
            stats["?"] = {"count": len(uncategorized), "size": sum(e.size for e in uncategorized)}

        if installers:
            stats["EXECUTABLES-ARCHIVES"] = {
                "count": len(installers),
                "size": sum(e.size for e in installers),
            }

        return stats


# ============================================================
# EXTRACTEUR DE MÉTADONNÉES — PASSE 1 (fichier natif)
# ============================================================

class MetadataExtractor:

    @staticmethod
    def extract(entry: ScanEntry) -> dict:
        meta = {
            "nom": entry.name,
            "extension": entry.ext,
            "taille_octets": entry.size,
            "date_modification": datetime.datetime.fromtimestamp(entry.mtime).isoformat(),
        }

        ext = entry.ext.lower()

        # --- Audio ---
        if ext in (".mp3", ".flac", ".ogg", ".m4a", ".aac", ".wav", ".wma", ".aiff", ".aif"):
            try:
                import mutagen
                mf = mutagen.File(entry.path, easy=True)
                if mf:
                    for key in ("title", "artist", "album", "date", "tracknumber", "genre", "bpm"):
                        val = mf.get(key)
                        if val:
                            meta[key] = val[0] if isinstance(val, list) else str(val)
                    if hasattr(mf, "info"):
                        info = mf.info
                        if hasattr(info, "length"):
                            meta["duree_secondes"] = round(info.length, 2)
                        if hasattr(info, "bitrate"):
                            meta["bitrate"] = info.bitrate
                        if hasattr(info, "sample_rate"):
                            meta["sample_rate"] = info.sample_rate
            except Exception:
                pass

        # --- Images ---
        elif ext in (".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tiff", ".tif", ".heic", ".heif"):
            try:
                import exifread
                with open(entry.path, "rb") as f:
                    tags = exifread.process_file(f, stop_tag="UNDEF", details=False)

                def get_tag(name):
                    t = tags.get(name)
                    return str(t) if t else None

                for label, tag_name in [
                    ("date_prise", "EXIF DateTimeOriginal"),
                    ("date_numerisation", "EXIF DateTimeDigitized"),
                    ("fabricant", "Image Make"),
                    ("modele_appareil", "Image Model"),
                    ("exposition", "EXIF ExposureTime"),
                    ("ouverture", "EXIF FNumber"),
                    ("iso", "EXIF ISOSpeedRatings"),
                    ("focale", "EXIF FocalLength"),
                    ("flash", "EXIF Flash"),
                    ("orientation", "Image Orientation"),
                    ("largeur", "EXIF ExifImageWidth"),
                    ("hauteur", "EXIF ExifImageLength"),
                    ("gps_latitude", "GPS GPSLatitude"),
                    ("gps_longitude", "GPS GPSLongitude"),
                ]:
                    val = get_tag(tag_name)
                    if val:
                        meta[label] = val
            except Exception:
                pass

        # --- Vidéo ---
        elif ext in (".mp4", ".avi", ".mov", ".mkv", ".webm", ".mts", ".m2ts", ".wmv", ".flv"):
            try:
                import av
                with av.open(entry.path) as container:
                    meta["duree_secondes"] = round(float(container.duration) / 1_000_000, 2) if container.duration else None
                    if container.streams.video:
                        vs = container.streams.video[0]
                        meta["resolution"] = f"{vs.width}x{vs.height}" if vs.width else None
                        meta["codec_video"] = vs.codec_context.name if vs.codec_context else None
                        if vs.average_rate:
                            meta["fps"] = float(vs.average_rate)
                    if container.streams.audio:
                        as_ = container.streams.audio[0]
                        meta["codec_audio"] = as_.codec_context.name if as_.codec_context else None
                        meta["sample_rate"] = as_.codec_context.sample_rate if as_.codec_context else None
            except Exception:
                pass

        # --- PDF ---
        elif ext == ".pdf":
            try:
                import PyPDF2
                with open(entry.path, "rb") as f:
                    reader = PyPDF2.PdfReader(f)
                    meta["nb_pages"] = len(reader.pages)
                    info = reader.metadata
                    if info:
                        for key in ("/Title", "/Author", "/Subject", "/Creator", "/Producer", "/CreationDate"):
                            val = info.get(key)
                            if val:
                                meta[key.lstrip("/")] = str(val)
            except Exception:
                pass

        # --- Word DOCX ---
        elif ext == ".docx":
            try:
                from docx import Document
                doc = Document(entry.path)
                cp = doc.core_properties
                if cp.title:
                    meta["titre"] = cp.title
                if cp.author:
                    meta["auteur"] = cp.author
                if cp.created:
                    meta["date_creation"] = cp.created.isoformat()
                if cp.modified:
                    meta["date_modification_doc"] = cp.modified.isoformat()
                meta["nb_paragraphes"] = len(doc.paragraphs)
            except Exception:
                pass

        # --- Excel XLSX ---
        elif ext == ".xlsx":
            try:
                import openpyxl
                wb = openpyxl.load_workbook(entry.path, read_only=True, data_only=True)
                meta["nb_feuilles"] = len(wb.sheetnames)
                meta["feuilles"] = wb.sheetnames
                props = wb.properties
                if props.title:
                    meta["titre"] = props.title
                if props.creator:
                    meta["auteur"] = props.creator
                wb.close()
            except Exception:
                pass

        return meta

    @staticmethod
    def get_audio_tags(path: str) -> dict:
        """Retourne artist/album pour la nomenclature, avec fallbacks intelligents."""
        tags = {"artist": None, "album": None}
        
        # Tentative 1: mutagen
        try:
            import mutagen
            mf = mutagen.File(path, easy=True)
            if mf:
                artist = mf.get("artist", [None])[0]
                album = mf.get("album", [None])[0]
                if artist:
                    tags["artist"] = artist
                if album:
                    tags["album"] = album
                if artist or album:
                    return tags
        except Exception:
            pass
        
        # Tentative 2: inference de la hiérarchie de dossiers (ex: /path/Artiste/Album/file.aif)
        try:
            parent = os.path.dirname(path)  # dossier contenant le fichier
            parent_name = os.path.basename(parent)  # "Album" ou "1985-07-27"
            grandparent = os.path.dirname(parent)  # dossier parent du dossier parent
            grandparent_name = os.path.basename(grandparent)  # "Artiste" ou "Shibumi"
            
            # Stratégie: prendre 2 niveaux de dossiers si disponibles
            # /path/Shibumi/1985-07-27/file.aif → artist="Shibumi", album="1985-07-27"
            if grandparent_name and not grandparent_name.startswith("."):
                # Vérifier si pattern "Artiste - Album" dans un seul niveau
                if " - " in parent_name:
                    parts = parent_name.split(" - ", 1)
                    tags["artist"] = parts[0].strip()
                    if len(parts) > 1:
                        tags["album"] = parts[1].strip()
                else:
                    # Utiliser 2 niveaux: grandparent = artist, parent = album
                    tags["artist"] = grandparent_name.strip()
                    if parent_name and not parent_name.startswith("."):
                        tags["album"] = parent_name.strip()
            elif parent_name and not parent_name.startswith("."):
                # Fallback: un seul dossier disponible
                tags["artist"] = parent_name.strip()
        except Exception:
            pass
        
        return tags

    @staticmethod
    def get_image_date(path: str) -> Optional[datetime.datetime]:
        """Retourne la date EXIF d'une image, ou None."""
        try:
            import exifread
            with open(path, "rb") as f:
                tags = exifread.process_file(f, stop_tag="EXIF DateTimeOriginal", details=False)
            date_tag = tags.get("EXIF DateTimeOriginal")
            if date_tag:
                return datetime.datetime.strptime(str(date_tag), "%Y:%m:%d %H:%M:%S")
        except Exception:
            pass
        return None

    @staticmethod
    def get_video_date(path: str) -> Optional[datetime.datetime]:
        """Date vidéo par défaut: date de création fichier (fallback metadata, puis mtime)."""
        # Priorité 1 (Windows): date de création du fichier
        try:
            return datetime.datetime.fromtimestamp(os.path.getctime(path))
        except Exception:
            pass

        # Priorité 2: métadonnée interne vidéo (si disponible)
        try:
            import av
            with av.open(path) as container:
                for stream in container.streams.video:
                    creation = container.metadata.get("creation_time")
                    if creation:
                        return datetime.datetime.fromisoformat(creation.replace("Z", "+00:00")).replace(tzinfo=None)
        except Exception:
            pass

        # Priorité 3: date de modification
        try:
            return datetime.datetime.fromtimestamp(os.path.getmtime(path))
        except Exception:
            return None


# ============================================================
# ORGANISATEUR DE FICHIERS
# ============================================================

class FileOrganizer:
    def __init__(self, destination: str, mode: str, log_callback, progress_callback):
        self.destination = destination
        self.mode = mode  # "copy" ou "move"
        self.log = log_callback
        self.set_progress = progress_callback

    def _resolve_dest(
        self,
        entry: ScanEntry,
        metadata_overrides: dict = None,
        video_overrides: dict = None,
        video_skip_date_hierarchy: bool = False,
        video_no_date_folder_name: str = "DiversVideo",
    ) -> str:
        """Calcule le chemin de destination pour un fichier.
        
        Args:
            metadata_overrides: {path: {"artist": ..., "album": ...}} pour les fichiers musicauxédités
        """
        metadata_overrides = metadata_overrides or {}
        video_overrides = video_overrides or {}
        cat = entry.category
        subcat = entry.subcategory
        base_dest = self._base_destination(entry.path)

        if cat == "MUSIQUE-AUDIO":
            # Vérifier d'abord les override utilisateur
            if entry.path in metadata_overrides:
                tags = metadata_overrides[entry.path]
            else:
                tags = MetadataExtractor.get_audio_tags(entry.path)
            artist = self._sanitize(tags.get("artist") or "Artiste-Inconnu")
            album = self._sanitize(tags.get("album") or "Album-Inconnu")
            if subcat == "AUDIO-DIVERS":
                # Si pas de tags → dossier Divers
                has_tags = bool(tags.get("artist"))
                if has_tags:
                    return os.path.join(base_dest, "MUSIQUE-AUDIO", "Artistes", artist, album, entry.name)
                else:
                    return os.path.join(base_dest, "MUSIQUE-AUDIO", subcat, entry.name)
            return os.path.join(base_dest, "MUSIQUE-AUDIO", "Artistes", artist, album, entry.name)

        elif cat == "PHOTO-VIDEO":
            override = video_overrides.get(entry.path, {}) if subcat == "Video" else {}
            # Appliquer la reclassification manuelle IA ↔ Vraies Photos
            if subcat in ("Images", "IA-Images"):
                subcat = state.image_reclassify_overrides.get(entry.path, subcat)
            custom_folder = (override.get("folder") or "").strip()
            custom_name = (override.get("filename") or "").strip()

            final_name = entry.name
            if custom_name:
                custom_ext = os.path.splitext(custom_name)[1]
                if custom_ext:
                    final_name = self._sanitize(custom_name)
                else:
                    final_name = f"{self._sanitize(custom_name)}{entry.ext}"

            if subcat in ("Images", "IA-Images"):
                dt = MetadataExtractor.get_image_date(entry.path)
            else:
                dt = MetadataExtractor.get_video_date(entry.path)

            if dt is None:
                dt = datetime.datetime.fromtimestamp(entry.mtime)

            date_folder = f"{dt.year:04d}-{dt.month:02d}-{dt.day:02d}"

            if subcat == "Video":
                folder_parts = [
                    self._sanitize(part)
                    for part in custom_folder.replace("\\", "/").split("/")
                    if part.strip()
                ] if custom_folder else []

                if video_skip_date_hierarchy:
                    folder_name = self._sanitize(video_no_date_folder_name or "DiversVideo")
                    return os.path.join(base_dest, "PHOTO-VIDEO", subcat, *folder_parts, folder_name, final_name)

                return os.path.join(base_dest, "PHOTO-VIDEO", subcat, *folder_parts, date_folder, final_name)

            return os.path.join(base_dest, "PHOTO-VIDEO", subcat, date_folder, final_name)

        elif cat == "DOCUMENTS":
            return os.path.join(base_dest, "DOCUMENTS", subcat, entry.name)

        # Fallback
        return os.path.join(base_dest, cat, subcat, entry.name)

    def _resolve_project_dest(self, proj_root: str, subcat: str) -> str:
        """Dossier destination pour un projet (préserve la hiérarchie)."""
        base_dest = self._base_destination(proj_root)
        proj_name = os.path.basename(proj_root)
        return os.path.join(base_dest, "PROJETS", subcat, proj_name)

    def _base_destination(self, source_path: str) -> str:
        """Retourne la racine destination (ou racine 'sur place' si vide)."""
        if self.destination:
            return self.destination
        abs_src = os.path.abspath(source_path)
        drive, _ = os.path.splitdrive(abs_src)
        if drive:
            return os.path.join(drive + os.sep, "Organizator")
        return os.path.join(os.path.dirname(abs_src), "Organizator")

    @staticmethod
    def _sanitize(name: str) -> str:
        """Nettoie un nom pour l'utiliser dans un chemin."""
        for ch in r'<>:"/\|?*':
            name = name.replace(ch, "_")
        return name.strip(". ")[:80] or "Inconnu"

    @staticmethod
    def _resolve_conflict(dest_path: str) -> str:
        """Si la destination existe déjà, ajoute un suffixe numérique."""
        if not os.path.exists(dest_path):
            return dest_path
        base, ext = os.path.splitext(dest_path)
        counter = 1
        while True:
            new_path = f"{base}_{counter}{ext}"
            if not os.path.exists(new_path):
                return new_path
            counter += 1

    def _do_transfer(self, src: str, dest: str) -> tuple[bool, str]:
        """Copie ou déplace src vers dest. Retourne (succès, destination réelle)."""
        # Même fichier source/destination → rien à faire (évite le renommage parasite)
        if os.path.normcase(os.path.normpath(src)) == os.path.normcase(os.path.normpath(dest)):
            return True, dest
        os.makedirs(os.path.dirname(dest), exist_ok=True)
        dest = self._resolve_conflict(dest)
        # Re-vérifier après résolution de conflit (cas extrême)
        if os.path.normcase(os.path.normpath(src)) == os.path.normcase(os.path.normpath(dest)):
            return True, dest
        try:
            if self.mode == "copy":
                shutil.copy2(src, dest)
            else:
                shutil.move(src, dest)
            return True, dest
        except Exception as e:
            self.log(f"❌ Erreur transfert {os.path.basename(src)} : {e}")
            return False, dest

    def organize(
        self,
        selected: dict,            # {(cat, subcat): bool}
        selected_files: dict,      # {path: bool}
        scan_results: dict,        # {cat: {subcat: [ScanEntry]}}
        project_roots: list,       # [(root_path, subcat)]
        uncategorized_decisions: dict,  # {path: "ignore" | "delete" | "assign:CAT"}
        uncategorized: list,       # [ScanEntry]
        installer_decisions: dict,  # {path: "keep" | "trash" | "delete" | "assign:CAT"}
        installer_files: list,      # [ScanEntry]
        db_archive=None,           # PostgreSQLArchive optionnel
        metadata_overrides: dict = None,  # {path: {"artist": ..., "album": ...}} pour MUSIQUE-AUDIOédité
        video_overrides: dict = None,
        video_skip_date_hierarchy: bool = False,
        video_no_date_folder_name: str = "DiversVideo",
    ) -> dict:
        """
        Exécute l'organisation. Retourne un rapport.
        """
        metadata_overrides = metadata_overrides or {}
        video_overrides = video_overrides or {}
        report = {
            "success": 0,
            "skipped": 0,
            "errors": 0,
            "deleted": 0,
            "trashed": 0,
            "organized": [],
            "deleted_files": [],
            "trashed_files": [],
            "installer_assigned": [],
        }
        state.cancel_flag = False

        # Collecter tous les fichiers à traiter
        tasks = []  # (entry, dest_path)

        for cat, subcats in scan_results.items():
            if cat == "PROJETS":
                continue  # Traitement séparé
            for subcat, entries in subcats.items():
                for entry in entries:
                    if not is_scan_entry_selected(entry, selected, selected_files):
                        report["skipped"] += 1
                        continue
                    dest = self._resolve_dest(
                        entry,
                        metadata_overrides,
                        video_overrides,
                        video_skip_date_hierarchy,
                        video_no_date_folder_name,
                    )
                    # Routage NSFW vers destination alternative
                    is_nsfw = entry.nsfw_status in ("SENSUEL", "EXPLICIT")
                    if is_nsfw:
                        if state.nsfw_destination:
                            # Remplacer la base de destination par nsfw_destination
                            # en préservant toute la structure (catégorie/sous-cat/date/fichier)
                            base_dest = self._base_destination(entry.path)
                            try:
                                rel = os.path.relpath(dest, base_dest)
                            except ValueError:
                                rel = os.path.join(entry.category, entry.subcategory, os.path.basename(dest))
                            dest = os.path.join(state.nsfw_destination, rel)
                            self.log(f"🔞 NSFW ({entry.nsfw_status}) → {state.nsfw_destination} : {entry.name}")
                        elif state.nsfw_exclude:
                            # Exclure complètement (pas de destination alternative)
                            report["skipped"] += 1
                            self.log(f"⏭️ NSFW exclu (pas de dossier dest.) : {entry.name}")
                            continue
                    tasks.append((entry, dest))

        # Projets sélectionnés
        selected_proj_subcats = {
            subcat for (cat, subcat), checked in selected.items()
            if cat == "PROJETS" and checked
        }
        project_tasks = []
        for proj_root, subcat in project_roots:
            if subcat in selected_proj_subcats:
                dest_root = self._resolve_project_dest(proj_root, subcat)
                project_tasks.append((proj_root, dest_root))
            else:
                # Compter les fichiers skippés
                for _, _, files in os.walk(proj_root):
                    report["skipped"] += len(files)

        # Fichiers non catégorisés
        uncategorized_delete_ops = []
        for entry in uncategorized:
            decision = uncategorized_decisions.get(entry.path, "ignore")
            if decision == "ignore":
                report["skipped"] += 1
                continue
            if decision == "delete":
                uncategorized_delete_ops.append(entry)
                continue
            if decision.startswith("assign:"):
                assigned_cat = decision[7:]
                base_dest = self._base_destination(entry.path)
                dest = os.path.join(base_dest, assigned_cat,
                                    entry.ext.lstrip(".").upper() or "DIVERS",
                                    entry.name)
                tasks.append((entry, dest))

        # Fichiers exécutables et archives (.msi/.exe/.zip/.rar)
        installer_delete_ops = []
        installer_trash_ops = []
        for entry in installer_files:
            decision = installer_decisions.get(entry.path, "keep")
            if decision == "keep":
                # Les archives conservées (.zip/.rar) sont organisées automatiquement vers DOCUMENTS.
                if entry.ext.lower() in {".zip", ".rar"}:
                    assigned_cat = "DOCUMENTS"
                    subcat = entry.ext.lstrip(".").upper() or "DIVERS"
                    assigned_entry = ScanEntry(
                        path=entry.path,
                        name=entry.name,
                        ext=entry.ext,
                        size=entry.size,
                        mtime=entry.mtime,
                        category=assigned_cat,
                        subcategory=subcat,
                    )
                    base_dest = self._base_destination(entry.path)
                    dest = os.path.join(base_dest, assigned_cat, subcat, entry.name)
                    tasks.append((assigned_entry, dest))
                    report["installer_assigned"].append({
                        "original": entry.path,
                        "destination": dest,
                        "categorie": assigned_cat,
                        "sous_categorie": subcat,
                        "decision": "auto_keep_archive_to_documents",
                    })
                    continue
                report["skipped"] += 1
                continue
            if decision == "trash":
                installer_trash_ops.append(entry)
                continue
            if decision == "delete":
                installer_delete_ops.append(entry)
                continue
            if decision.startswith("assign:"):
                assigned_cat = decision[7:]
                subcat = entry.ext.lstrip(".").upper() or "DIVERS"
                assigned_entry = ScanEntry(
                    path=entry.path,
                    name=entry.name,
                    ext=entry.ext,
                    size=entry.size,
                    mtime=entry.mtime,
                    category=assigned_cat,
                    subcategory=subcat,
                )
                base_dest = self._base_destination(entry.path)
                dest = os.path.join(base_dest, assigned_cat, subcat, entry.name)
                tasks.append((assigned_entry, dest))
                report["installer_assigned"].append({
                    "original": entry.path,
                    "destination": dest,
                    "categorie": assigned_cat,
                    "sous_categorie": subcat,
                })

        total = len(tasks) + len(project_tasks) + len(installer_trash_ops) + len(installer_delete_ops) + len(uncategorized_delete_ops)
        if total == 0:
            self.log("ℹ️ Aucun fichier à organiser.")
            return report

        self.log(f"Organisation de {total} éléments en mode '{self.mode}'...")
        processed = 0

        # Envoi à la corbeille des exécutables/archives marqués
        for entry in installer_trash_ops:
            if state.cancel_flag:
                self.log("⚠️ Organisation annulée.")
                break
            try:
                if not HAS_SEND2TRASH:
                    raise RuntimeError("Package send2trash non installé")
                send2trash(entry.path)
                report["trashed"] += 1
                report["trashed_files"].append(entry.path)
                self.log(f"🧺 Corbeille : {entry.path}")
            except Exception as e:
                report["errors"] += 1
                self.log(f"❌ Erreur corbeille {entry.path} : {e}")

            processed += 1
            if processed % 50 == 0 or processed == total:
                self.set_progress(processed / total, f"Organisation : {processed:,} / {total:,}")

        # Suppression des exécutables/archives marqués
        for entry in installer_delete_ops:
            if state.cancel_flag:
                self.log("⚠️ Organisation annulée.")
                break
            try:
                os.remove(entry.path)
                report["deleted"] += 1
                report["deleted_files"].append(entry.path)
                self.log(f"🗑️ Supprimé : {entry.path}")
            except Exception as e:
                report["errors"] += 1
                self.log(f"❌ Erreur suppression {entry.path} : {e}")

            processed += 1
            if processed % 50 == 0 or processed == total:
                self.set_progress(processed / total, f"Organisation : {processed:,} / {total:,}")

        # Suppression des non catégorisés marqués
        for entry in uncategorized_delete_ops:
            if state.cancel_flag:
                self.log("⚠️ Organisation annulée.")
                break
            try:
                os.remove(entry.path)
                report["deleted"] += 1
                report["deleted_files"].append(entry.path)
                self.log(f"🗑️ Supprimé (non catégorisé) : {entry.path}")
            except Exception as e:
                report["errors"] += 1
                self.log(f"❌ Erreur suppression {entry.path} : {e}")

            processed += 1
            if processed % 50 == 0 or processed == total:
                self.set_progress(processed / total, f"Organisation : {processed:,} / {total:,}")

        # Transfert des fichiers normaux
        for entry, dest_path in tasks:
            if state.cancel_flag:
                self.log("⚠️ Organisation annulée.")
                break

            ok, actual_dest = self._do_transfer(entry.path, dest_path)
            if ok:
                report["success"] += 1
                report["organized"].append({
                    "original": entry.path,
                    "destination": actual_dest,
                    "categorie": entry.category,
                    "sous_categorie": entry.subcategory,
                })
                # Copier/déplacer les fichiers compagnons (.txt tags, .json, _validation.json)
                if entry.companions:
                    dest_dir = os.path.dirname(actual_dest)
                    # Si l'image a été renommée (conflit), adapter le stem des compagnons
                    orig_stem = os.path.splitext(entry.name)[0]
                    actual_stem = os.path.splitext(os.path.basename(actual_dest))[0]
                    stem_changed = (orig_stem != actual_stem)
                    for companion_path in entry.companions:
                        try:
                            cname = os.path.basename(companion_path)
                            if stem_changed:
                                # Renommer le compagnon pour rester synchrone avec la nouvelle image
                                c_stem, c_ext = os.path.splitext(cname)
                                # Remplacer le préfixe original par le nouveau stem
                                if c_stem.lower().startswith(orig_stem.lower()):
                                    suffix = c_stem[len(orig_stem):]  # ex: "_validation"
                                    cname = actual_stem + suffix + c_ext
                            companion_dest = os.path.join(dest_dir, cname)
                            self._do_transfer(companion_path, companion_dest)
                        except Exception as e:
                            self.log(f"⚠️ Compagnon non transféré {os.path.basename(companion_path)} : {e}")
                # Archivage PostgreSQL si disponible
                if db_archive:
                    checksum = self._md5(entry.path) if self.mode == "copy" else ""
                    db_archive.register_file_sync(entry, actual_dest, checksum)
            else:
                report["errors"] += 1

            processed += 1
            if processed % 50 == 0 or processed == len(tasks):
                self.set_progress(processed / total, f"Organisation : {processed:,} / {total:,}")

        # Transfert des projets (copie récursive)
        for src_root, dest_root in project_tasks:
            if state.cancel_flag:
                break
            try:
                if self.mode == "copy":
                    if os.path.exists(dest_root):
                        dest_root = self._resolve_conflict(dest_root)
                    shutil.copytree(src_root, dest_root)
                else:
                    if os.path.exists(dest_root):
                        dest_root = self._resolve_conflict(dest_root)
                    shutil.move(src_root, dest_root)
                report["success"] += 1
                self.log(f"✅ Projet organisé : {os.path.basename(src_root)} → {dest_root}")
            except Exception as e:
                report["errors"] += 1
                self.log(f"❌ Erreur projet {os.path.basename(src_root)} : {e}")

            processed += 1
            self.set_progress(processed / total, f"Organisation : {processed:,} / {total:,}")

        self.log(
            f"🎉 Organisation terminée ! {report['success']} succès, "
            f"{report['trashed']} en corbeille, {report['deleted']} supprimés, "
            f"{report['errors']} erreurs, {report['skipped']} ignorés."
        )
        self.set_progress(1.0, "Organisation terminée !")
        return report

    @staticmethod
    def _md5(path: str) -> str:
        try:
            h = hashlib.md5()
            with open(path, "rb") as f:
                for chunk in iter(lambda: f.read(65536), b""):
                    h.update(chunk)
            return h.hexdigest()
        except Exception:
            return ""


# ============================================================
# ARCHIVE POSTGRESQL
# ============================================================

class PostgreSQLArchive:
    """Gestion de la base de données PostgreSQL."""

    DB_URL = "postgresql://organizator:organizator2026@localhost:5432/organizator"

    CREATE_TABLE_SQL = """
    CREATE TABLE IF NOT EXISTS fichiers (
        id                BIGSERIAL PRIMARY KEY,
        chemin_original   TEXT NOT NULL,
        chemin_destination TEXT,
        nom_fichier       TEXT NOT NULL,
        extension         TEXT,
        categorie         TEXT,
        sous_categorie    TEXT,
        taille_octets     BIGINT,
        date_modification TIMESTAMPTZ,
        checksum_md5      TEXT,
        metadata_json     JSONB DEFAULT '{}',
        ai_metadata_json  JSONB DEFAULT '{}',
        ai_traite_le      TIMESTAMPTZ,
        statut            TEXT DEFAULT 'organisé',
        cree_le           TIMESTAMPTZ DEFAULT NOW()
    );
    CREATE INDEX IF NOT EXISTS idx_fichiers_categorie ON fichiers(categorie);
    CREATE INDEX IF NOT EXISTS idx_fichiers_extension ON fichiers(extension);
    CREATE INDEX IF NOT EXISTS idx_fichiers_statut    ON fichiers(statut);
    CREATE INDEX IF NOT EXISTS idx_fichiers_chemin    ON fichiers(chemin_original);
    """

    def __init__(self):
        self._pool = None
        self._available = False
        self._main_loop = None

    def set_main_loop(self, loop):
        """Capture la boucle asyncio principale (UI) pour les callbacks thread-safe."""
        self._main_loop = loop

    async def connect(self) -> bool:
        try:
            import asyncpg
            self._pool = await asyncpg.create_pool(self.DB_URL, min_size=2, max_size=10)
            async with self._pool.acquire() as conn:
                await conn.execute(self.CREATE_TABLE_SQL)
            self._available = True
            return True
        except Exception as e:
            state.add_log(f"⚠️ PostgreSQL indisponible : {e}")
            self._available = False
            return False

    async def close(self):
        if self._pool:
            await self._pool.close()

    def register_file_sync(self, entry: ScanEntry, dest_path: str, checksum: str):
        """Enregistre un fichier de manière synchrone (wrapper pour usage dans thread)."""
        if not self._available or not self._pool:
            return

        loop = self._main_loop
        if loop is None or loop.is_closed():
            state.add_log("⚠️ Boucle asyncio principale indisponible pour l'archivage BD.")
            return

        future = asyncio.run_coroutine_threadsafe(
            self._register(entry, dest_path, checksum),
            loop,
        )

        # Lire le résultat pour éviter les exceptions silencieuses en arrière-plan.
        def _on_done(fut):
            try:
                fut.result()
            except Exception as e:
                state.add_log(f"⚠️ BD insert async erreur : {e}")

        future.add_done_callback(_on_done)

    async def _register(self, entry: ScanEntry, dest_path: str, checksum: str):
        if not self._available or not self._pool:
            return
        try:
            async with self._pool.acquire() as conn:
                await conn.execute("""
                    INSERT INTO fichiers
                        (chemin_original, chemin_destination, nom_fichier, extension,
                         categorie, sous_categorie, taille_octets, date_modification,
                         checksum_md5, statut)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10)
                    ON CONFLICT DO NOTHING
                """,
                    entry.path, dest_path, entry.name, entry.ext,
                    entry.category, entry.subcategory, entry.size,
                    datetime.datetime.fromtimestamp(entry.mtime),
                    checksum, "en_attente_validation"
                )
        except Exception as e:
            state.add_log(f"⚠️ BD insert erreur : {e}")

    async def register_batch(self, entries: list[dict]):
        """Enregistre un lot d'entrées après organisation."""
        if not self._available or not self._pool:
            return
        try:
            async with self._pool.acquire() as conn:
                await conn.executemany("""
                    INSERT INTO fichiers
                        (chemin_original, chemin_destination, nom_fichier, extension,
                         categorie, sous_categorie, taille_octets, checksum_md5, statut)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)
                    ON CONFLICT DO NOTHING
                """, [
                    (e["chemin_original"], e["chemin_destination"], e["nom_fichier"],
                     e["extension"], e["categorie"], e["sous_categorie"],
                     e["taille_octets"], e.get("checksum_md5", ""), e.get("statut", "en_attente_validation"))
                    for e in entries
                ])
        except Exception as e:
            state.add_log(f"⚠️ BD batch insert erreur : {e}")

    async def update_metadata(
        self,
        path_lookup: str,
        metadata: dict,
        is_ai: bool = False,
        set_archived: bool = False,
    ):
        if not self._available or not self._pool:
            return
        try:
            import json as _json
            col = "ai_metadata_json" if is_ai else "metadata_json"
            extra_parts = []
            if is_ai:
                extra_parts.append("ai_traite_le = NOW()")
                if set_archived:
                    extra_parts.append("statut = 'archivé'")
            extra = (", " + ", ".join(extra_parts)) if extra_parts else ""
            async with self._pool.acquire() as conn:
                await conn.execute(
                    f"UPDATE fichiers SET {col} = $1{extra} WHERE chemin_original = $2 OR chemin_destination = $2",
                    _json.dumps(metadata, ensure_ascii=False), path_lookup
                )
        except Exception as e:
            state.add_log(f"⚠️ BD update metadata erreur : {e}")

    async def get_stats(self) -> dict:
        if not self._available or not self._pool:
            return {}
        try:
            async with self._pool.acquire() as conn:
                total = await conn.fetchval("SELECT COUNT(*) FROM fichiers")
                by_cat = await conn.fetch(
                    "SELECT categorie, COUNT(*), SUM(taille_octets) FROM fichiers GROUP BY categorie"
                )
                pending_ai = await conn.fetchval(
                    "SELECT COUNT(*) FROM fichiers WHERE ai_traite_le IS NULL"
                )
                return {
                    "total": total,
                    "par_categorie": {r["categorie"]: {"count": r["count"], "taille": r["sum"]} for r in by_cat},
                    "en_attente_ia": pending_ai,
                }
        except Exception:
            return {}

    async def get_pending_ai_files(self, limit: int = 500) -> list[dict]:
        """Retourne les fichiers dont l'enrichissement IA n'est pas encore traité."""
        if not self._available or not self._pool:
            return []
        try:
            async with self._pool.acquire() as conn:
                rows = await conn.fetch(
                    """
                    SELECT
                        chemin_original,
                        chemin_destination,
                        nom_fichier,
                        categorie,
                        sous_categorie,
                        extension,
                        taille_octets,
                        cree_le
                    FROM fichiers
                    WHERE ai_traite_le IS NULL
                    ORDER BY cree_le DESC NULLS LAST
                    LIMIT $1
                    """,
                    max(1, int(limit)),
                )
                return [dict(r) for r in rows]
        except Exception as e:
            state.add_log(f"⚠️ BD lecture attente IA impossible: {e}")
            return []

    async def migrate_existing_records_to_validation(self) -> int:
        """Marque les anciens enregistrements comme en_attente_validation si pas encore traités IA."""
        if not self._available or not self._pool:
            return 0
        try:
            async with self._pool.acquire() as conn:
                result = await conn.execute(
                    """
                    UPDATE fichiers
                    SET statut = 'en_attente_validation'
                    WHERE (statut = 'organisé' OR statut IS NULL OR statut = '')
                      AND ai_traite_le IS NULL
                    """
                )
            # asyncpg renvoie "UPDATE <n>"
            count = int(str(result).split()[-1]) if result else 0
            if count > 0:
                state.add_log(f"🛠️ Migration statuts BD: {count} enregistrements -> en_attente_validation")
            return count
        except Exception as e:
            state.add_log(f"⚠️ Migration statuts BD impossible: {e}")
            return 0


# ============================================================
# FILE DE MÉTADONNÉES LLM (PASSE 2) — BACKGROUND
# ============================================================

class LLMMetadataQueue:
    """
    File d'attente pour l'enrichissement LLM (Passe 2).
    Traitement en arrière-plan, non bloquant.
    Réutilise Qwen3-VL de MediaMind AI si disponible.
    """

    def __init__(self, db_archive: PostgreSQLArchive):
        self.db = db_archive
        self.queue: asyncio.Queue = asyncio.Queue()
        self._running = False
        self._task = None
        self.required_signals = {"aesthetic", "nsfw", "tags"}
        self.archive_roots: list[str] = []

    def configure_validation(self, validate_aesthetic: bool, validate_nsfw: bool, validate_tags: bool, archive_roots: list[str]):
        required = set()
        if validate_aesthetic:
            required.add("aesthetic")
        if validate_nsfw:
            required.add("nsfw")
        if validate_tags:
            required.add("tags")
        self.required_signals = required
        self.archive_roots = list(archive_roots or [])

    def enqueue(self, paths: list[str]):
        jobs = []
        for p in paths:
            jobs.append({
                "analyze_path": p,
                "original_path": p,
                "destination_path": p,
            })
        self.enqueue_jobs(jobs)

    def enqueue_jobs(self, jobs: list[dict]):
        count = 0
        for job in jobs:
            if not isinstance(job, dict):
                continue
            analyze = str(job.get("analyze_path") or "").strip()
            original = str(job.get("original_path") or analyze).strip()
            destination = str(job.get("destination_path") or analyze).strip()
            if not analyze:
                continue
            self.queue.put_nowait({
                "analyze_path": analyze,
                "original_path": original,
                "destination_path": destination,
            })
            count += 1
        state.add_log(f"🤖 {count} fichiers en file pour enrichissement LLM.")

    def pending_count(self) -> int:
        return self.queue.qsize()

    async def start(self):
        if self._running:
            return
        self._running = True
        self._task = asyncio.create_task(self._worker())

    async def stop(self):
        self._running = False
        if self._task:
            self._task.cancel()

    async def _worker(self):
        _POLL_DELAY_S = 30
        _SERVER_WAIT_S = 60

        pending: list[dict] = []
        submitted: set[str] = set()

        while self._running:
            new_jobs: list[dict] = []
            try:
                while True:
                    job = self.queue.get_nowait()
                    analyze = str((job or {}).get("analyze_path", ""))
                    if is_image_path(analyze):
                        pending.append(job)
                        new_jobs.append(job)
                    else:
                        display = os.path.basename(str((job or {}).get("destination_path", analyze)))
                        state.add_log(f"⏸️ Validation IA différée (non-image): {display}")
                    self.queue.task_done()
            except asyncio.QueueEmpty:
                pass

            if not pending:
                try:
                    job = await asyncio.wait_for(self.queue.get(), timeout=1.0)
                    analyze = str((job or {}).get("analyze_path", ""))
                    if is_image_path(analyze):
                        pending.append(job)
                        new_jobs.append(job)
                    else:
                        display = os.path.basename(str((job or {}).get("destination_path", analyze)))
                        state.add_log(f"⏸️ Validation IA différée (non-image): {display}")
                    self.queue.task_done()
                except asyncio.TimeoutError:
                    pass
                continue

            new_paths = [
                str(j.get("analyze_path", ""))
                for j in new_jobs
                if str(j.get("analyze_path", "")) and str(j.get("analyze_path", "")) not in submitted
            ]
            if new_paths:
                ok = await asyncio.get_event_loop().run_in_executor(
                    None, self._submit_for_analysis, new_paths
                )
                if ok:
                    submitted.update(new_paths)
                    state.add_log(
                        f"🔍 {len(new_paths)} fichier(s) soumis à MediaMind AI pour indexation."
                    )
                else:
                    state.add_log(
                        f"⚠️ MediaMind AI non joignable — {len(pending)} fichier(s) en attente."
                        f" Démarrez MediaMind AI pour continuer."
                    )

            still_pending: list[dict] = []
            validated = 0
            not_indexed_count = 0
            server_down = False

            for job in pending:
                analyze_path = str((job or {}).get("analyze_path", ""))
                original_path = str((job or {}).get("original_path", analyze_path))
                destination_path = str((job or {}).get("destination_path", analyze_path))
                display_name = os.path.basename(destination_path or analyze_path or "?")

                try:
                    meta = await asyncio.get_event_loop().run_in_executor(
                        None, self._process_file, analyze_path
                    )

                    if meta is None:
                        still_pending.append(job)
                        server_down = True
                        continue

                    if meta.get("_not_indexed"):
                        poll_count = int(job.get("_poll_count", 0)) + 1
                        job = dict(job)
                        job["_poll_count"] = poll_count
                        still_pending.append(job)
                        not_indexed_count += 1
                        if poll_count >= 3:
                            # Retirer du submitted pour permettre la resoumission
                            submitted.discard(analyze_path)
                            state.add_log(
                                f"⚠️ {display_name} non indexé après {poll_count} tentatives — resoumission forcée"
                                f" | chemin: {analyze_path}"
                            )
                        else:
                            state.add_log(f"⌛ Métadonnées IA pas encore prêtes : {display_name}")
                        continue

                    present = {k for k in ("aesthetic", "nsfw", "tags", "prompt") if k in meta}
                    effective_present = set(present)
                    if "prompt" in effective_present:
                        effective_present.add("tags")
                    missing = sorted(list(self.required_signals - effective_present))
                    if missing:
                        still_pending.append(job)
                        not_indexed_count += 1
                        state.add_log(
                            f"⌛ Données IA partielles pour {display_name} — présents: {sorted(list(present))}, manquants: {missing}"
                        )
                        continue

                    if self.archive_roots and not is_under_archive_roots(destination_path, self.archive_roots):
                        state.add_log(
                            f"⏳ Fichier non archivé final ({display_name}) — hors racines archive"
                        )
                        continue

                    meta["validation_required"] = sorted(list(self.required_signals))
                    meta["validation_present"] = sorted(list(present))
                    meta["validation_effective_present"] = sorted(list(effective_present))
                    meta["archive_destination"] = destination_path
                    meta["archive_roots"] = self.archive_roots

                    self._write_sidecars(destination_path, meta)
                    await self.db.update_metadata(original_path, meta, is_ai=True, set_archived=True)
                    submitted.discard(analyze_path)
                    validated += 1
                    state.add_log(f"✅ Validation IA finalisée : {display_name}")

                except Exception as e:
                    state.add_log(f"⚠️ LLM erreur sur {display_name} : {e}")

            pending = still_pending

            if validated > 0:
                state.add_log(
                    f"🤖 {validated} fichier(s) validé(s)."
                    + (f" {len(pending)} encore en attente d'indexation." if pending else " Tous validés ✓")
                )

            if pending:
                wait = _SERVER_WAIT_S if server_down else _POLL_DELAY_S
                state.add_log(
                    f"⏱️ Vérification dans {wait}s ({len(pending)} fichier(s) en attente)."
                )
                await asyncio.sleep(wait)

    @staticmethod
    def _write_sidecars(destination_path: str, payload: dict):
        if not destination_path:
            return
        try:
            target = Path(destination_path)
            if not target.exists() or not target.is_file():
                return

            json_sidecar = target.with_name(f"{target.stem}_archive_validation.json")
            txt_sidecar = target.with_name(f"{target.stem}_archive_validation.txt")

            with open(json_sidecar, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)

            lines = [
                f"file: {target.name}",
                f"path: {destination_path}",
                f"validated_at: {datetime.datetime.now().isoformat(timespec='seconds')}",
            ]
            if "aesthetic" in payload:
                aes = payload.get("aesthetic") or {}
                lines.append(f"aesthetic_avg: {aes.get('avg_score', 'n/a')}")
                lines.append(f"aesthetic_max: {aes.get('max_score', 'n/a')}")
            if "nsfw" in payload:
                ns = payload.get("nsfw") or {}
                lines.append(f"nsfw_label: {ns.get('top_label', 'n/a')}")
                lines.append(f"nsfw_danger: {ns.get('danger_score', 'n/a')}")
            if "tags" in payload:
                tag_values = ((payload.get("tags") or {}).get("values") or {})
                lines.append(f"tags_count: {len(tag_values)}")
            if "prompt" in payload:
                prompt_text = str(((payload.get("prompt") or {}).get("text") or "")).strip()
                lines.append(f"prompt_present: {'yes' if prompt_text else 'no'}")
                if prompt_text:
                    lines.append(f"prompt_preview: {prompt_text[:240].replace(chr(10), ' ')}")
            with open(txt_sidecar, "w", encoding="utf-8") as f:
                f.write("\n".join(lines))
        except Exception as e:
            state.add_log(f"⚠️ Sidecar IA impossible ({os.path.basename(destination_path)}): {e}")

    @staticmethod
    def _server_base_url() -> str:
        api_base = os.getenv("MEDIAMIND_API_URL", "http://127.0.0.1:8190/api/llm/enrich").strip()
        from urllib.parse import urlparse, urlunparse
        parsed = urlparse(api_base)
        return urlunparse((parsed.scheme, parsed.netloc, "", "", "", ""))

    @staticmethod
    def _submit_for_analysis(paths: list[str]) -> bool:
        """Soumet une liste de chemins à MediaMind AI pour analyse."""
        try:
            if not paths:
                return True
            analyze_url = f"{LLMMetadataQueue._server_base_url()}/api/llm/request_analysis"
            body = json.dumps({"paths": paths}).encode("utf-8")
            req = urllib_request.Request(
                analyze_url,
                data=body,
                headers={"Content-Type": "application/json"},
                method="POST",
            )
            with urllib_request.urlopen(req, timeout=8) as resp:
                resp.read()
            return True
        except Exception:
            return False

    @staticmethod
    def _process_file(path: str) -> Optional[dict]:
        """
        Lit les données IA depuis l'API MediaMind AI ou le cache SQLite local.
        Retourne None si le serveur est injoignable ou les données absentes.
        Retourne {"_not_indexed": True} si le serveur est joignable mais le fichier n'est pas encore analysé.
        Ne déclenche PAS d'analyse — utiliser _submit_for_analysis pour cela.
        """
        server_reachable = False
        api_url = os.getenv("MEDIAMIND_API_URL", "http://127.0.0.1:8190/api/llm/enrich").strip()
        if api_url:
            try:
                body = json.dumps({"path": path}).encode("utf-8")
                req = urllib_request.Request(
                    api_url,
                    data=body,
                    headers={"Content-Type": "application/json"},
                    method="POST",
                )
                with urllib_request.urlopen(req, timeout=12) as resp:
                    raw = resp.read().decode("utf-8", errors="ignore")
                parsed = json.loads(raw) if raw else {}
                server_reachable = True
                if parsed.get("ok") and parsed.get("has_signal"):
                    payload = parsed.get("payload") or {}
                    has_required = any(k in payload for k in ("tags", "prompt", "nsfw", "aesthetic"))
                    if has_required and isinstance(payload, dict):
                        payload.setdefault("source", "media_mind_ai_api")
                        payload.setdefault("path", path)
                        payload.setdefault("analyzed_at", datetime.datetime.now().isoformat(timespec="seconds"))
                        return payload
                # Serveur joignable mais données absentes ou mismatch chemin —
                # on tente le fallback SQLite direct avant de déclarer _not_indexed
            except (urllib_error.URLError, urllib_error.HTTPError, TimeoutError, ValueError, OSError):
                pass

        cache_db = os.path.join(os.path.dirname(os.path.abspath(__file__)), "image_cache.db")
        if not os.path.isfile(cache_db):
            return {"_not_indexed": True} if server_reachable else None

        def _query_sqlite(lookup_path: str) -> Optional[dict]:
            try:
                conn = sqlite3.connect(cache_db, check_same_thread=False)
                cur = conn.cursor()

                def safe_json(value):
                    if value is None:
                        return None
                    try:
                        return json.loads(value)
                    except Exception:
                        return None

                tags_row = cur.execute(
                    "SELECT model, tags FROM tags_cache WHERE path = ? ORDER BY rowid DESC LIMIT 1",
                    (lookup_path,),
                ).fetchone()
                prompt_row = cur.execute(
                    "SELECT source, prompt FROM prompt_cache WHERE path = ? LIMIT 1",
                    (lookup_path,),
                ).fetchone()
                nsfw_row = cur.execute(
                    "SELECT model, top_label, danger_score, details FROM nsfw_cache WHERE path = ? ORDER BY rowid DESC LIMIT 1",
                    (lookup_path,),
                ).fetchone()
                aes_row = cur.execute(
                    "SELECT model, avg_score, max_score FROM aes_cache WHERE path = ? ORDER BY rowid DESC LIMIT 1",
                    (lookup_path,),
                ).fetchone()
                face_row = cur.execute(
                    "SELECT COUNT(*) FROM face_cache WHERE path = ? AND face_idx >= 0",
                    (lookup_path,),
                ).fetchone()
                conn.close()

                pl: dict = {
                    "source": "media_mind_ai_sqlite",
                    "analyzed_at": datetime.datetime.now().isoformat(timespec="seconds"),
                    "path": path,
                }
                if tags_row:
                    pl["tags"] = {"model": tags_row[0], "values": safe_json(tags_row[1]) or {}}
                if prompt_row and prompt_row[1]:
                    pl["prompt"] = {"source": prompt_row[0] or "image_metadata", "text": str(prompt_row[1])}
                if nsfw_row:
                    pl["nsfw"] = {
                        "model": nsfw_row[0], "top_label": nsfw_row[1],
                        "danger_score": float(nsfw_row[2]) if nsfw_row[2] is not None else None,
                        "details": safe_json(nsfw_row[3]) or {},
                    }
                if aes_row:
                    pl["aesthetic"] = {
                        "model": aes_row[0],
                        "avg_score": float(aes_row[1]) if aes_row[1] is not None else None,
                        "max_score": float(aes_row[2]) if aes_row[2] is not None else None,
                    }
                if face_row and face_row[0] is not None:
                    pl["faces"] = {"count": int(face_row[0])}
                return pl if any(k in pl for k in ("tags", "prompt", "nsfw", "aesthetic")) else None
            except Exception:
                return None

        # Essai 1 : chemin tel quel
        result = _query_sqlite(path)
        if result:
            return result
        # Essai 2 : chemin normalisé (rétro-compat avec backslash/forward-slash)
        path_norm = os.path.normpath(path)
        if path_norm != path:
            result = _query_sqlite(path_norm)
            if result:
                return result

        return {"_not_indexed": True} if server_reachable else None
def is_scan_entry_selected(entry, selected_categories: dict, selected_files: Optional[dict] = None) -> bool:
    if selected_files and entry.path in selected_files:
        return bool(selected_files[entry.path])
    return bool(selected_categories.get((entry.category, entry.subcategory), False))


@lru_cache(maxsize=256)
def get_video_thumbnail_data_url(path: str, mtime: float, size: int) -> Optional[str]:
    try:
        import av
        from PIL import Image

        with av.open(path) as container:
            if not container.streams.video:
                return None

            stream = container.streams.video[0]
            if container.duration and stream.time_base:
                target_seconds = min(max(float(container.duration) / 1_000_000 * 0.15, 0.0), 5.0)
                target_pts = int(target_seconds / float(stream.time_base))
                container.seek(target_pts, backward=True, any_frame=False, stream=stream)

            for frame in container.decode(stream):
                image = frame.to_image().convert("RGB")
                image.thumbnail((240, 135), Image.Resampling.LANCZOS)
                buffer = io.BytesIO()
                image.save(buffer, format="JPEG", quality=72, optimize=True)
                return "data:image/jpeg;base64," + base64.b64encode(buffer.getvalue()).decode("ascii")
    except Exception:
        return None
    return None


def pick_folder_native() -> str:
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk()
    root.attributes("-topmost", True)
    root.withdraw()
    folder = filedialog.askdirectory()
    root.destroy()
    return folder


async def select_folder_async() -> str:
    return await run.io_bound(pick_folder_native)


def build_video_preview_url(path: str) -> str:
    return f"/preview/video?path={quote(path, safe='')}"


@app.get("/preview/video")
def preview_video_file(path: str):
    decoded_path = unquote(path)
    if not os.path.isfile(decoded_path):
        raise HTTPException(status_code=404, detail="Fichier vidéo introuvable")
    return FileResponse(decoded_path)


@app.get("/preview/image")
def preview_image_file(path: str):
    decoded_path = unquote(path)
    if not os.path.isfile(decoded_path):
        raise HTTPException(status_code=404, detail="Fichier image introuvable")
    return FileResponse(decoded_path)


def get_image_thumbnail_data_url(path: str) -> Optional[str]:
    """Génère une miniature base64 pour une image (JPEG, PNG, WEBP…)."""
    try:
        from PIL import Image as PILImage
        with PILImage.open(path) as img:
            img = img.convert("RGB")
            img.thumbnail((240, 180), PILImage.Resampling.LANCZOS)
            buffer = io.BytesIO()
            img.save(buffer, format="JPEG", quality=72, optimize=True)
            return "data:image/jpeg;base64," + base64.b64encode(buffer.getvalue()).decode("ascii")
    except Exception:
        return None


# ============================================================
# PAGE PRINCIPALE NiceGUI
# ============================================================

def fmt_size(n: int) -> str:
    if not n:
        return "0 o"
    if n >= 1_073_741_824:
        return f"{n / 1_073_741_824:.1f} Go"
    if n >= 1_048_576:
        return f"{n / 1_048_576:.1f} Mo"
    if n >= 1024:
        return f"{n / 1024:.1f} Ko"
    return f"{n} o"


@ui.page("/")
def main_page():
    db_archive = PostgreSQLArchive()
    llm_queue = LLMMetadataQueue(db_archive)

    # ---------------------------------------------------------------
    # Conversion dossiers YYYY/Mois/JJ → YYYY-MM-DD
    # ---------------------------------------------------------------
    def open_convert_date_folders_dialog():
        _MONTHS_FR_REV = {v.lower(): str(k).zfill(2) for k, v in MONTHS_FR.items()}

        def _detect_date_triplets(root: str):
            """Retourne une liste de (year_dir, month_dir, day_dir, year_int, month_int, day_int)."""
            import re
            triplets = []
            for dirpath, dirnames, _ in os.walk(root):
                # Chercher sous-dossiers YYYY
                year_dirs = [d for d in dirnames if re.fullmatch(r"\d{4}", d)]
                for ydir in year_dirs:
                    year_path = os.path.join(dirpath, ydir)
                    year_int = int(ydir)
                    try:
                        month_dirs = os.listdir(year_path)
                    except OSError:
                        continue
                    for mdir in month_dirs:
                        month_path = os.path.join(year_path, mdir)
                        if not os.path.isdir(month_path):
                            continue
                        # Mois FR ou numérique
                        month_num = _MONTHS_FR_REV.get(mdir.lower())
                        if not month_num:
                            if re.fullmatch(r"0?[1-9]|1[0-2]", mdir):
                                month_num = mdir.zfill(2)
                        if not month_num:
                            continue
                        try:
                            day_dirs = os.listdir(month_path)
                        except OSError:
                            continue
                        for ddir in day_dirs:
                            day_path = os.path.join(month_path, ddir)
                            if not os.path.isdir(day_path):
                                continue
                            if re.fullmatch(r"0?[1-9]|[12]\d|3[01]", ddir):
                                triplets.append((
                                    dirpath, ydir, mdir, ddir,
                                    year_int, int(month_num), int(ddir),
                                ))
            return triplets

        with ui.dialog() as dlg, ui.card().classes("w-[680px] max-w-[95vw] bg-gray-900 gap-3"):
            ui.label("Convertir YYYY/Mois/DD → YYYY-MM-DD").classes("text-lg font-bold text-orange-300")
            ui.label(
                "Fusionne les sous-dossiers de date en un seul dossier au format YYYY-MM-DD. "
                "Les dossiers vides intermédiaires seront supprimés."
            ).classes("text-xs text-gray-400")

            conv_root_input = ui.input(
                "Dossier racine à convertir",
                placeholder="ex: D:\\Media-Organisé\\PHOTO-VIDEO",
            ).classes("w-full")

            async def _pick_root():
                folder = await select_folder_async()
                if folder:
                    conv_root_input.set_value(folder)
            ui.button(icon="folder_open", on_click=_pick_root).props("flat dense color=orange").tooltip("Choisir dossier")

            conv_status = ui.label("").classes("text-sm text-gray-300")
            conv_log = ui.log(max_lines=20).classes("w-full text-xs bg-black/40 rounded h-40")

            async def _run_conversion():
                root = conv_root_input.value.strip()
                if not root or not os.path.isdir(root):
                    ui.notify("Dossier introuvable.", type="negative")
                    return
                conv_status.set_text("⏳ Analyse en cours…")
                conv_log.clear()

                def task():
                    triplets = _detect_date_triplets(root)
                    if not triplets:
                        return 0, 0, 0, ["Aucun dossier YYYY/Mois/DD trouvé dans ce répertoire."]
                    moved = 0
                    errors = 0
                    logs = [f"Triplets trouvés : {len(triplets)}"]
                    for base, ydir, mdir, ddir, y, m, d in triplets:
                        src_day = os.path.join(base, ydir, mdir, ddir)
                        date_folder = f"{y:04d}-{m:02d}-{d:02d}"
                        dest_date = os.path.join(base, date_folder)
                        try:
                            os.makedirs(dest_date, exist_ok=True)
                            for fname in os.listdir(src_day):
                                src_file = os.path.join(src_day, fname)
                                dst_file = os.path.join(dest_date, fname)
                                if os.path.exists(dst_file):
                                    dst_file = os.path.join(dest_date, f"{os.path.splitext(fname)[0]}_dup{os.path.splitext(fname)[1]}")
                                shutil.move(src_file, dst_file)
                                moved += 1
                            # Nettoyer dossiers vides
                            for empty in (src_day, os.path.join(base, ydir, mdir), os.path.join(base, ydir)):
                                try:
                                    if os.path.isdir(empty) and not os.listdir(empty):
                                        os.rmdir(empty)
                                except OSError:
                                    pass
                            logs.append(f"✅ {os.path.join(ydir, mdir, ddir)} → {date_folder}")
                        except Exception as e:
                            errors += 1
                            logs.append(f"❌ {os.path.join(ydir, mdir, ddir)} : {e}")
                    return moved, errors, len(triplets), logs

                moved, errors, total, logs = await run.io_bound(task)
                for line in logs:
                    conv_log.push(line)
                conv_status.set_text(
                    f"{'✅' if not errors else '⚠️'} Terminé : {moved} fichiers déplacés, "
                    f"{total} triplets traités, {errors} erreur(s)."
                )

            with ui.row().classes("w-full gap-3 justify-end"):
                ui.button("Annuler", on_click=dlg.close).props("flat color=gray")
                ui.button("Convertir", icon="swap_horiz", on_click=_run_conversion).props("color=orange")

        dlg.open()

    with ui.right_drawer(value=False).classes("bg-gray-900 border-l border-gray-800 w-[38rem] max-w-[95vw]") as log_drawer:
        ui.label("Journaux").classes("text-sm text-gray-300 mb-2")
        log_el = ui.log().classes("w-full h-[50vh] text-xs font-mono")

    with ui.header().classes("bg-gray-900 border-b border-gray-800 px-4 py-2"):
        with ui.row().classes("w-full items-center justify-between"):
            ui.label("Organisateur de disques durs").classes("text-sm text-gray-400")
            with ui.row().classes("items-center gap-3"):
                db_status_label = ui.label("BD: —").classes("text-xs text-gray-500 font-mono")
                ui.button(icon="table_view", on_click=lambda: ui.navigate.to("/bd")).props(
                    "flat text-color=cyan size=sm dense"
                ).tooltip("Explorer la base de données")
                ui.button(icon="drive_file_rename_outline", on_click=lambda: open_convert_date_folders_dialog()).props(
                    "flat text-color=orange size=sm dense"
                ).tooltip("Convertir dossiers YYYY/Mois/DD → YYYY-MM-DD")
                ui.button(icon="terminal", on_click=log_drawer.toggle).props(
                    "flat text-color=white size=sm dense"
                ).tooltip("Journaux")

    # ---- Mise à jour des logs dans l'UI ----
    _scan_log_refs: list = []  # log widgets de l'étape 2 (scan), rempli plus bas

    def update_logs():
        if state.logs:
            for msg in state.logs:
                log_el.push(msg)
                for _ref in _scan_log_refs:
                    try:
                        _ref.push(msg)
                    except Exception:
                        pass
            state.logs.clear()

    def open_custom_rules_dialog(on_saved=None, on_rescan=None):
        """Dialogue léger pour ajouter une catégorie ou un type de projet depuis n'importe quelle étape."""
        with ui.dialog() as dlg, ui.card().classes("w-[780px] max-w-[95vw] bg-gray-900"):
            ui.label("Ajouter une catégorie / un type projet").classes("text-lg font-bold text-cyan-300")

            with ui.card().classes("w-full bg-gray-800 border border-gray-700 p-3"):
                ui.label("Nouvelle catégorie").classes("text-sm font-bold text-cyan-200")
                with ui.row().classes("w-full gap-3 items-end flex-wrap"):
                    cat_name_input = ui.input("Catégorie", placeholder="ex: ARCHIVES")
                    cat_sub_input = ui.input("Sous-catégorie", placeholder="ex: ZIP")
                    cat_ext_input = ui.input("Extensions", placeholder="ex: .zip,.rar,.7z")

                    def add_category_from_dialog():
                        cat_name = (cat_name_input.value or "").strip().upper()
                        subcat = (cat_sub_input.value or "").strip() or "DIVERS"
                        exts = parse_extensions_csv(cat_ext_input.value or "")
                        if not cat_name:
                            ui.notify("Nom de catégorie requis.", type="warning")
                            return
                        if cat_name == "PROJETS":
                            ui.notify("Utilisez l'ajout de type projet pour PROJETS.", type="warning")
                            return
                        if not exts:
                            ui.notify("Ajoutez au moins une extension.", type="warning")
                            return

                        if cat_name not in CATEGORIES:
                            CATEGORIES[cat_name] = {
                                "icon": "category",
                                "color": "cyan",
                                "extensions": {},
                            }

                        for ext in exts:
                            CATEGORIES[cat_name]["extensions"][ext] = subcat

                        save_custom_rules()
                        state.add_log(f"➕ Catégorie ajoutée (dialog): {cat_name}/{subcat} ({', '.join(exts)})")
                        ui.notify(
                            f"Catégorie {cat_name}/{subcat} ajoutée. Pensez à re-scanner.",
                            type="positive",
                        )
                        if on_saved:
                            on_saved()
                        cat_name_input.value = ""
                        cat_sub_input.value = ""
                        cat_ext_input.value = ""

                    ui.button("Ajouter catégorie", icon="add", on_click=add_category_from_dialog).props(
                        "outline color=cyan"
                    )

            with ui.card().classes("w-full bg-gray-800 border border-gray-700 p-3"):
                ui.label("Nouveau type de projet").classes("text-sm font-bold text-orange-200")
                with ui.row().classes("w-full gap-3 items-end flex-wrap"):
                    proj_name_input = ui.input("Type projet", placeholder="ex: FL Studio")
                    proj_ext_input = ui.input("Extension signature", placeholder="ex: .flp")

                    def add_project_from_dialog():
                        proj_name = (proj_name_input.value or "").strip()
                        proj_ext = normalize_extension(proj_ext_input.value or "")
                        if not proj_name:
                            ui.notify("Nom de type projet requis.", type="warning")
                            return
                        if not proj_ext:
                            ui.notify("Extension projet requise.", type="warning")
                            return

                        PROJECT_SIGNATURES[proj_ext] = (proj_name, f"Projet {proj_name}")
                        save_custom_rules()
                        state.add_log(f"➕ Type projet ajouté (dialog): {proj_name} ({proj_ext})")
                        ui.notify(
                            f"Type projet {proj_name} ajouté. Pensez à re-scanner.",
                            type="positive",
                        )
                        if on_saved:
                            on_saved()
                        proj_name_input.value = ""
                        proj_ext_input.value = ""

                    ui.button("Ajouter type projet", icon="add_box", on_click=add_project_from_dialog).props(
                        "outline color=orange"
                    )

            with ui.row().classes("w-full justify-between items-center mt-2"):
                if on_rescan:
                    async def rescan_now_from_dialog():
                        result = on_rescan()
                        if asyncio.iscoroutine(result):
                            await result

                    ui.button(
                        "Re-scanner maintenant",
                        icon="refresh",
                        on_click=rescan_now_from_dialog,
                    ).props("outline color=blue")

                ui.button("Fermer", icon="close", on_click=dlg.close).props("flat color=gray")

        dlg.open()

    ui.timer(0.4, update_logs)

    # ---- Pied de page (barre de progression) ----
    with ui.footer().classes(
        "bg-gray-900 border-t border-gray-800 px-4 py-0 flex items-center gap-4 h-8"
    ):
        ui.label().bind_text_from(state, "status_text").classes(
            "text-blue-300 font-mono text-xs truncate max-w-[35%]"
        )
        ui.spinner(size="xs", color="blue").bind_visibility_from(
            state, "is_loading"
        ).classes("shrink-0")
        ui.linear_progress(value=0, show_value=False).bind_value_from(
            state, "progress"
        ).classes("flex-grow h-1.5 rounded")
        ui.button(
            "ANNULER", icon="cancel",
            on_click=lambda: setattr(state, "cancel_flag", True)
        ).props("color=red size=sm dense outline").classes("text-xs").bind_visibility_from(
            state, "is_scanning"
        )

    # ==================================================================
    # STEPPER PRINCIPAL
    # ==================================================================

    with ui.stepper(
        value=("Étape 7" if START_IN_LLM_MODE else "Étape 1")
    ).props("vertical=false animated flat header-nav").classes(
        "w-full h-[calc(100vh-88px)] bg-[#0f0f17]"
    ) as stepper:

        step_refresh_refs = {
            "uncategorized_table": None,
            "uncategorized_label": None,
            "installers_ui": None,
            "summary_ui": None,
            "video_selection_ui": None,
            "ia_images_ui": None,
            "real_images_ui": None,
        }

        async def go_from_step2():
            """Lazy-refresh étape 3 au moment de la navigation depuis l'étape 2."""
            state.is_loading = True
            state.status_text = "⏳ Chargement sélection..."
            await asyncio.sleep(0)
            selection_ui.refresh()
            await asyncio.sleep(0)
            if step_refresh_refs["video_selection_ui"]:
                step_refresh_refs["video_selection_ui"]()
            if step_refresh_refs["ia_images_ui"]:
                step_refresh_refs["ia_images_ui"]()
            if step_refresh_refs["real_images_ui"]:
                step_refresh_refs["real_images_ui"]()
            state.is_loading = False
            state.status_text = "✅ Prêt"
            stepper.next()

        async def go_from_step3():
            if state.uncategorized:
                state.is_loading = True
                state.status_text = "⏳ Chargement fichiers non catégorisés..."
                await asyncio.sleep(0)
                if step_refresh_refs["uncategorized_table"]:
                    step_refresh_refs["uncategorized_table"]()
                if step_refresh_refs["uncategorized_label"]:
                    step_refresh_refs["uncategorized_label"]()
                state.is_loading = False
                state.status_text = "✅ Prêt"
                stepper.set_value("Étape 4")
                return
            if state.installer_files:
                state.is_loading = True
                state.status_text = "⏳ Chargement installeurs..."
                await asyncio.sleep(0)
                if step_refresh_refs["installers_ui"]:
                    step_refresh_refs["installers_ui"]()
                state.is_loading = False
                state.status_text = "✅ Prêt"
                stepper.set_value("Étape 5")
                return
            state.is_loading = True
            state.status_text = "⏳ Chargement résumé..."
            await asyncio.sleep(0)
            if step_refresh_refs["summary_ui"]:
                step_refresh_refs["summary_ui"]()
            state.is_loading = False
            state.status_text = "✅ Prêt"
            stepper.set_value("Étape 6")

        async def go_from_step4():
            if state.installer_files:
                state.is_loading = True
                state.status_text = "⏳ Chargement installeurs..."
                await asyncio.sleep(0)
                if step_refresh_refs["installers_ui"]:
                    step_refresh_refs["installers_ui"]()
                state.is_loading = False
                state.status_text = "✅ Prêt"
                stepper.set_value("Étape 5")
                return
            state.is_loading = True
            state.status_text = "⏳ Chargement résumé..."
            await asyncio.sleep(0)
            if step_refresh_refs["summary_ui"]:
                step_refresh_refs["summary_ui"]()
            state.is_loading = False
            state.status_text = "✅ Prêt"
            stepper.set_value("Étape 6")

        async def go_from_step5():
            state.is_loading = True
            state.status_text = "⏳ Chargement résumé..."
            await asyncio.sleep(0)
            if step_refresh_refs["summary_ui"]:
                step_refresh_refs["summary_ui"]()
            state.is_loading = False
            state.status_text = "✅ Prêt"
            stepper.set_value("Étape 6")

        # ================================================================
        # ÉTAPE 1 — CONFIGURATION
        # ================================================================
        with ui.step("Étape 1", title="Configuration", icon="settings"):
            with ui.column().classes("w-full max-w-4xl mx-auto p-6 gap-6"):

                ui.label("Configurer les sources et la destination").classes(
                    "text-2xl font-bold text-orange-400"
                )

                # Sources
                with ui.card().classes("w-full bg-gray-800 border border-gray-700 p-4 gap-3"):
                    ui.label("Dossiers / Disques sources").classes("text-lg font-semibold text-blue-300")
                    ui.label("Ajoutez les disques ou dossiers à organiser.").classes("text-sm text-gray-400")

                    sources_container = ui.column().classes("w-full gap-2")

                    def refresh_sources():
                        sources_container.clear()
                        with sources_container:
                            for i, src in enumerate(state.sources):
                                with ui.row().classes("w-full items-center gap-2 bg-gray-900 px-3 py-2 rounded-lg"):
                                    ui.icon("hard_drive", size="1.2rem").classes("text-blue-400")
                                    ui.label(src).classes("flex-grow font-mono text-sm truncate")
                                    ui.button(
                                        icon="delete",
                                        on_click=lambda _, idx=i: (state.sources.pop(idx), refresh_sources())
                                    ).props("flat round dense text-color=red")

                    refresh_sources()

                    async def add_source():
                        folder = await select_folder_async()
                        if folder and folder not in state.sources:
                            state.sources.append(folder)
                            refresh_sources()

                    with ui.row().classes("gap-2 mt-2"):
                        ui.button("+ Ajouter un dossier", icon="add", on_click=add_source).props(
                            "outline color=blue"
                        )
                        ui.button("+ Disque entier (C:)", icon="storage",
                            on_click=lambda: (
                                state.sources.append("C:\\") or state.sources.__class__,
                                refresh_sources()
                            ) if "C:\\" not in state.sources else None
                        ).props("outline color=gray")

                # Destination
                with ui.card().classes("w-full bg-gray-800 border border-gray-700 p-4 gap-3"):
                    ui.label("Dossier de destination").classes("text-lg font-semibold text-green-300")
                    ui.label(
                        "Choisissez où les fichiers organisés seront copiés/déplacés."
                    ).classes("text-sm text-gray-400")

                    dest_input = ui.input(
                        "Chemin de destination", value=state.destination
                    ).classes("w-full font-mono text-sm").bind_value(state, "destination")

                    async def pick_dest():
                        folder = await select_folder_async()
                        if folder:
                            dest_input.value = folder
                            state.destination = folder

                    with ui.row().classes("gap-2 mt-1"):
                        ui.button("Parcourir", icon="folder_open", on_click=pick_dest).props(
                            "outline color=green"
                        )
                        ui.label("Ou réorganiser sur place (laisser vide)").classes(
                            "text-xs text-gray-500 self-center"
                        )

                # Mode
                with ui.card().classes("w-full bg-gray-800 border border-gray-700 p-4 gap-3"):
                    ui.label("Mode de transfert").classes("text-lg font-semibold text-yellow-300")
                    with ui.row().classes("gap-4"):
                        ui.radio(
                            {"copy": "📋 Copier (conserve l'original)", "move": "✂️ Déplacer (supprime l'original)"},
                            value="copy"
                        ).bind_value(state, "mode").classes("text-sm")

                    ui.switch(
                        "Supprimer les dossiers source devenus vides (après déplacement)",
                        value=False,
                    ).bind_value(state, "delete_empty_sources_after_move").bind_visibility_from(
                        state, "mode", backward=lambda m: m == "move"
                    ).classes("text-sm text-amber-200")
                    ui.label(
                        "Sécurité: un dossier source n'est supprimé que s'il est réellement vide en fin d'opération."
                    ).bind_visibility_from(
                        state, "mode", backward=lambda m: m == "move"
                    ).classes("text-xs text-gray-400")

                with ui.stepper_navigation():
                    ui.button(
                        "Suivant → Scanner",
                        icon="arrow_forward",
                        on_click=stepper.next
                    ).classes("bg-orange-600 hover:bg-orange-500 font-bold px-6").props(
                        "size=lg"
                    )

        # ================================================================
        # ÉTAPE 2 — SCAN
        # ================================================================
        with ui.step("Étape 2", title="Scan", icon="search"):
            with ui.column().classes("w-full max-w-5xl mx-auto p-6 gap-4"):

                ui.label("Analyse des fichiers").classes("text-2xl font-bold text-blue-400")

                scan_status_label = ui.label("Prêt à scanner.").classes("text-gray-300")

                # Tableau de résultats
                scan_table_container = ui.column().classes("w-full")

                @ui.refreshable
                def scan_results_table():
                    scan_table_container.clear()
                    with scan_table_container:
                        if not state.scan_stats:
                            ui.label("Les résultats du scan apparaîtront ici.").classes(
                                "text-gray-500 text-sm m-4"
                            )
                            return

                        rows = []
                        for cat, stats in state.scan_stats.items():
                            icon = CATEGORIES.get(cat, {}).get("icon", "folder")
                            rows.append({
                                "categorie": cat,
                                "icon": icon,
                                "count": f"{stats.get('count', 0):,}",
                                "taille": fmt_size(stats.get("size", 0)),
                            })

                        with ui.table(
                            columns=[
                                {"name": "categorie", "label": "Catégorie", "field": "categorie", "align": "left"},
                                {"name": "count", "label": "Fichiers", "field": "count", "align": "right"},
                                {"name": "taille", "label": "Taille totale", "field": "taille", "align": "right"},
                            ],
                            rows=rows,
                        ).classes("w-full bg-gray-800 text-white"):
                            pass

                scan_results_table()

                # Journal inline visible pendant le scan
                scan_inline_log = ui.log(max_lines=30).classes(
                    'w-full text-xs font-mono bg-black/50 rounded border border-gray-700 h-36'
                ).bind_visibility_from(state, 'is_scanning')
                _scan_log_refs.append(scan_inline_log)
                ui.linear_progress(show_value=False).bind_value_from(
                    state, 'progress'
                ).bind_visibility_from(state, 'is_scanning').classes('w-full h-1.5 rounded')

                async def run_scan():
                    if not state.sources:
                        ui.notify("Ajoutez au moins une source !", type="warning")
                        return

                    state.reset_scan()
                    state.is_scanning = True
                    scan_status_label.set_text("Scan en cours...")
                    scan_results_table.refresh()

                    def task():
                        def _prog_cb(v, msg='', *_):
                            state.progress = v
                            if msg:
                                state.status_text = msg
                        scanner = DiskScanner(
                            log_callback=state.add_log,
                            progress_callback=_prog_cb,
                        )
                        results, project_roots, uncategorized, installers, project_evidence = scanner.scan(state.sources)
                        stats = DiskScanner.compute_stats(results, project_roots, uncategorized, installers)
                        state.scan_results = results
                        state.project_roots = project_roots
                        state.project_detection_evidence = project_evidence
                        state.uncategorized = uncategorized
                        state.installer_files = installers
                        state.installer_decisions = {
                            e.path: ("assign:DOCUMENTS" if e.ext.lower() in {".zip", ".rar"} else "keep")
                            for e in installers
                        }
                        state.scan_stats = stats
                        state.add_log(
                            f"✅ Scan terminé : {sum(s['count'] for s in stats.values()):,} fichiers trouvés."
                        )

                    await run.io_bound(task)
                    state.is_scanning = False
                    nb = sum(s.get('count', 0) for s in state.scan_stats.values())
                    scan_status_label.set_text(f"Scan terminé : {nb:,} fichiers trouvés.")
                    state.status_text = f"✅ {nb:,} fichiers — Prêt"
                    state.progress = 0.0

                    # Rafraîchir uniquement le tableau visible (léger : ~10 lignes de catégories).
                    # Les autres étapes se rafraîchissent en lazy via go_from_step2/3/4/5.
                    await asyncio.sleep(0.1)
                    scan_results_table.refresh()

                with ui.row().classes("gap-4"):
                    ui.button(
                        "🔍 Lancer le scan", icon="search", on_click=run_scan
                    ).classes("bg-blue-600 hover:bg-blue-500 font-bold px-6").props("size=lg")

                with ui.stepper_navigation():
                    ui.button("← Retour", on_click=stepper.previous).props("flat color=gray")
                    ui.button(
                        "Suivant → Sélection", icon="arrow_forward", on_click=go_from_step2
                    ).classes("bg-orange-600 hover:bg-orange-500 font-bold px-6")

        # ================================================================
        # ÉTAPE 3 — SÉLECTION DES CATÉGORIES
        # ================================================================
        with ui.step("Étape 3", title="Sélection", icon="checklist"):
            with ui.column().classes("w-full max-w-none p-6 gap-4"):

                ui.label("Sélectionner les catégories à organiser").classes(
                    "text-2xl font-bold text-green-400"
                )
                ui.label(
                    "Cochez les types de fichiers que vous souhaitez organiser."
                ).classes("text-sm text-gray-400")

                async def rescan_from_step3():
                    if not state.sources:
                        ui.notify("Ajoutez au moins une source dans l'Étape 1.", type="warning")
                        return
                    stepper.set_value("Étape 2")
                    await run_scan()
                    stepper.set_value("Étape 3")
                    selection_ui.refresh()

                with ui.row().classes("w-full justify-end mb-2"):
                    ui.button(
                        "Ajouter catégorie/projet",
                        icon="playlist_add",
                        on_click=lambda: open_custom_rules_dialog(
                            on_saved=lambda: selection_ui.refresh(),
                            on_rescan=rescan_from_step3,
                        ),
                    ).props("outline color=cyan")

                # ---- zone sélection ----
                with ui.column().classes("w-full gap-2"):
                    
                    # Placeholder pour référence croisée (refresh du panel musical)
                    music_metadata_refresh_ref = {}
                    video_selection_refresh_ref = {}
                    video_filters = {"query": ""}

                    def get_video_entries():
                        return state.scan_results.get("PHOTO-VIDEO", {}).get("Video", [])

                    def get_filtered_video_entries(limit: int = 60):
                        query = (video_filters.get("query") or "").strip().lower()
                        entries = [
                            entry for entry in get_video_entries()
                            if not query or query in f"{entry.name} {entry.path}".lower()
                        ]
                        return entries[:limit]

                    def sync_video_subcategory_state():
                        video_entries = get_video_entries()
                        if not video_entries:
                            return
                        state.selected_categories[("PHOTO-VIDEO", "Video")] = all(
                            is_scan_entry_selected(entry, state.selected_categories, state.selected_files)
                            for entry in video_entries
                        )

                    async def refresh_selection_dependent_panels():
                        state.is_loading = True
                        await asyncio.sleep(0)
                        selection_ui.refresh()
                        if video_selection_refresh_ref.get("fn"):
                            video_selection_refresh_ref["fn"]()
                        if music_metadata_refresh_ref.get("fn"):
                            music_metadata_refresh_ref["fn"]()
                        state.is_loading = False
                    _sel_expand: dict = {}   # état collapse/expand des catégories
                    _sub_expand: dict = {}   # état collapse/expand des listes de fichiers par sous-catégorie
                    _sub_page:   dict = {}   # page courante par sous-catégorie (pagination)
                    _file_checked: dict = {} # état checkbox par fichier (True=coché, défaut True)

                    @ui.refreshable
                    def selection_ui():
                        if not state.scan_results and not state.project_roots:
                            ui.label("⬅ Lancez d'abord le scan à l'Étape 2.").classes(
                                "text-gray-500 m-4"
                            )
                            return

                        with ui.column().classes("w-full gap-3"):
                            for cat, subcats in state.scan_results.items():
                                if cat in ("PROJETS", "PHOTO-VIDEO", "MUSIQUE-AUDIO"):
                                    continue
                                cat_info = CATEGORIES.get(cat, {})
                                color = cat_info.get("color", "gray")
                                icon_name = cat_info.get("icon", "folder")
                                stats = state.scan_stats.get(cat, {})
                                sub_keys = list(subcats.keys())
                                n_sel = sum(
                                    1 for s in sub_keys
                                    if state.selected_categories.get((cat, s), True)
                                )
                                is_expanded = _sel_expand.get(cat, True)

                                with ui.card().classes("w-full bg-gray-800 border border-gray-700 p-0 overflow-hidden"):
                                    # ── En-tête cliquable ────────────────────────────────
                                    with ui.row().classes(
                                        "w-full items-center gap-3 p-3 cursor-pointer select-none"
                                    ).on('click', lambda c=cat: (
                                        _sel_expand.update({c: not _sel_expand.get(c, True)}),
                                        selection_ui.refresh(),
                                    )):
                                        ui.icon(icon_name, size="1.4rem").classes(f"text-{color}-400")
                                        ui.label(cat).classes("text-lg font-bold flex-grow")
                                        ui.label(
                                            f"{n_sel}/{len(sub_keys)} types · "
                                            f"{stats.get('count', 0):,} fichiers — "
                                            f"{fmt_size(stats.get('size', 0))}"
                                        ).classes("text-sm text-gray-400")
                                        ui.icon(
                                            "expand_less" if is_expanded else "expand_more",
                                            size="1.2rem",
                                        ).classes("text-gray-500 ml-auto shrink-0")

                                    # ── Corps (sous-catégories) ──────────────────────────
                                    if is_expanded:
                                        with ui.column().classes("w-full px-3 pb-3 gap-2"):

                                            # Boutons Tout / Rien
                                            with ui.row().classes("gap-2 mb-1"):
                                                async def _select_all(c=cat, subs=sub_keys):
                                                    for s in subs:
                                                        state.selected_categories[(c, s)] = True
                                                    if c == "PHOTO-VIDEO" and "Video" in subs:
                                                        for entry in get_video_entries():
                                                            state.selected_files[entry.path] = True
                                                    await refresh_selection_dependent_panels()

                                                async def _deselect_all(c=cat, subs=sub_keys):
                                                    for s in subs:
                                                        state.selected_categories[(c, s)] = False
                                                    if c == "PHOTO-VIDEO" and "Video" in subs:
                                                        for entry in get_video_entries():
                                                            state.selected_files[entry.path] = False
                                                    await refresh_selection_dependent_panels()

                                                ui.button("✅ Tout organiser", on_click=_select_all).props(
                                                    f"flat dense size=sm color={color}"
                                                )
                                                ui.button("🚫 Tout ignorer", on_click=_deselect_all).props(
                                                    "flat dense size=sm color=gray"
                                                )

                                            # Grille des sous-catégories
                                            with ui.grid(columns=3).classes("w-full gap-2"):
                                                for subcat, entries in subcats.items():
                                                    key = (cat, subcat)
                                                    if key not in state.selected_categories:
                                                        state.selected_categories[key] = True
                                                    count = len(entries)
                                                    size = sum(e.size for e in entries)
                                                    is_sel = state.selected_categories[key]

                                                    bg_cls  = "bg-gray-700" if is_sel else "bg-gray-900"
                                                    bdr_cls = f"border-{color}-600" if is_sel else "border-gray-700"

                                                    async def _on_action(e, k=key, c=cat, sc=subcat, es=entries):
                                                        new_is_org = (e.value == "Organiser")
                                                        state.selected_categories[k] = new_is_org
                                                        # Recalculer les overrides par fichier selon la nouvelle action
                                                        if c != "PHOTO-VIDEO" or sc != "Video":
                                                            for entry in es:
                                                                chk = _file_checked.get(entry.path, True)
                                                                will_org = (chk == new_is_org)
                                                                if will_org == new_is_org:  # égal au défaut catégorie
                                                                    state.selected_files.pop(entry.path, None)
                                                                else:
                                                                    state.selected_files[entry.path] = will_org
                                                        else:
                                                            for entry in es:
                                                                state.selected_files[entry.path] = new_is_org
                                                        await refresh_selection_dependent_panels()

                                                    with ui.card().classes(
                                                        f"w-full {bg_cls} border {bdr_cls} p-3"
                                                    ):
                                                        with ui.row().classes("w-full items-center gap-2 mb-1"):
                                                            ui.label(subcat).classes(
                                                                "flex-grow font-mono font-bold text-sm text-white"
                                                            )
                                                            sub_exp_key = (cat, subcat)
                                                            is_sub_exp = _sub_expand.get(sub_exp_key, False)
                                                            ui.label(
                                                                f"{'▲' if is_sub_exp else '▼'} {count:,} fichiers"
                                                            ).classes(
                                                                "text-xs text-blue-400 cursor-pointer underline shrink-0"
                                                            ).on('click', lambda k=sub_exp_key: (
                                                                _sub_expand.update({k: not _sub_expand.get(k, False)}),
                                                                selection_ui.refresh(),
                                                            ))
                                                        ui.label(fmt_size(size)).classes("text-xs text-gray-500 mb-2")
                                                        ui.select(
                                                            options=["Organiser", "Ignorer"],
                                                            value="Organiser" if is_sel else "Ignorer",
                                                            on_change=_on_action,
                                                        ).classes("w-full text-xs").props("dense outlined")

                                                        # ── Liste dépliable des fichiers ──────────────────
                                                        if _sub_expand.get((cat, subcat), False):
                                                            _PAGE_SIZE = 200
                                                            sub_key = (cat, subcat)
                                                            page = _sub_page.get(sub_key, 0)
                                                            total_pages = max(1, (len(entries) + _PAGE_SIZE - 1) // _PAGE_SIZE)
                                                            page = max(0, min(page, total_pages - 1))
                                                            shown = entries[page * _PAGE_SIZE:(page + 1) * _PAGE_SIZE]

                                                            with ui.column().classes(
                                                                "w-full mt-2 gap-0.5 bg-gray-950 rounded "
                                                                "border border-gray-700"
                                                            ):
                                                                # Boutons Sélectionner tout / Désélectionner tout
                                                                with ui.row().classes(
                                                                    "w-full items-center gap-2 px-3 py-2 border-b border-gray-700"
                                                                ):
                                                                    async def _chk_all(es=entries, k=sub_key, io=is_sel):
                                                                        for ent in es:
                                                                            _file_checked.pop(ent.path, None)
                                                                            # checked=True + is_org = is_org (défaut) → retirer override
                                                                            state.selected_files.pop(ent.path, None)
                                                                        await asyncio.sleep(0)
                                                                        selection_ui.refresh()

                                                                    async def _unchk_all(es=entries, k=sub_key, io=is_sel):
                                                                        for ent in es:
                                                                            _file_checked[ent.path] = False
                                                                            # checked=False → will_org = not is_org
                                                                            will_org = not io
                                                                            if will_org == io:
                                                                                state.selected_files.pop(ent.path, None)
                                                                            else:
                                                                                state.selected_files[ent.path] = will_org
                                                                        await asyncio.sleep(0)
                                                                        selection_ui.refresh()

                                                                    ui.button(
                                                                        "✅ Tout sélectionner", on_click=_chk_all
                                                                    ).props("flat dense size=sm color=positive")
                                                                    ui.button(
                                                                        "❌ Tout désélectionner", on_click=_unchk_all
                                                                    ).props("flat dense size=sm color=grey-7")
                                                                    ui.label(
                                                                        f"{sum(1 for ent in entries if _file_checked.get(ent.path, True)):,} "
                                                                        f"/ {len(entries):,} sélectionnés"
                                                                    ).classes("text-xs text-gray-400 ml-auto")

                                                                # Barre de pagination
                                                                if total_pages > 1:
                                                                    with ui.row().classes(
                                                                        "w-full items-center gap-2 px-3 py-2 "
                                                                        "border-b border-gray-700"
                                                                    ):
                                                                        ui.button(
                                                                            icon="chevron_left",
                                                                            on_click=lambda k=sub_key, p=page: (
                                                                                _sub_page.update({k: max(0, p - 1)}),
                                                                                selection_ui.refresh(),
                                                                            ),
                                                                        ).props("flat dense round").classes(
                                                                            "text-gray-400"
                                                                        ).set_enabled(page > 0)
                                                                        ui.label(
                                                                            f"Page {page + 1} / {total_pages}  "
                                                                            f"({page * _PAGE_SIZE + 1}–"
                                                                            f"{min((page + 1) * _PAGE_SIZE, len(entries)):,} "
                                                                            f"sur {len(entries):,})"
                                                                        ).classes("text-xs text-gray-400 flex-grow text-center")
                                                                        ui.button(
                                                                            icon="chevron_right",
                                                                            on_click=lambda k=sub_key, p=page, tp=total_pages: (
                                                                                _sub_page.update({k: min(tp - 1, p + 1)}),
                                                                                selection_ui.refresh(),
                                                                            ),
                                                                        ).props("flat dense round").classes(
                                                                            "text-gray-400"
                                                                        ).set_enabled(page < total_pages - 1)

                                                                # Fichiers de la page avec checkboxes
                                                                with ui.column().classes(
                                                                    "w-full max-h-96 overflow-y-auto gap-0 p-2"
                                                                ):
                                                                    for entry in shown:
                                                                        chk_val = _file_checked.get(entry.path, True)

                                                                        def _on_file_chk(ev, p=entry.path, io=is_sel):
                                                                            _file_checked[p] = ev.value
                                                                            will_org = (ev.value == io)
                                                                            if will_org == io:
                                                                                state.selected_files.pop(p, None)
                                                                            else:
                                                                                state.selected_files[p] = will_org

                                                                        with ui.row().classes(
                                                                            "w-full items-center gap-2 py-0.5 "
                                                                            "hover:bg-gray-800/60 rounded px-1"
                                                                        ):
                                                                            ui.checkbox(
                                                                                value=chk_val,
                                                                                on_change=_on_file_chk,
                                                                            ).props("dense")
                                                                            ui.label(entry.name).classes(
                                                                                "text-xs font-mono flex-grow break-all leading-tight " +
                                                                                ("text-gray-200" if chk_val else "text-gray-500 line-through")
                                                                            )
                                                                            ui.label(fmt_size(entry.size)).classes(
                                                                                "text-xs text-gray-500 shrink-0"
                                                                            )

                            # ---- Projets ----
                            if state.project_roots:
                                proj_subcats: dict = defaultdict(list)
                                for root, subcat in state.project_roots:
                                    proj_subcats[subcat].append(root)

                                with ui.card().classes(
                                    "w-full bg-gray-800 border border-orange-700 p-4"
                                ):
                                    with ui.row().classes("w-full items-center gap-3 mb-2"):
                                        ui.icon("folder_special", size="1.4rem").classes("text-orange-400")
                                        ui.label("PROJETS").classes("text-lg font-bold flex-grow")
                                        ui.label(
                                            f"{len(state.project_roots)} projets détectés"
                                        ).classes("text-sm text-orange-300")

                                    with ui.grid(columns=3).classes("w-full gap-2"):
                                        for subcat, roots in proj_subcats.items():
                                            key = ("PROJETS", subcat)
                                            if key not in state.selected_categories:
                                                state.selected_categories[key] = False

                                            with ui.row().classes(
                                                "items-center gap-2 bg-gray-900 px-3 py-2 rounded-lg"
                                            ):
                                                chk = ui.checkbox(
                                                    value=state.selected_categories[key]
                                                )
                                                
                                                def _on_proj_checkbox_change(e, k=key):
                                                    state.selected_categories.update({k: e.value})
                                                    if music_metadata_refresh_ref.get('fn'):
                                                        music_metadata_refresh_ref['fn']()
                                                
                                                chk.on_value_change(_on_proj_checkbox_change)
                                                with ui.column().classes("flex-grow gap-0"):
                                                    ui.label(subcat).classes("text-sm font-mono font-bold")
                                                    ui.label(f"{len(roots)} projet(s)").classes(
                                                        "text-xs text-gray-400"
                                                    )

                                    if state.project_detection_evidence:
                                        with ui.expansion("Vérifier les projets détectés", icon="fact_check").classes("w-full mt-2"):
                                            ui.label("Détection par extension marqueur (vérifiez que chaque projet est valide avant exécution).\n").classes(
                                                "text-xs text-gray-400"
                                            )
                                            with ui.column().classes("w-full max-h-48 overflow-y-auto gap-1"):
                                                for root_path in sorted(state.project_detection_evidence.keys()):
                                                    info = state.project_detection_evidence.get(root_path, {})
                                                    subcat = info.get("subcategory", "?")
                                                    marker = info.get("marker_ext", "?")
                                                    with ui.row().classes("w-full items-center gap-2 bg-gray-900 px-2 py-1 rounded"):
                                                        ui.label(marker).classes("text-xs font-mono text-orange-300 w-14")
                                                        ui.label(subcat).classes("text-xs text-orange-200 w-36")
                                                        ui.label(root_path).classes("text-xs font-mono text-gray-400 break-all")

                    selection_ui()

                with ui.tabs().classes("w-full bg-gray-900/60 rounded-t-lg border border-gray-700 mt-3") as main_media_tabs:
                    tab_videos = ui.tab("Vidéos", icon="movie").props("no-caps")
                    tab_images = ui.tab("Images", icon="image").props("no-caps")
                    tab_music  = ui.tab("Musique & Audio", icon="music_note").props("no-caps")

                image_panel_refresh_refs = {}

                with ui.tab_panels(main_media_tabs, value=tab_videos).classes("w-full bg-gray-800 border border-gray-700 border-t-0 rounded-b-lg p-0"):

                    # ===== TAB VIDÉOS =====
                    with ui.tab_panel(tab_videos):
                        with ui.column().classes("w-full gap-3 p-3"):

                            # Checkbox sélection globale vidéos
                            video_entries_tab = state.scan_results.get("PHOTO-VIDEO", {}).get("Video", [])
                            if video_entries_tab:
                                vid_key = ("PHOTO-VIDEO", "Video")
                                if vid_key not in state.selected_categories:
                                    state.selected_categories[vid_key] = True
                                with ui.row().classes("items-center gap-3 flex-wrap p-2 bg-gray-900/50 rounded mb-1"):
                                    chk_vid_all = ui.checkbox(
                                        "Inclure les vidéos dans l'organisation",
                                        value=state.selected_categories.get(vid_key, True),
                                    )
                                    async def _on_vid_all_tab(e):
                                        state.selected_categories[("PHOTO-VIDEO", "Video")] = e.value
                                        for ent in get_video_entries():
                                            state.selected_files[ent.path] = e.value
                                        await refresh_selection_dependent_panels()
                                    chk_vid_all.on_value_change(_on_vid_all_tab)
                                    ui.label(
                                        f"{len(video_entries_tab):,} vidéos — {fmt_size(sum(e.size for e in video_entries_tab))}"
                                    ).classes("text-sm text-teal-300 font-bold")
                            else:
                                ui.label("Aucune vidéo détectée lors du scan.").classes("text-gray-400 text-sm italic p-2")

                            # Contrôles sélection individuelle des vidéos
                            with ui.row().classes("w-full items-end gap-3 flex-wrap mb-2"):
                                video_search_input = ui.input(
                                    "Rechercher une vidéo",
                                    placeholder="Nom ou chemin",
                                ).classes("min-w-[340px] flex-grow")
                                video_search_input.bind_value(video_filters, "query")

                                skip_date_chk = ui.checkbox(
                                    "Déposer dans DiversVideo (sans date)",
                                    value=state.video_skip_date_hierarchy,
                                ).classes("text-xs text-teal-200")
                                skip_date_chk.on_value_change(
                                    lambda e: (
                                        setattr(state, "video_skip_date_hierarchy", bool(e.value)),
                                        video_selection_refresh_ref.get("fn") and video_selection_refresh_ref["fn"](),
                                    )
                                )

                                no_date_folder_chk = ui.checkbox(
                                    "Utiliser le dossier VideosSansDate",
                                    value=(state.video_no_date_folder_name == "VideosSansDate"),
                                ).classes("text-xs text-teal-200")
                                no_date_folder_chk.on_value_change(
                                    lambda e: (
                                        setattr(
                                            state,
                                            "video_no_date_folder_name",
                                            "VideosSansDate" if bool(e.value) else "DiversVideo",
                                        ),
                                        video_selection_refresh_ref.get("fn") and video_selection_refresh_ref["fn"](),
                                    )
                                )

                                ui.button(
                                    "Tout sélectionner",
                                    icon="done_all",
                                    on_click=lambda: (
                                        [state.selected_files.update({entry.path: True}) for entry in get_video_entries()],
                                        sync_video_subcategory_state(),
                                        refresh_selection_dependent_panels(),
                                    ),
                                ).props("outline color=teal")

                                ui.button(
                                    "Tout désélectionner",
                                    icon="remove_done",
                                    on_click=lambda: (
                                        [state.selected_files.update({entry.path: False}) for entry in get_video_entries()],
                                        sync_video_subcategory_state(),
                                        refresh_selection_dependent_panels(),
                                    ),
                                ).props("outline color=gray")

                                def open_bulk_video_edit_dialog():
                                    targets = get_filtered_video_entries()
                                    if not targets:
                                        ui.notify("Aucune vidéo visible à modifier.", type="warning")
                                        return

                                    with ui.dialog() as dlg, ui.card().classes("w-[740px] max-w-[96vw] bg-gray-900 border border-teal-700"):
                                        ui.label("Édition en lot des vidéos visibles").classes("text-lg font-bold text-teal-200")
                                        ui.label(f"{len(targets)} vidéo(s) seront modifiées.").classes("text-sm text-gray-400")

                                        bulk_folder = ui.input(
                                            "Dossier final relatif (ex: cosmikpratique)",
                                            placeholder=(
                                                f"Laisser vide pour: {state.video_no_date_folder_name}"
                                                if state.video_skip_date_hierarchy
                                                else "Laisser vide pour: YYYY/MM/DD ; sinon: dossier/YYYY/MM/DD"
                                            ),
                                        ).classes("w-full")
                                        bulk_name = ui.input(
                                            "Nom de base (optionnel)",
                                            placeholder="ex: PratiqueCosmikP",
                                        ).classes("w-full")

                                        with ui.row().classes("w-full justify-end gap-2 mt-2"):
                                            ui.button("Annuler", on_click=dlg.close).props("flat color=gray")

                                            def apply_bulk_changes():
                                                folder_value = (bulk_folder.value or "").strip()
                                                name_value = (bulk_name.value or "").strip()
                                                if not folder_value and not name_value:
                                                    ui.notify("Rien à appliquer.", type="warning")
                                                    return

                                                for idx, entry in enumerate(targets, start=1):
                                                    override = state.video_output_overrides.get(entry.path, {})
                                                    if folder_value:
                                                        override["folder"] = folder_value
                                                    if name_value:
                                                        override["filename"] = f"{name_value}{idx:03d}" if len(targets) > 1 else name_value
                                                    state.video_output_overrides[entry.path] = override

                                                if video_selection_refresh_ref.get("fn"):
                                                    video_selection_refresh_ref["fn"]()
                                                ui.notify(f"{len(targets)} vidéo(s) mises à jour.", type="positive")
                                                dlg.close()

                                            ui.button("Appliquer", on_click=apply_bulk_changes).props("color=teal")

                                    dlg.open()

                                ui.button(
                                    "Éditer visibles",
                                    icon="edit",
                                    on_click=open_bulk_video_edit_dialog,
                                ).props("outline color=teal")

                                def clear_bulk_video_edits():
                                    targets = get_filtered_video_entries(limit=100000)
                                    if not targets:
                                        ui.notify("Aucune vidéo visible.", type="warning")
                                        return
                                    for entry in targets:
                                        state.video_output_overrides.pop(entry.path, None)
                                    if video_selection_refresh_ref.get("fn"):
                                        video_selection_refresh_ref["fn"]()
                                    ui.notify("Personnalisations vidéo supprimées (filtre courant).", type="info")

                                ui.button(
                                    "Réinitialiser visibles",
                                    icon="restart_alt",
                                    on_click=clear_bulk_video_edits,
                                ).props("outline color=gray")

                            @ui.refreshable
                            def video_selection_ui():
                                video_entries = get_video_entries()
                                if not video_entries:
                                    ui.label("Aucune vidéo détectée.").classes("text-gray-400 text-sm")
                                    return

                                query = (video_filters.get("query") or "").strip().lower()
                                filtered_entries = [
                                    entry for entry in video_entries
                                    if not query or query in f"{entry.name} {entry.path}".lower()
                                ]
                                visible_entries = filtered_entries[:60]

                                def open_video_preview(current_index: int):
                                    current_state = {"index": current_index}

                                    with ui.dialog() as dlg, ui.card().classes("w-[1100px] max-w-[96vw] bg-gray-900 border border-teal-700 p-4"):
                                        @ui.refreshable
                                        def render_preview():
                                            entry = visible_entries[current_state["index"]]
                                            thumb = get_video_thumbnail_data_url(entry.path, entry.mtime, entry.size)
                                            preview_url = build_video_preview_url(entry.path)

                                            with ui.column().classes("w-full gap-3"):
                                                with ui.row().classes("w-full items-start justify-between gap-4"):
                                                    with ui.column().classes("flex-grow gap-1"):
                                                        ui.label(entry.name).classes("text-lg font-bold text-white break-all")
                                                        ui.label(entry.path).classes("text-xs font-mono text-gray-400 break-all")
                                                        ui.label(
                                                            f"Vidéo {current_state['index'] + 1:,} sur {len(visible_entries):,}"
                                                        ).classes("text-xs text-teal-300")
                                                    ui.button(icon="close", on_click=dlg.close).props("flat round color=gray")

                                                ui.html(
                                                    f'<video src="{preview_url}" controls preload="metadata" autoplay style="width:100%; max-height:72vh; background:#000; border-radius:12px;"></video>'
                                                )

                                                with ui.row().classes("w-full items-center justify-between gap-3"):
                                                    prev_btn = ui.button(
                                                        "Précédente",
                                                        icon="arrow_back",
                                                        on_click=lambda: (
                                                            current_state.update({"index": current_state["index"] - 1}),
                                                            render_preview.refresh(),
                                                        ),
                                                    ).props("outline color=teal")
                                                    if current_state["index"] == 0:
                                                        prev_btn.disable()

                                                    next_btn = ui.button(
                                                        "Suivante",
                                                        icon="arrow_forward",
                                                        on_click=lambda: (
                                                            current_state.update({"index": current_state["index"] + 1}),
                                                            render_preview.refresh(),
                                                        ),
                                                    ).props("outline color=teal")
                                                    if current_state["index"] >= len(visible_entries) - 1:
                                                        next_btn.disable()

                                                if thumb:
                                                    ui.label("Repère visuel rapide").classes("text-xs text-gray-400")
                                                    ui.image(thumb).classes("w-56 h-32 object-cover rounded bg-black")

                                        render_preview()

                                    dlg.open()
                                selected_count = sum(
                                    1 for entry in video_entries
                                    if is_scan_entry_selected(entry, state.selected_categories, state.selected_files)
                                )

                                ui.label(
                                    f"{selected_count:,} vidéo(s) sélectionnée(s) sur {len(video_entries):,}"
                                ).classes("text-sm text-teal-200 font-bold mb-2")

                                if not filtered_entries:
                                    ui.label("Aucune vidéo ne correspond au filtre.").classes("text-gray-400 text-sm")
                                    return

                                if len(filtered_entries) > 60:
                                    ui.label("Affichage limité aux 60 premières vidéos pour garder l'interface fluide.").classes(
                                        "text-xs text-gray-500 mb-2"
                                    )

                                with ui.scroll_area().classes("w-full h-72 bg-black/25 rounded p-2"):
                                    with ui.grid(columns=4).classes("w-full gap-3"):
                                        for idx, entry in enumerate(visible_entries):
                                            checked = is_scan_entry_selected(entry, state.selected_categories, state.selected_files)
                                            thumb = get_video_thumbnail_data_url(entry.path, entry.mtime, entry.size)
                                            override = state.video_output_overrides.get(entry.path, {})
                                            custom_folder = (override.get("folder") or "").strip()
                                            custom_name = (override.get("filename") or "").strip()

                                            with ui.card().classes("w-full bg-gray-900/90 border border-gray-700 p-2 gap-2"):
                                                if thumb:
                                                    preview_img = ui.image(thumb).classes("w-full h-32 object-cover rounded bg-black cursor-pointer")
                                                    preview_img.on("click", lambda _, current_index=idx: open_video_preview(current_index))
                                                else:
                                                    with ui.column().classes("w-full h-32 items-center justify-center bg-black rounded gap-2"):
                                                        ui.icon("movie", size="2rem").classes("text-teal-300")
                                                        ui.label("Aperçu indisponible").classes("text-xs text-gray-400")
                                                        ui.button(
                                                            "Ouvrir aperçu",
                                                            icon="open_in_new",
                                                            on_click=lambda current_index=idx: open_video_preview(current_index),
                                                        ).props("flat dense color=teal")

                                                chk = ui.checkbox(value=checked)
                                                chk.props("dense")
                                                ui.label(entry.name).classes("text-xs font-bold text-white break-all")
                                                ui.label(entry.path).classes("text-[10px] font-mono text-gray-500 break-all")
                                                with ui.row().classes("gap-2 text-[10px]"):
                                                    ui.label(entry.ext.lower()).classes("text-teal-300")
                                                    ui.label(fmt_size(entry.size)).classes("text-gray-300")

                                                if custom_folder or custom_name or state.video_skip_date_hierarchy:
                                                    date_or_bucket = state.video_no_date_folder_name if state.video_skip_date_hierarchy else "YYYY/MM/DD"
                                                    ui.label(
                                                        f"Sortie: {((custom_folder + '/') if custom_folder else '')}{date_or_bucket}/{(custom_name or entry.name)}"
                                                    ).classes("text-[10px] text-lime-300 break-all")

                                                ui.button(
                                                    "Agrandir",
                                                    icon="play_circle",
                                                    on_click=lambda current_index=idx: open_video_preview(current_index),
                                                ).props("flat dense color=teal")

                                                def open_single_video_edit_dialog(current_entry=entry):
                                                    current_override = state.video_output_overrides.get(current_entry.path, {})
                                                    with ui.dialog() as dlg, ui.card().classes("w-[700px] max-w-[96vw] bg-gray-900 border border-teal-700"):
                                                        ui.label("Éditer destination vidéo").classes("text-lg font-bold text-teal-200")
                                                        ui.label(current_entry.name).classes("text-sm text-gray-300")
                                                        ui.label(current_entry.path).classes("text-xs font-mono text-gray-500 break-all")

                                                        folder_input = ui.input(
                                                            "Dossier final relatif (avant YYYY/MM/DD)",
                                                            value=current_override.get("folder", ""),
                                                            placeholder="ex: cosmikpratique",
                                                        ).classes("w-full")

                                                        name_input = ui.input(
                                                            "Nom final",
                                                            value=current_override.get("filename", ""),
                                                            placeholder="ex: PratiqueCosmikP1 (extension optionnelle)",
                                                        ).classes("w-full")

                                                        with ui.row().classes("w-full justify-between gap-2 mt-2"):
                                                            def clear_override():
                                                                state.video_output_overrides.pop(current_entry.path, None)
                                                                if video_selection_refresh_ref.get("fn"):
                                                                    video_selection_refresh_ref["fn"]()
                                                                ui.notify("Personnalisation supprimée.", type="info")
                                                                dlg.close()

                                                            ui.button("Réinitialiser", on_click=clear_override).props("flat color=gray")

                                                            with ui.row().classes("gap-2"):
                                                                ui.button("Annuler", on_click=dlg.close).props("flat color=gray")

                                                                def save_override():
                                                                    folder_value = (folder_input.value or "").strip()
                                                                    name_value = (name_input.value or "").strip()
                                                                    if not folder_value and not name_value:
                                                                        state.video_output_overrides.pop(current_entry.path, None)
                                                                    else:
                                                                        state.video_output_overrides[current_entry.path] = {
                                                                            "folder": folder_value,
                                                                            "filename": name_value,
                                                                        }
                                                                    if video_selection_refresh_ref.get("fn"):
                                                                        video_selection_refresh_ref["fn"]()
                                                                    ui.notify("Sortie vidéo mise à jour.", type="positive")
                                                                    dlg.close()

                                                                ui.button("Enregistrer", on_click=save_override).props("color=teal")

                                                    dlg.open()

                                                ui.button(
                                                    "Éditer sortie",
                                                    icon="edit",
                                                    on_click=open_single_video_edit_dialog,
                                                ).props("flat dense color=teal")

                                                def _on_video_toggle(e, current_entry=entry):
                                                    state.selected_files[current_entry.path] = e.value
                                                    sync_video_subcategory_state()
                                                    refresh_selection_dependent_panels()

                                                chk.on_value_change(_on_video_toggle)

                            video_search_input.on_value_change(lambda _: video_selection_ui.refresh())
                            video_selection_ui()
                            video_selection_refresh_ref["fn"] = video_selection_ui.refresh
                            step_refresh_refs["video_selection_ui"] = video_selection_ui.refresh

                    # ===== TAB IMAGES (IA + Vraies Photos) =====
                    with ui.tab_panel(tab_images):
                        with ui.column().classes("w-full gap-3 p-3"):

                            def _get_effective_subcat(entry_path: str, scan_subcat: str) -> str:
                                return state.image_reclassify_overrides.get(entry_path, scan_subcat)

                            def _get_image_entries_by_effective_subcat(target_subcat: str) -> list:
                                all_entries = []
                                for scan_subcat in ("Images", "IA-Images"):
                                    for e in state.scan_results.get("PHOTO-VIDEO", {}).get(scan_subcat, []):
                                        if _get_effective_subcat(e.path, scan_subcat) == target_subcat:
                                            all_entries.append(e)
                                return all_entries

                            def _apply_nsfw_filter(entries: list) -> list:
                                f = state.nsfw_filter_ui
                                if f == "sain":
                                    return [e for e in entries if e.nsfw_status not in ("SENSUEL", "EXPLICIT")]
                                if f == "nsfw":
                                    return [e for e in entries if e.nsfw_status in ("SENSUEL", "EXPLICIT")]
                                return entries

                            # ---- Barre NSFW ----
                            with ui.card().classes("w-full bg-gray-900/60 border border-rose-900/50 p-3 gap-2"):
                                with ui.row().classes("w-full items-center gap-3 flex-wrap"):
                                    ui.icon("shield", size="1.2rem").classes("text-rose-400")
                                    ui.label("Filtrage NSFW").classes("text-sm font-bold text-rose-300")

                                    def _set_nsfw_filter(val: str):
                                        state.nsfw_filter_ui = val
                                        if image_panel_refresh_refs.get("ia"):
                                            image_panel_refresh_refs["ia"]()
                                        if image_panel_refresh_refs.get("real"):
                                            image_panel_refresh_refs["real"]()

                                    ui.button("Tout", on_click=lambda: _set_nsfw_filter("all")).props("flat dense size=sm color=gray")
                                    ui.button("✅ Sain seulement", on_click=lambda: _set_nsfw_filter("sain")).props("flat dense size=sm color=green")
                                    ui.button("🔞 NSFW seulement", on_click=lambda: _set_nsfw_filter("nsfw")).props("flat dense size=sm color=red")

                                with ui.row().classes("w-full items-center gap-3 flex-wrap"):
                                    nsfw_excl_chk = ui.checkbox(
                                        "Exclure NSFW de l'organisation principale",
                                        value=state.nsfw_exclude,
                                    ).classes("text-xs text-rose-200")
                                    nsfw_excl_chk.on_value_change(lambda e: setattr(state, "nsfw_exclude", bool(e.value)))

                                    nsfw_dest_input = ui.input(
                                        "Destination alternative NSFW (optionnel)",
                                        placeholder="ex: D:\\Media-NSFW",
                                        value=state.nsfw_destination,
                                    ).classes("flex-grow text-xs").tooltip(
                                        "Si renseigné, les images NSFW (Sensuel+Explicit) seront copiées ici au lieu de la destination principale."
                                    )
                                    nsfw_dest_input.on_value_change(lambda e: setattr(state, "nsfw_destination", e.value.strip()))

                                    async def _pick_nsfw_dest():
                                        folder = await select_folder_async()
                                        if folder:
                                            nsfw_dest_input.set_value(folder)
                                            state.nsfw_destination = folder
                                    ui.button(icon="folder_open", on_click=_pick_nsfw_dest).props("flat dense color=rose").tooltip("Choisir dossier")

                            # Sous-onglets IA / Vraies Photos
                            with ui.tabs().classes("w-full bg-gray-900/40 rounded-t") as img_sub_tabs:
                                tab_ia  = ui.tab("IA Générées", icon="auto_awesome").props("no-caps")
                                tab_real = ui.tab("Vraies Photos", icon="photo_camera").props("no-caps")

                            img_search_state = {"ia": "", "real": ""}

                            with ui.tab_panels(img_sub_tabs, value=tab_ia).classes("w-full bg-gray-800/50 rounded-b p-0"):

                                # ----- Sous-panel IA -----
                                with ui.tab_panel(tab_ia):
                                    with ui.column().classes("w-full gap-2 p-2"):

                                        @ui.refreshable
                                        def ia_images_ui():
                                            ia_entries = _get_image_entries_by_effective_subcat("IA-Images")
                                            ia_entries = _apply_nsfw_filter(ia_entries)
                                            if not ia_entries:
                                                ui.label("Aucune image IA détectée.").classes("text-gray-400 text-sm italic p-2")
                                                return

                                            ia_key = ("PHOTO-VIDEO", "IA-Images")
                                            if ia_key not in state.selected_categories:
                                                state.selected_categories[ia_key] = True

                                            query = img_search_state.get("ia", "").strip().lower()
                                            filtered = [e for e in ia_entries if not query or query in f"{e.name} {e.path}".lower()]
                                            visible = filtered[:80]

                                            sel_count = sum(
                                                1 for e in ia_entries
                                                if state.selected_files.get(e.path, state.selected_categories.get(ia_key, True))
                                            )

                                            with ui.row().classes("w-full items-center gap-3 flex-wrap mb-2"):
                                                ui.label(f"{sel_count:,} / {len(ia_entries):,} sélectionnées").classes("text-sm text-violet-300 font-bold flex-grow")
                                                ui.button(
                                                    "Tout sélectionner",
                                                    icon="done_all",
                                                    on_click=lambda: (
                                                        [state.selected_files.update({e.path: True}) for e in ia_entries],
                                                        state.selected_categories.update({("PHOTO-VIDEO", "IA-Images"): True}),
                                                        ia_images_ui.refresh(),
                                                    ),
                                                ).props("flat dense color=violet")
                                                ui.button(
                                                    "Tout désélectionner",
                                                    icon="remove_done",
                                                    on_click=lambda: (
                                                        [state.selected_files.update({e.path: False}) for e in ia_entries],
                                                        state.selected_categories.update({("PHOTO-VIDEO", "IA-Images"): False}),
                                                        ia_images_ui.refresh(),
                                                    ),
                                                ).props("flat dense color=gray")
                                                ui.button(
                                                    "Tout → Vraies Photos",
                                                    icon="swap_horiz",
                                                    on_click=lambda: (
                                                        [state.image_reclassify_overrides.update({e.path: "Images"}) for e in ia_entries],
                                                        ia_images_ui.refresh(),
                                                        image_panel_refresh_refs.get("real") and image_panel_refresh_refs["real"](),
                                                    ),
                                                ).props("flat dense color=orange").tooltip("Reclasser toutes ces images comme Vraies Photos")

                                            if len(filtered) > 80:
                                                ui.label(f"Affichage limité à 80 / {len(filtered):,} résultats.").classes("text-xs text-gray-500 mb-1")

                                            with ui.scroll_area().classes("w-full h-96 bg-black/25 rounded p-2"):
                                                with ui.grid(columns=5).classes("w-full gap-2"):
                                                    for entry in visible:
                                                        checked = state.selected_files.get(
                                                            entry.path,
                                                            state.selected_categories.get(("PHOTO-VIDEO", "IA-Images"), True),
                                                        )
                                                        thumb = get_image_thumbnail_data_url(entry.path)

                                                        with ui.card().classes("w-full bg-gray-900/90 border border-violet-900/50 p-1 gap-1"):
                                                            if thumb:
                                                                ui.image(thumb).classes("w-full h-28 object-cover rounded bg-black")
                                                            else:
                                                                with ui.column().classes("w-full h-28 items-center justify-center bg-black rounded"):
                                                                    ui.icon("auto_awesome", size="1.8rem").classes("text-violet-400")
                                                                    ui.label("Aperçu N/D").classes("text-[10px] text-gray-400")

                                                            chk_ia = ui.checkbox(value=checked).props("dense")
                                                            ui.label(entry.name).classes("text-[10px] font-bold text-white break-all leading-tight")
                                                            with ui.row().classes("gap-1 text-[10px] items-center flex-wrap"):
                                                                ui.label(entry.ext.lower()).classes("text-violet-300")
                                                                ui.label(fmt_size(entry.size)).classes("text-gray-400")
                                                                # Badge NSFW
                                                                if entry.nsfw_status == "SAIN":
                                                                    ui.label("✅ SAIN").classes("text-[9px] text-green-400 font-bold")
                                                                elif entry.nsfw_status == "SENSUEL":
                                                                    ui.label("🟡 SENSUEL").classes("text-[9px] text-yellow-400 font-bold")
                                                                elif entry.nsfw_status == "EXPLICIT":
                                                                    ui.label("🔞 EXPLICIT").classes("text-[9px] text-red-400 font-bold")
                                                                if entry.companions:
                                                                    ui.label(f"📎 {len(entry.companions)}").classes("text-[9px] text-blue-300").tooltip(
                                                                        "Fichiers compagnons : " + ", ".join(os.path.basename(c) for c in entry.companions)
                                                                    )

                                                            # Badge reclassifié ?
                                                            if state.image_reclassify_overrides.get(entry.path) == "Images":
                                                                ui.label("→ Vraie photo").classes("text-[10px] text-orange-400 font-bold")

                                                            with ui.row().classes("w-full gap-1"):
                                                                def _on_ia_chk(e, p=entry.path):
                                                                    state.selected_files[p] = e.value
                                                                chk_ia.on_value_change(_on_ia_chk)

                                                                ui.button(
                                                                    "→ Vraie",
                                                                    on_click=lambda p=entry.path: (
                                                                        state.image_reclassify_overrides.update({p: "Images"}),
                                                                        ia_images_ui.refresh(),
                                                                        image_panel_refresh_refs.get("real") and image_panel_refresh_refs["real"](),
                                                                    ),
                                                                ).props("flat dense size=xs color=orange").tooltip("Reclasser comme Vraie Photo")

                                        ia_images_ui()
                                        image_panel_refresh_refs["ia"] = ia_images_ui.refresh
                                        step_refresh_refs["ia_images_ui"] = ia_images_ui.refresh

                                        # Recherche IA
                                        ia_search = ui.input("Rechercher", placeholder="nom ou chemin").classes("w-full mt-1")
                                        ia_search.on_value_change(lambda e: (
                                            img_search_state.update({"ia": e.value}),
                                            ia_images_ui.refresh(),
                                        ))

                                # ----- Sous-panel Vraies Photos -----
                                with ui.tab_panel(tab_real):
                                    with ui.column().classes("w-full gap-2 p-2"):

                                        @ui.refreshable
                                        def real_images_ui():
                                            real_entries = _get_image_entries_by_effective_subcat("Images")
                                            real_entries = _apply_nsfw_filter(real_entries)
                                            if not real_entries:
                                                ui.label("Aucune vraie photo détectée.").classes("text-gray-400 text-sm italic p-2")
                                                return

                                            real_key = ("PHOTO-VIDEO", "Images")
                                            if real_key not in state.selected_categories:
                                                state.selected_categories[real_key] = True

                                            query = img_search_state.get("real", "").strip().lower()
                                            filtered = [e for e in real_entries if not query or query in f"{e.name} {e.path}".lower()]
                                            visible = filtered[:80]

                                            sel_count = sum(
                                                1 for e in real_entries
                                                if state.selected_files.get(e.path, state.selected_categories.get(real_key, True))
                                            )

                                            with ui.row().classes("w-full items-center gap-3 flex-wrap mb-2"):
                                                ui.label(f"{sel_count:,} / {len(real_entries):,} sélectionnées").classes("text-sm text-teal-300 font-bold flex-grow")
                                                ui.button(
                                                    "Tout sélectionner",
                                                    icon="done_all",
                                                    on_click=lambda: (
                                                        [state.selected_files.update({e.path: True}) for e in real_entries],
                                                        state.selected_categories.update({("PHOTO-VIDEO", "Images"): True}),
                                                        real_images_ui.refresh(),
                                                    ),
                                                ).props("flat dense color=teal")
                                                ui.button(
                                                    "Tout désélectionner",
                                                    icon="remove_done",
                                                    on_click=lambda: (
                                                        [state.selected_files.update({e.path: False}) for e in real_entries],
                                                        state.selected_categories.update({("PHOTO-VIDEO", "Images"): False}),
                                                        real_images_ui.refresh(),
                                                    ),
                                                ).props("flat dense color=gray")
                                                ui.button(
                                                    "Tout → IA",
                                                    icon="swap_horiz",
                                                    on_click=lambda: (
                                                        [state.image_reclassify_overrides.update({e.path: "IA-Images"}) for e in real_entries],
                                                        real_images_ui.refresh(),
                                                        image_panel_refresh_refs.get("ia") and image_panel_refresh_refs["ia"](),
                                                    ),
                                                ).props("flat dense color=orange").tooltip("Reclasser toutes ces images comme IA")

                                            if len(filtered) > 80:
                                                ui.label(f"Affichage limité à 80 / {len(filtered):,} résultats.").classes("text-xs text-gray-500 mb-1")

                                            with ui.scroll_area().classes("w-full h-96 bg-black/25 rounded p-2"):
                                                with ui.grid(columns=5).classes("w-full gap-2"):
                                                    for entry in visible:
                                                        checked = state.selected_files.get(
                                                            entry.path,
                                                            state.selected_categories.get(("PHOTO-VIDEO", "Images"), True),
                                                        )
                                                        thumb = get_image_thumbnail_data_url(entry.path)

                                                        with ui.card().classes("w-full bg-gray-900/90 border border-teal-900/50 p-1 gap-1"):
                                                            if thumb:
                                                                ui.image(thumb).classes("w-full h-28 object-cover rounded bg-black")
                                                            else:
                                                                with ui.column().classes("w-full h-28 items-center justify-center bg-black rounded"):
                                                                    ui.icon("photo_camera", size="1.8rem").classes("text-teal-400")
                                                                    ui.label("Aperçu N/D").classes("text-[10px] text-gray-400")

                                                            chk_real = ui.checkbox(value=checked).props("dense")
                                                            ui.label(entry.name).classes("text-[10px] font-bold text-white break-all leading-tight")
                                                            with ui.row().classes("gap-1 text-[10px] items-center flex-wrap"):
                                                                ui.label(entry.ext.lower()).classes("text-teal-300")
                                                                ui.label(fmt_size(entry.size)).classes("text-gray-400")
                                                                # Badge NSFW
                                                                if entry.nsfw_status == "SAIN":
                                                                    ui.label("✅ SAIN").classes("text-[9px] text-green-400 font-bold")
                                                                elif entry.nsfw_status == "SENSUEL":
                                                                    ui.label("🟡 SENSUEL").classes("text-[9px] text-yellow-400 font-bold")
                                                                elif entry.nsfw_status == "EXPLICIT":
                                                                    ui.label("🔞 EXPLICIT").classes("text-[9px] text-red-400 font-bold")
                                                                if entry.companions:
                                                                    ui.label(f"📎 {len(entry.companions)}").classes("text-[9px] text-blue-300").tooltip(
                                                                        "Fichiers compagnons : " + ", ".join(os.path.basename(c) for c in entry.companions)
                                                                    )

                                                            # Badge reclassifié ?
                                                            if state.image_reclassify_overrides.get(entry.path) == "IA-Images":
                                                                ui.label("→ IA").classes("text-[10px] text-orange-400 font-bold")

                                                            with ui.row().classes("w-full gap-1"):
                                                                def _on_real_chk(e, p=entry.path):
                                                                    state.selected_files[p] = e.value
                                                                chk_real.on_value_change(_on_real_chk)

                                                                ui.button(
                                                                    "→ IA",
                                                                    on_click=lambda p=entry.path: (
                                                                        state.image_reclassify_overrides.update({p: "IA-Images"}),
                                                                        real_images_ui.refresh(),
                                                                        image_panel_refresh_refs.get("ia") and image_panel_refresh_refs["ia"](),
                                                                    ),
                                                                ).props("flat dense size=xs color=orange").tooltip("Reclasser comme image IA")

                                        real_images_ui()
                                        image_panel_refresh_refs["real"] = real_images_ui.refresh
                                        step_refresh_refs["real_images_ui"] = real_images_ui.refresh

                                        # Recherche Vraies Photos
                                        real_search = ui.input("Rechercher", placeholder="nom ou chemin").classes("w-full mt-1")
                                        real_search.on_value_change(lambda e: (
                                            img_search_state.update({"real": e.value}),
                                            real_images_ui.refresh(),
                                        ))

                    # ===== TAB MUSIQUE & AUDIO =====
                    with ui.tab_panel(tab_music):
                        with ui.column().classes("w-full gap-3 p-3"):

                            # Checkboxes sélection sous-catégories musique
                            music_subcats_data = state.scan_results.get("MUSIQUE-AUDIO", {})
                            if music_subcats_data:
                                music_cat_info = CATEGORIES.get("MUSIQUE-AUDIO", {})
                                music_stats = state.scan_stats.get("MUSIQUE-AUDIO", {})
                                with ui.card().classes("w-full bg-gray-800 border border-purple-700 p-3"):
                                    with ui.row().classes("w-full items-center gap-3 mb-2"):
                                        ui.icon("music_note", size="1.2rem").classes("text-purple-400")
                                        ui.label("MUSIQUE-AUDIO").classes("text-base font-bold flex-grow")
                                        ui.label(
                                            f"{music_stats.get('count', 0):,} fichiers — {fmt_size(music_stats.get('size', 0))}"
                                        ).classes("text-xs text-gray-400")
                                        music_sub_keys = list(music_subcats_data.keys())

                                        def _music_all():
                                            for s in music_sub_keys:
                                                state.selected_categories[("MUSIQUE-AUDIO", s)] = True
                                            if music_metadata_refresh_ref.get("fn"):
                                                music_metadata_refresh_ref["fn"]()
                                        def _music_none():
                                            for s in music_sub_keys:
                                                state.selected_categories[("MUSIQUE-AUDIO", s)] = False
                                            if music_metadata_refresh_ref.get("fn"):
                                                music_metadata_refresh_ref["fn"]()

                                        ui.button("Tout", on_click=_music_all).props("flat dense size=sm color=purple")
                                        ui.button("Rien", on_click=_music_none).props("flat dense size=sm color=gray")

                                    with ui.grid(columns=3).classes("w-full gap-2"):
                                        for msubcat, mentries in music_subcats_data.items():
                                            mkey = ("MUSIQUE-AUDIO", msubcat)
                                            if mkey not in state.selected_categories:
                                                state.selected_categories[mkey] = True
                                            mcount = len(mentries)
                                            msize = sum(e.size for e in mentries)
                                            with ui.row().classes("items-center gap-2 bg-gray-900 px-3 py-2 rounded-lg"):
                                                mchk = ui.checkbox(value=state.selected_categories[mkey])
                                                def _on_music_chk(e, k=mkey):
                                                    state.selected_categories[k] = e.value
                                                    if music_metadata_refresh_ref.get("fn"):
                                                        music_metadata_refresh_ref["fn"]()
                                                mchk.on_value_change(_on_music_chk)
                                                ui.label(msubcat).classes("flex-grow text-sm font-mono")
                                                ui.label(f"{mcount:,}").classes("text-xs text-gray-400")
                                                ui.label(fmt_size(msize)).classes("text-xs text-gray-500")
                            else:
                                ui.label("Aucun fichier audio détecté lors du scan.").classes("text-gray-400 text-sm italic p-2")

                            # Éditeur métadonnées musicales
                            with ui.expansion("✏️ Édition métadonnées musicales", icon="music_note").classes("w-full bg-gray-800 border border-purple-700/50 mt-1"):
                                @ui.refreshable
                                def music_metadata_editor():
                                    music_entries = []
                                    selected = {k: v for k, v in state.selected_categories.items() if v}
                                    for (cat, subcat), _ in selected.items():
                                        if cat == "MUSIQUE-AUDIO":
                                            entries = state.scan_results.get(cat, {}).get(subcat, [])
                                            music_entries.extend(entries)

                                    if not music_entries:
                                        ui.label("Aucun fichier MUSIQUE-AUDIO sélectionné.").classes("text-gray-400 text-sm")
                                        return

                                    ui.label(f"{len(music_entries):,} fichiers musicaux détectés").classes("text-sm text-purple-300 font-bold mb-2")

                                    with ui.scroll_area().classes("w-full h-64 bg-black/30 rounded p-2"):
                                        with ui.column().classes("gap-2"):
                                            for entry in music_entries[:50]:
                                                tags = MetadataExtractor.get_audio_tags(entry.path)
                                                current_artist = state.music_metadata_overrides.get(entry.path, {}).get("artist") or tags.get("artist") or ""
                                                current_album = state.music_metadata_overrides.get(entry.path, {}).get("album") or tags.get("album") or ""

                                                with ui.card().classes("w-full bg-gray-900/80 border border-gray-700 p-2"):
                                                    with ui.row().classes("w-full items-start gap-2 flex-wrap"):
                                                        with ui.column().classes("flex-grow min-w-[300px] gap-1"):
                                                            ui.label(entry.name).classes("text-xs font-bold text-white break-all")
                                                            ui.label(entry.path).classes("text-xs font-mono text-gray-500 break-all")

                                                        artist_input = ui.input(
                                                            "Artiste",
                                                            value=current_artist,
                                                        ).classes("min-w-[200px] text-xs")

                                                        album_input = ui.input(
                                                            "Album",
                                                            value=current_album,
                                                        ).classes("min-w-[200px] text-xs")

                                                        def save_music_metadata(path=entry.path, artist_field=artist_input, album_field=album_input):
                                                            state.music_metadata_overrides[path] = {
                                                                "artist": artist_field.value or "",
                                                                "album": album_field.value or "",
                                                            }
                                                            ui.notify(f"✓ {artist_field.value or '(artiste)'} / {album_field.value or '(album)'}", type="positive", timeout=1500)

                                                        ui.button(
                                                            "✓ Enregistrer",
                                                            icon="check",
                                                            on_click=save_music_metadata,
                                                        ).props("dense flat size=sm color=purple")

                                music_metadata_editor()
                                music_metadata_refresh_ref['fn'] = music_metadata_editor.refresh

                with ui.stepper_navigation():
                    ui.button("← Retour", on_click=stepper.previous).props("flat color=gray")
                    ui.button(
                        "Suivant", icon="arrow_forward", on_click=go_from_step3
                    ).classes("bg-orange-600 hover:bg-orange-500 font-bold px-6")

        # ================================================================
        # ÉTAPE 4 — FICHIERS NON CATÉGORISÉS
        # ================================================================
        with ui.step("Étape 4", title="Non catégorisés", icon="help"):
            with ui.column().classes("w-full max-w-5xl mx-auto p-6 gap-4"):

                ui.label("Décision sur les fichiers non catégorisés").classes(
                    "text-2xl font-bold text-yellow-400"
                )

                async def rescan_from_step4():
                    if not state.sources:
                        ui.notify("Ajoutez au moins une source dans l'Étape 1.", type="warning")
                        return
                    stepper.set_value("Étape 2")
                    await run_scan()
                    stepper.set_value("Étape 4")
                    selection_ui.refresh()
                    uncategorized_table.refresh()
                    refresh_uncategorized_label()

                with ui.row().classes("w-full justify-end"):
                    ui.button(
                        "Ajouter catégorie/projet",
                        icon="playlist_add",
                        on_click=lambda: open_custom_rules_dialog(
                            on_saved=lambda: (selection_ui.refresh(), uncategorized_table.refresh()),
                            on_rescan=rescan_from_step4,
                        ),
                    ).props("outline color=cyan")

                uncategorized_filters = {
                    "query": "",
                    "ext": "",
                }
                _unc_ext_expand: dict = {}  # extension → bool (déplié)
                _unc_ext_page:   dict = {}  # extension → page courante

                with ui.row().classes("w-full items-end gap-3 flex-wrap"):
                    unc_search_input = ui.input(
                        "Rechercher un fichier",
                        placeholder="Nom ou chemin",
                    ).classes("min-w-[360px] flex-grow")
                    unc_search_input.bind_value(uncategorized_filters, "query")

                    unc_ext_input = ui.input(
                        "Filtre extension",
                        placeholder="ex: .plist, .band, (aucune), TOUS",
                    ).classes("min-w-[260px]")
                    unc_ext_input.bind_value(uncategorized_filters, "ext")

                    ui.button(
                        "Effacer filtre",
                        icon="clear",
                        on_click=lambda: (
                            uncategorized_filters.update({"query": "", "ext": ""}),
                            unc_search_input.set_value(""),
                            unc_ext_input.set_value(""),
                            uncategorized_table.refresh(),
                        ),
                    ).props("outline color=gray")

                uncategorized_count_label = ui.label().classes("text-gray-300")

                def refresh_uncategorized_label():
                    n = len(state.uncategorized)
                    uncategorized_count_label.set_text(
                        f"{n:,} fichiers n'appartiennent à aucune catégorie connue."
                        if n > 0 else "Aucun fichier non catégorisé. 🎉"
                    )

                step_refresh_refs["uncategorized_label"] = refresh_uncategorized_label

                refresh_uncategorized_label()

                @ui.refreshable
                def uncategorized_table():
                    if not state.uncategorized:
                        ui.label("Aucun fichier non catégorisé ! 🎉").classes(
                            "text-green-400 m-4"
                        )
                        return

                    cat_options = [c for c in CATEGORIES.keys() if c != "PROJETS"]
                    default_cat = "DOCUMENTS" if "DOCUMENTS" in cat_options else (cat_options[0] if cat_options else "")

                    query = (uncategorized_filters.get("query") or "").strip().lower()
                    ext_filter = (uncategorized_filters.get("ext") or "").strip().lower()

                    # Grouper par extension
                    by_ext: dict = defaultdict(list)
                    for entry in state.uncategorized:
                        ext_key = (entry.ext or "(aucune)").lower()
                        if query and query not in f"{entry.name} {entry.path} {ext_key}".lower():
                            continue
                        if ext_filter and ext_filter not in {"tous", "*"} and ext_filter != ext_key:
                            continue
                        by_ext[ext_key].append(entry)

                    filtered_count = sum(len(v) for v in by_ext.values())
                    ui.label(
                        f"{filtered_count:,} affichés sur {len(state.uncategorized):,} fichiers non catégorisés"
                    ).classes("text-sm text-yellow-200 mb-1")

                    if filtered_count == 0:
                        ui.label("Aucun fichier ne correspond au filtre.").classes("text-yellow-300 text-sm")
                        return

                    with ui.column().classes("w-full max-h-[52vh] overflow-y-auto gap-2"):
                        for ext, entries in sorted(by_ext.items()):
                            total_size = sum(e.size for e in entries)
                            with ui.card().classes("w-full bg-gray-800 border border-gray-700 p-3"):
                                with ui.row().classes("w-full items-center gap-3 mb-2"):
                                    ui.label(ext if ext else "(sans extension)").classes(
                                        "font-mono text-yellow-300 font-bold w-24 shrink-0"
                                    )
                                    ui.label(
                                        f"{len(entries):,} fichiers — {fmt_size(total_size)}"
                                    ).classes("flex-grow text-sm text-gray-300")

                                    ui.button(
                                        "Ignorer tout",
                                        on_click=lambda _, es=entries, x=ext: (
                                            [state.uncategorized_decisions.update({e.path: "ignore"}) for e in es],
                                            ui.notify(f"{x} : ignoré", type="info", timeout=1000)
                                        )
                                    ).props("flat dense size=sm color=gray")

                                    ui.button(
                                        "Effacer tout",
                                        on_click=lambda _, es=entries, x=ext: (
                                            [state.uncategorized_decisions.update({e.path: "delete"}) for e in es],
                                            ui.notify(f"{x} : suppression sélectionnée", type="warning", timeout=1200)
                                        )
                                    ).props("flat dense size=sm color=red")

                                    assign_sel = ui.select(
                                        options=["→ Assigner à..."] + cat_options,
                                        value="→ Assigner à...",
                                    ).classes("min-w-[180px] text-xs")

                                    def on_assign(ev, es=entries, x=ext, sel=assign_sel):
                                        if ev.value.startswith("→"):
                                            return
                                        for ent in es:
                                            state.uncategorized_decisions[ent.path] = f"assign:{ev.value}"
                                        ui.notify(f"{x} → {ev.value}", type="positive", timeout=1200)
                                        sel.set_value("→ Assigner à...")

                                    assign_sel.on_value_change(on_assign)

                                    # Toggle expand/collapse des fichiers individuels
                                    is_ext_exp = _unc_ext_expand.get(ext, False)
                                    ui.label(
                                        f"{'\u25b2' if is_ext_exp else '\u25bc'} {len(entries):,} fichiers"
                                    ).classes(
                                        "text-xs text-blue-400 cursor-pointer underline shrink-0 ml-2"
                                    ).on("click", lambda x=ext: (
                                        _unc_ext_expand.update({x: not _unc_ext_expand.get(x, False)}),
                                        uncategorized_table.refresh(),
                                    ))

                                # Liste dépliable des fichiers (rendue seulement si expanded)
                                if _unc_ext_expand.get(ext, False):
                                    _UNC_PAGE_SIZE = 50
                                    sorted_unc = sorted(entries, key=lambda e: e.name.lower())
                                    unc_total_pages = max(1, (len(sorted_unc) + _UNC_PAGE_SIZE - 1) // _UNC_PAGE_SIZE)
                                    unc_page = max(0, min(_unc_ext_page.get(ext, 0), unc_total_pages - 1))
                                    shown_unc = sorted_unc[unc_page * _UNC_PAGE_SIZE:(unc_page + 1) * _UNC_PAGE_SIZE]

                                    with ui.column().classes("w-full gap-2 mt-2"):
                                        if unc_total_pages > 1:
                                            with ui.row().classes("w-full items-center gap-2 py-1 border-t border-gray-700"):
                                                ui.button(
                                                    icon="chevron_left",
                                                    on_click=lambda x=ext, p=unc_page: (
                                                        _unc_ext_page.update({x: max(0, p - 1)}),
                                                        uncategorized_table.refresh(),
                                                    ),
                                                ).props("flat dense round").set_enabled(unc_page > 0)
                                                ui.label(
                                                    f"Page {unc_page + 1} / {unc_total_pages}  "
                                                    f"({unc_page * _UNC_PAGE_SIZE + 1}–"
                                                    f"{min((unc_page + 1) * _UNC_PAGE_SIZE, len(sorted_unc)):,} "
                                                    f"sur {len(sorted_unc):,})"
                                                ).classes("text-xs text-gray-400 flex-grow text-center")
                                                ui.button(
                                                    icon="chevron_right",
                                                    on_click=lambda x=ext, p=unc_page, tp=unc_total_pages: (
                                                        _unc_ext_page.update({x: min(tp - 1, p + 1)}),
                                                        uncategorized_table.refresh(),
                                                    ),
                                                ).props("flat dense round").set_enabled(unc_page < unc_total_pages - 1)

                                        for entry in shown_unc:
                                            current_decision = state.uncategorized_decisions.get(entry.path, "ignore")
                                            current_action = "assign" if current_decision.startswith("assign:") else current_decision
                                            current_cat = current_decision[7:] if current_decision.startswith("assign:") else default_cat

                                            with ui.card().classes("w-full bg-gray-900/80 border border-gray-700 p-2"):
                                                with ui.row().classes("w-full items-start gap-3 flex-wrap"):
                                                    with ui.column().classes("flex-grow min-w-[420px] gap-1"):
                                                        ui.label(entry.name).classes("text-sm font-bold text-white")
                                                        ui.label(entry.path).classes("text-xs font-mono text-gray-400 break-all")
                                                        with ui.row().classes("gap-3"):
                                                            ui.label((entry.ext or "(aucune)").lower()).classes("text-xs font-mono text-yellow-300")
                                                            ui.label(fmt_size(entry.size)).classes("text-xs text-gray-300")
                                                            ui.label(
                                                                datetime.datetime.fromtimestamp(entry.mtime).strftime("%Y-%m-%d %H:%M")
                                                            ).classes("text-xs text-gray-500")

                                                    action_sel = ui.select(
                                                        options={
                                                            "ignore": "Ignorer",
                                                            "assign": "Catégoriser",
                                                            "delete": "Effacer",
                                                        },
                                                        value=current_action,
                                                    ).classes("min-w-[170px] text-xs")

                                                    cat_sel = ui.select(
                                                        options=cat_options,
                                                        value=current_cat,
                                                    ).classes("min-w-[180px] text-xs")

                                                    def apply_uncategorized_decision(
                                                        _,
                                                        file_path=entry.path,
                                                        action=action_sel,
                                                        cat=cat_sel,
                                                    ):
                                                        if action.value == "assign":
                                                            chosen_cat = cat.value or default_cat
                                                            state.uncategorized_decisions[file_path] = f"assign:{chosen_cat}"
                                                        elif action.value == "delete":
                                                            state.uncategorized_decisions[file_path] = "delete"
                                                        else:
                                                            state.uncategorized_decisions[file_path] = "ignore"

                                                    action_sel.on_value_change(apply_uncategorized_decision)
                                                    cat_sel.on_value_change(apply_uncategorized_decision)

                unc_search_input.on_value_change(lambda _: uncategorized_table.refresh())
                unc_ext_input.on_value_change(lambda _: uncategorized_table.refresh())

                uncategorized_table()
                step_refresh_refs["uncategorized_table"] = uncategorized_table.refresh

                with ui.row().classes("gap-3"):
                    ui.button(
                        "Ignorer TOUT",
                        icon="block",
                        on_click=lambda: (
                            [state.uncategorized_decisions.update({e.path: "ignore"}) for e in state.uncategorized],
                            ui.notify("Tous ignorés", type="info")
                        )
                    ).props("outline color=gray")

                    ui.button(
                        "Effacer TOUT",
                        icon="delete_forever",
                        on_click=lambda: (
                            [state.uncategorized_decisions.update({e.path: "delete"}) for e in state.uncategorized],
                            ui.notify("Tous les non catégorisés seront supprimés.", type="warning")
                        )
                    ).props("outline color=red")

                with ui.stepper_navigation():
                    ui.button("← Retour", on_click=stepper.previous).props("flat color=gray")
                    ui.button(
                        "Suivant", icon="arrow_forward", on_click=go_from_step4
                    ).classes("bg-orange-600 hover:bg-orange-500 font-bold px-6")

        # ================================================================
        # ÉTAPE 5 — FICHIERS EXÉCUTABLES ET ARCHIVES (.MSI/.EXE/.ZIP/.RAR)
        # ================================================================
        with ui.step("Étape 5", title="Exécutables & Archives", icon="inventory_2"):
            with ui.column().classes("w-full max-w-6xl mx-auto p-6 gap-4"):

                ui.label("Gestion des fichiers exécutables et archives").classes(
                    "text-2xl font-bold text-amber-400"
                )
                ui.label(
                    "Décidez quoi faire avec les .msi / .exe / .zip / .rar : conserver, catégoriser, corbeille ou suppression. (archives .zip/.rar conservées -> DOCUMENTS)"
                ).classes("text-sm text-gray-400")

                async def rescan_from_step5():
                    if not state.sources:
                        ui.notify("Ajoutez au moins une source dans l'Étape 1.", type="warning")
                        return
                    stepper.set_value("Étape 2")
                    await run_scan()
                    stepper.set_value("Étape 5")
                    selection_ui.refresh()
                    uncategorized_table.refresh()
                    refresh_uncategorized_label()
                    installers_ui.refresh()

                with ui.row().classes("w-full justify-end"):
                    ui.button(
                        "Ajouter catégorie/projet",
                        icon="playlist_add",
                        on_click=lambda: open_custom_rules_dialog(
                            on_saved=lambda: (selection_ui.refresh(), installers_ui.refresh()),
                            on_rescan=rescan_from_step5,
                        ),
                    ).props("outline color=cyan")

                installer_filters = {
                    "query": "",
                    "ext": "TOUS",
                }

                with ui.row().classes("w-full items-end gap-3 flex-wrap"):
                    search_input = ui.input(
                        "Rechercher un fichier",
                        placeholder="Nom ou chemin (ex: setup, nvidia, archive)",
                    ).classes("min-w-[380px] flex-grow")
                    search_input.bind_value(installer_filters, "query")

                    ext_select = ui.select(
                        options=["TOUS", ".exe", ".msi", ".zip", ".rar"],
                        value="TOUS",
                        label="Extension",
                    ).classes("min-w-[160px]")
                    ext_select.bind_value(installer_filters, "ext")

                    ui.button(
                        "Effacer filtre",
                        icon="clear",
                        on_click=lambda: (
                            installer_filters.update({"query": "", "ext": "TOUS"}),
                            search_input.set_value(""),
                            ext_select.set_value("TOUS"),
                            installers_ui.refresh(),
                        ),
                    ).props("outline color=gray")

                @ui.refreshable
                def installers_ui():
                    if not state.installer_files:
                        ui.label("Aucun fichier .msi/.exe/.zip/.rar détecté.").classes("text-green-400 m-2")
                        return

                    cat_options = [c for c in CATEGORIES.keys() if c != "PROJETS"]
                    default_cat = "DOCUMENTS" if "DOCUMENTS" in cat_options else (cat_options[0] if cat_options else "")

                    query = (installer_filters.get("query") or "").strip().lower()
                    ext_filter = installer_filters.get("ext") or "TOUS"

                    filtered_installers = []
                    for entry in state.installer_files:
                        if ext_filter != "TOUS" and entry.ext.lower() != ext_filter.lower():
                            continue
                        if query and query not in f"{entry.name} {entry.path}".lower():
                            continue
                        filtered_installers.append(entry)

                    # Résumé rapide
                    total_size = sum(e.size for e in state.installer_files)
                    ui.label(
                        f"{len(filtered_installers):,} affichés sur {len(state.installer_files):,} fichiers détectés ({fmt_size(total_size)})"
                    ).classes("text-sm text-amber-200 mb-1")

                    if not filtered_installers:
                        ui.label("Aucun fichier ne correspond au filtre.").classes("text-yellow-300 text-sm")
                        return

                    # Liste détaillée par fichier (nom visible + décision individuelle)
                    with ui.column().classes("w-full max-h-[55vh] overflow-y-auto gap-2"):
                        for entry in sorted(filtered_installers, key=lambda e: (e.ext.lower(), e.name.lower())):
                            current_decision = state.installer_decisions.get(entry.path, "keep")
                            current_action = "assign" if current_decision.startswith("assign:") else current_decision
                            current_cat = current_decision[7:] if current_decision.startswith("assign:") else default_cat

                            with ui.card().classes("w-full bg-gray-800 border border-amber-700 p-3"):
                                with ui.row().classes("w-full items-start gap-3 flex-wrap"):
                                    with ui.column().classes("flex-grow min-w-[420px] gap-1"):
                                        ui.label(entry.name).classes("text-sm font-bold text-white")
                                        ui.label(entry.path).classes("text-xs font-mono text-gray-400 break-all")
                                        with ui.row().classes("gap-3"):
                                            ui.label(entry.ext.lower()).classes("text-xs font-mono text-amber-300")
                                            ui.label(fmt_size(entry.size)).classes("text-xs text-gray-300")
                                            ui.label(
                                                datetime.datetime.fromtimestamp(entry.mtime).strftime("%Y-%m-%d %H:%M")
                                            ).classes("text-xs text-gray-500")

                                    action_sel = ui.select(
                                        options={
                                            "keep": "Conserver (archives -> DOCUMENTS)",
                                            "assign": "Catégoriser",
                                            "trash": "Corbeille",
                                            "delete": "Supprimer",
                                        },
                                        value=current_action,
                                    ).classes("min-w-[170px] text-xs")

                                    cat_sel = ui.select(
                                        options=cat_options,
                                        value=current_cat,
                                    ).classes("min-w-[180px] text-xs")

                                    def apply_one_file_decision(
                                        _,
                                        file_path=entry.path,
                                        action=action_sel,
                                        cat=cat_sel,
                                    ):
                                        if action.value == "delete":
                                            state.installer_decisions[file_path] = "delete"
                                        elif action.value == "trash":
                                            state.installer_decisions[file_path] = "trash"
                                        elif action.value == "assign":
                                            chosen_cat = cat.value or default_cat
                                            state.installer_decisions[file_path] = f"assign:{chosen_cat}"
                                        else:
                                            state.installer_decisions[file_path] = "keep"

                                    action_sel.on_value_change(apply_one_file_decision)
                                    cat_sel.on_value_change(apply_one_file_decision)

                search_input.on_value_change(lambda _: installers_ui.refresh())
                ext_select.on_value_change(lambda _: installers_ui.refresh())

                installers_ui()
                step_refresh_refs["installers_ui"] = installers_ui.refresh

                with ui.row().classes("gap-2"):
                    ui.button(
                        "Tout en corbeille (.msi/.exe/.zip/.rar)",
                        icon="delete_sweep",
                        on_click=lambda: (
                            [state.installer_decisions.update({e.path: "trash"}) for e in state.installer_files],
                            installers_ui.refresh(),
                            ui.notify("Tous les exécutables/archives iront à la corbeille.", type="warning"),
                        ),
                    ).props("outline color=amber")

                    ui.button(
                        "Tout supprimer (.msi/.exe/.zip/.rar)",
                        icon="delete_forever",
                        on_click=lambda: (
                            [state.installer_decisions.update({e.path: "delete"}) for e in state.installer_files],
                            installers_ui.refresh(),
                            ui.notify("Tous les exécutables/archives seront supprimés.", type="warning"),
                        ),
                    ).props("outline color=red")

                    ui.button(
                        "Tout conserver",
                        icon="inventory",
                        on_click=lambda: (
                            [state.installer_decisions.update({e.path: "keep"}) for e in state.installer_files],
                            installers_ui.refresh(),
                            ui.notify("Tous les exécutables/archives seront conservés.", type="info"),
                        ),
                    ).props("outline color=gray")

                with ui.stepper_navigation():
                    ui.button("← Retour", on_click=stepper.previous).props("flat color=gray")
                    ui.button(
                        "Suivant", icon="arrow_forward", on_click=go_from_step5
                    ).classes("bg-orange-600 hover:bg-orange-500 font-bold px-6")

        # ================================================================
        # ÉTAPE 6 — APERÇU ET EXÉCUTION
        # ================================================================
        with ui.step("Étape 6", title="Exécution", icon="play_circle"):
            with ui.column().classes("w-full max-w-4xl mx-auto p-6 gap-4"):

                ui.label("Aperçu et lancement de l'organisation").classes(
                    "text-2xl font-bold text-red-400"
                )

                # Résumé des sélections
                with ui.card().classes("w-full bg-gray-800 border border-gray-700 p-4"):
                    ui.label("Résumé des opérations").classes("font-bold text-lg mb-2")

                    @ui.refreshable
                    def summary_ui():
                        selected = {k: v for k, v in state.selected_categories.items() if v}
                        n_assign_installers = sum(
                            1 for e in state.installer_files
                            if state.installer_decisions.get(e.path, "keep").startswith("assign:")
                        )
                        n_auto_archive_move = sum(
                            1 for e in state.installer_files
                            if e.ext.lower() in {".zip", ".rar"}
                            and state.installer_decisions.get(e.path, "keep") == "keep"
                        )
                        n_assign_uncategorized = sum(
                            1 for v in state.uncategorized_decisions.values()
                            if isinstance(v, str) and v.startswith("assign:")
                        )
                        n_delete_uncategorized = sum(
                            1 for v in state.uncategorized_decisions.values()
                            if v == "delete"
                        )

                        has_selected_scan_entries = any(
                            is_scan_entry_selected(entry, state.selected_categories, state.selected_files)
                            for subcats in state.scan_results.values()
                            for entries in subcats.values()
                            for entry in entries
                        )

                        if not has_selected_scan_entries and not selected and (n_assign_installers + n_auto_archive_move) == 0 and (n_assign_uncategorized + n_delete_uncategorized) == 0:
                            ui.label("Aucune catégorie sélectionnée.").classes("text-red-400")
                            ui.label("Astuce: vous pouvez aussi organiser uniquement des .exe/.msi/.zip/.rar depuis l'étape précédente.").classes(
                                "text-xs text-gray-400"
                            )
                            return

                        total_files = 0
                        total_size = 0
                        for cat, subcats in state.scan_results.items():
                            if cat == "PROJETS":
                                continue
                            for entries in subcats.values():
                                for entry in entries:
                                    if is_scan_entry_selected(entry, state.selected_categories, state.selected_files):
                                        total_files += 1
                                        total_size += entry.size

                        # Projets
                        proj_selected = [
                            subcat for (cat, subcat), v in selected.items()
                            if cat == "PROJETS" and v
                        ]
                        n_proj = sum(
                            1 for _, s in state.project_roots if s in proj_selected
                        )

                        n_delete_installers = sum(
                            1 for e in state.installer_files
                            if state.installer_decisions.get(e.path, "keep") == "delete"
                        )
                        n_trash_installers = sum(
                            1 for e in state.installer_files
                            if state.installer_decisions.get(e.path, "keep") == "trash"
                        )

                        dest_label = state.destination or "(sur place)"
                        mode_label = "Copie" if state.mode == "copy" else "Déplacement"

                        with ui.column().classes("gap-1"):
                            ui.label(
                                f"Mode : {mode_label}"
                            ).classes("text-sm font-mono")
                            ui.label(
                                f"Destination : {dest_label}"
                            ).classes("text-sm font-mono text-green-300")
                            ui.label(
                                f"Fichiers à organiser : {total_files:,} ({fmt_size(total_size)})"
                            ).classes("text-sm font-mono text-blue-300")
                            if n_assign_installers:
                                ui.label(
                                    f"Exécutables/archives à catégoriser : {n_assign_installers}"
                                ).classes("text-sm font-mono text-cyan-300")
                            if n_auto_archive_move:
                                ui.label(
                                    f"Archives .zip/.rar à déplacer vers DOCUMENTS : {n_auto_archive_move}"
                                ).classes("text-sm font-mono text-cyan-300")
                            if n_assign_uncategorized:
                                ui.label(
                                    f"Non catégorisés à assigner : {n_assign_uncategorized}"
                                ).classes("text-sm font-mono text-cyan-300")
                            if n_delete_uncategorized:
                                ui.label(
                                    f"Non catégorisés à supprimer : {n_delete_uncategorized}"
                                ).classes("text-sm font-mono text-red-300")
                            if n_proj:
                                ui.label(
                                    f"Projets à organiser : {n_proj}"
                                ).classes("text-sm font-mono text-orange-300")
                            if n_delete_installers:
                                ui.label(
                                    f"Exécutables/archives à supprimer : {n_delete_installers}"
                                ).classes("text-sm font-mono text-red-300")
                            if n_trash_installers:
                                ui.label(
                                    f"Exécutables/archives en corbeille : {n_trash_installers}"
                                ).classes("text-sm font-mono text-amber-300")
                            ui.label(
                                f"Catégories sélectionnées : {', '.join(f'{c}/{s}' for (c,s) in selected)}"
                            ).classes("text-xs text-gray-400 mt-1")

                    summary_ui()
                    step_refresh_refs["summary_ui"] = summary_ui.refresh

                # Assure un résumé toujours à jour même avec navigation directe du stepper
                ui.timer(0.8, lambda: summary_ui.refresh())

                # Exécution
                exec_log_label = ui.label("").classes("text-gray-400 text-sm font-mono")

                async def run_organize():
                    has_scan_data = bool(state.scan_results or state.project_roots or state.uncategorized or state.installer_files)
                    if not has_scan_data:
                        ui.notify("Lancez d'abord le scan !", type="warning")
                        return

                    if not state.destination and state.mode == "copy":
                        ui.notify(
                            "Définissez un dossier destination (ou choisissez le mode Déplacer sur place).",
                            type="warning"
                        )
                        return

                    installer_delete_paths = [
                        e.path for e in state.installer_files
                        if state.installer_decisions.get(e.path, "keep") == "delete"
                    ]
                    uncategorized_delete_paths = [
                        e.path for e in state.uncategorized
                        if state.uncategorized_decisions.get(e.path, "ignore") == "delete"
                    ]
                    to_delete_paths = installer_delete_paths + uncategorized_delete_paths
                    to_trash = [
                        e for e in state.installer_files
                        if state.installer_decisions.get(e.path, "keep") == "trash"
                    ]

                    organizer = FileOrganizer(
                        destination=state.destination,
                        mode=state.mode,
                        log_callback=lambda m: state.add_log(m),
                        progress_callback=lambda p, m: (
                            setattr(state, "progress", p),
                            setattr(state, "status_text", m)
                        )
                    )

                    # Prévisualisation des 10 premiers éléments à déplacer/copier
                    preview_move_items = []

                    # Fichiers classés par catégories cochées ou sélection individuelle
                    for cat, subcats in state.scan_results.items():
                        if cat == "PROJETS":
                            continue
                        for entries in subcats.values():
                            for entry in entries:
                                if is_scan_entry_selected(entry, state.selected_categories, state.selected_files):
                                    preview_move_items.append((
                                        entry.path,
                                        organizer._resolve_dest(
                                            entry,
                                            state.music_metadata_overrides,
                                            state.video_output_overrides,
                                            state.video_skip_date_hierarchy,
                                            state.video_no_date_folder_name,
                                        ),
                                    ))

                    # Non catégorisés assignés
                    for entry in state.uncategorized:
                        decision = state.uncategorized_decisions.get(entry.path, "ignore")
                        if decision.startswith("assign:"):
                            assigned_cat = decision[7:]
                            base_dest = organizer._base_destination(entry.path)
                            dest = os.path.join(base_dest, assigned_cat, entry.ext.lstrip(".").upper() or "DIVERS", entry.name)
                            preview_move_items.append((entry.path, dest))

                    # Exécutables/archives assignés
                    for entry in state.installer_files:
                        decision = state.installer_decisions.get(entry.path, "keep")
                        if decision.startswith("assign:"):
                            assigned_cat = decision[7:]
                            base_dest = organizer._base_destination(entry.path)
                            dest = os.path.join(base_dest, assigned_cat, entry.ext.lstrip(".").upper() or "DIVERS", entry.name)
                            preview_move_items.append((entry.path, dest))
                        elif decision == "keep" and entry.ext.lower() in {".zip", ".rar"}:
                            # Archives conservées: déplacement auto vers DOCUMENTS.
                            base_dest = organizer._base_destination(entry.path)
                            dest = os.path.join(base_dest, "DOCUMENTS", entry.ext.lstrip(".").upper() or "DIVERS", entry.name)
                            preview_move_items.append((entry.path, dest))

                    # Projets sélectionnés
                    proj_selected = {
                        subcat for (cat, subcat), checked in state.selected_categories.items()
                        if cat == "PROJETS" and checked
                    }
                    for proj_root, subcat in state.project_roots:
                        if subcat in proj_selected:
                            preview_move_items.append((proj_root, organizer._resolve_project_dest(proj_root, subcat)))

                    async def confirm_execution() -> bool:
                        with ui.dialog() as dlg, ui.card().classes("w-[900px] max-w-[95vw] bg-gray-900"):
                            ui.label("Confirmation d'organisation").classes("text-xl font-bold text-orange-300")
                            ui.label(
                                f"Mode: {state.mode} | Destination: {state.destination or 'sur place'}"
                            ).classes("text-sm text-gray-300")

                            if preview_move_items:
                                ui.separator()
                                ui.label(
                                    f"Aperçu des 10 premiers éléments à {'déplacer' if state.mode == 'move' else 'copier'} :"
                                ).classes("text-sm text-cyan-300 font-bold")
                                with ui.scroll_area().classes("w-full h-44 bg-black/40 rounded p-2"):
                                    with ui.column().classes("gap-0"):
                                        for src, dst in preview_move_items[:10]:
                                            ui.label(f"{src} -> {dst}").classes("text-xs font-mono text-cyan-100 break-all")

                            if to_delete_paths:
                                ui.separator()
                                ui.label(
                                    f"{len(to_delete_paths)} fichiers seront supprimés définitivement :"
                                ).classes("text-sm text-red-300 font-bold")
                                with ui.scroll_area().classes("w-full h-52 bg-black/40 rounded p-2"):
                                    with ui.column().classes("gap-0"):
                                        for path in to_delete_paths:
                                            ui.label(path).classes("text-xs font-mono text-red-200")

                            if to_trash:
                                ui.separator()
                                ui.label(
                                    f"{len(to_trash)} fichiers seront envoyés à la corbeille :"
                                ).classes("text-sm text-amber-300 font-bold")
                                with ui.scroll_area().classes("w-full h-40 bg-black/40 rounded p-2"):
                                    with ui.column().classes("gap-0"):
                                        for e in to_trash:
                                            ui.label(e.path).classes("text-xs font-mono text-amber-200")

                            if to_trash and not HAS_SEND2TRASH:
                                ui.label(
                                    "⚠️ Corbeille indisponible: package send2trash manquant."
                                ).classes("text-xs text-red-300")

                            with ui.row().classes("w-full justify-end gap-2 mt-3"):
                                ui.button("Annuler", on_click=lambda: dlg.submit(False)).props("flat color=gray")
                                ui.button("Confirmer", on_click=lambda: dlg.submit(True)).props("color=red")
                        dlg.open()
                        return await dlg

                    confirm = await confirm_execution()
                    if not confirm:
                        return

                    state.is_organizing = True
                    exec_log_label.set_text("Organisation en cours...")

                    def task():
                        report = organizer.organize(
                            selected=state.selected_categories,
                            selected_files=state.selected_files,
                            scan_results=state.scan_results,
                            project_roots=state.project_roots,
                            uncategorized_decisions=state.uncategorized_decisions,
                            uncategorized=state.uncategorized,
                            installer_decisions=state.installer_decisions,
                            installer_files=state.installer_files,
                            db_archive=db_archive if db_archive._available else None,
                            metadata_overrides=state.music_metadata_overrides,
                            video_overrides=state.video_output_overrides,
                            video_skip_date_hierarchy=state.video_skip_date_hierarchy,
                            video_no_date_folder_name=state.video_no_date_folder_name,
                        )
                        state.report = report
                        return report

                    report = await run.io_bound(task)

                    if state.mode == "move" and state.delete_empty_sources_after_move:
                        cleanup_result = await run.io_bound(lambda: cleanup_empty_source_dirs(state.sources))
                        report["source_cleanup"] = cleanup_result
                        report["deleted_empty_source_dirs"] = cleanup_result.get("deleted", [])
                        if cleanup_result.get("deleted"):
                            state.add_log(
                                f"🧹 Dossiers source supprimés (vides): {len(cleanup_result['deleted'])}"
                            )
                        else:
                            state.add_log("🧹 Aucun dossier source vide à supprimer.")

                    detailed_context = {
                        "selected_categories_count": len([k for k, v in state.selected_categories.items() if v]),
                        "selected_files_count": len([k for k, v in state.selected_files.items() if v]),
                        "uncategorized_assign_count": len([
                            v for v in state.uncategorized_decisions.values()
                            if isinstance(v, str) and v.startswith("assign:")
                        ]),
                        "uncategorized_delete_count": len([
                            v for v in state.uncategorized_decisions.values()
                            if v == "delete"
                        ]),
                        "installer_assign_count": len([
                            v for v in state.installer_decisions.values()
                            if isinstance(v, str) and v.startswith("assign:")
                        ]),
                        "installer_delete_count": len([
                            v for v in state.installer_decisions.values()
                            if v == "delete"
                        ]),
                        "installer_trash_count": len([
                            v for v in state.installer_decisions.values()
                            if v == "trash"
                        ]),
                        "video_skip_date_hierarchy": state.video_skip_date_hierarchy,
                        "video_no_date_folder_name": state.video_no_date_folder_name,
                    }

                    full_logs_snapshot = list(state.full_log_history)

                    json_report_path = export_report_json(
                        report=report,
                        destination=state.destination,
                        mode=state.mode,
                        sources=state.sources,
                        detailed_context=detailed_context,
                        full_logs=full_logs_snapshot,
                    )
                    json_compact_report_path = export_report_json_compact(
                        report=report,
                        destination=state.destination,
                        mode=state.mode,
                        sources=state.sources,
                    )
                    txt_report_path = export_report_txt_detailed(
                        report=report,
                        destination=state.destination,
                        mode=state.mode,
                        sources=state.sources,
                        full_logs=full_logs_snapshot,
                        detailed_context=detailed_context,
                    )
                    if json_report_path:
                        report["json_export_path"] = json_report_path
                        state.add_log(f"🧾 Rapport JSON exporté: {json_report_path}")
                    if json_compact_report_path:
                        report["json_compact_export_path"] = json_compact_report_path
                        state.add_log(f"🧾 Rapport JSON compact exporté: {json_compact_report_path}")
                    if txt_report_path:
                        report["txt_detailed_export_path"] = txt_report_path
                        state.add_log(f"🧾 Rapport TXT détaillé exporté: {txt_report_path}")

                    state.is_organizing = False
                    exec_log_label.set_text(
                        f"✅ Terminé ! {report['success']} succès | {report['trashed']} corbeille | "
                        f"{report['deleted']} supprimés | "
                        f"{report['errors']} erreurs | {report['skipped']} ignorés"
                    )
                    ui.notify(
                        f"Organisation terminée : {report['success']} organisés, "
                        f"{report['trashed']} corbeille, {report['deleted']} supprimés !",
                        type="positive"
                    )

                    # Validation explicite: .zip assignés à DOCUMENTS ont-ils bien été déplacés ?
                    expected_zip_docs = [
                        e for e in state.installer_files
                        if e.ext.lower() in {".zip", ".rar"}
                        and state.installer_decisions.get(e.path, "keep") == "assign:DOCUMENTS"
                    ]
                    moved_zip_docs = [
                        o for o in report.get("organized", [])
                        if o.get("original", "").lower().endswith((".zip", ".rar"))
                        and o.get("categorie") == "DOCUMENTS"
                        and os.path.exists(o.get("destination", ""))
                    ]
                    if expected_zip_docs:
                        state.add_log(
                            f"🔎 Vérification ZIP/RAR→DOCUMENTS : {len(moved_zip_docs)}/{len(expected_zip_docs)} déplacés."
                        )
                        if len(moved_zip_docs) != len(expected_zip_docs):
                            ui.notify(
                                "Attention: certains .zip/.rar assignés à DOCUMENTS ne se sont pas déplacés.",
                                type="warning",
                                timeout=5000,
                            )

                    # Afficher le rapport AVANT la passe métadonnées : l'utilisateur
                    # voit immédiatement le résultat de l'organisation. La passe
                    # métadonnées (potentiellement longue) tourne ensuite en arrière-plan.
                    try:
                        report_ui.refresh()
                    except Exception as e:
                        state.add_log(f"⚠️ Refresh rapport impossible : {e}")
                    try:
                        stepper.next()
                    except Exception as e:
                        state.add_log(f"⚠️ Stepper.next() impossible : {e}")
                        ui.notify(
                            "Organisation terminée. Le rafraîchissement de l'UI a échoué — "
                            "rechargez la page (F5) pour voir le rapport final.",
                            type="warning",
                            timeout=8000,
                        )

                    # Lancer la passe de métadonnées en tâche de fond (ne bloque plus l'UI)
                    async def _metadata_bg():
                        try:
                            await run_metadata_pass1(report.get("organized", []))
                            try:
                                report_ui.refresh()  # mise à jour finale avec compteurs métadonnées
                            except Exception:
                                pass
                        except Exception as e:
                            state.add_log(f"⚠️ Erreur passe métadonnées (arrière-plan) : {e}")

                    background_tasks.create(_metadata_bg(), name="organizator.metadata_pass1")

                async def run_metadata_pass1(organized: list):
                    """Passe 1 : extraction des métadonnées fichier."""
                    if not db_archive._available:
                        state.add_log("ℹ️ PostgreSQL non disponible — métadonnées non archivées.")
                        return

                    state.add_log(f"🔬 Extraction métadonnées Passe 1 : {len(organized)} fichiers...")
                    total = len(organized)
                    for i, item in enumerate(organized):
                        try:
                            # Reconstruire un ScanEntry minimal pour l'extraction
                            path = item["original"]
                            dest = item["destination"]
                            name = os.path.basename(dest)
                            ext = os.path.splitext(name)[1].lower()
                            try:
                                s = os.stat(dest)
                                size = s.st_size
                                mtime = s.st_mtime
                            except OSError:
                                size, mtime = 0, 0.0

                            entry = ScanEntry(
                                path=path, name=name, ext=ext,
                                size=size, mtime=mtime,
                                category=item["categorie"],
                                subcategory=item["sous_categorie"]
                            )
                            meta = MetadataExtractor.extract(entry)
                            await db_archive.update_metadata(path, meta, is_ai=False)
                        except Exception as e:
                            state.add_log(f"⚠️ Métadonnées {os.path.basename(item.get('original','?'))} : {e}")

                        if i % 100 == 0:
                            state.progress = i / total
                            state.status_text = f"Métadonnées Passe 1 : {i}/{total}"
                            await asyncio.sleep(0)  # Céder le contrôle

                    state.add_log("✅ Passe 1 métadonnées terminée.")

                ui.button(
                    "🚀 Lancer l'organisation",
                    icon="play_circle",
                    on_click=run_organize
                ).classes("bg-red-700 hover:bg-red-600 font-bold px-8 text-lg").props("size=lg")

                with ui.stepper_navigation():
                    ui.button("← Retour", on_click=stepper.previous).props("flat color=gray")

        # ================================================================
        # ÉTAPE 7 — RAPPORT FINAL
        # ================================================================
        with ui.step("Étape 7", title="Rapport", icon="assessment"):
            with ui.column().classes("w-full max-w-4xl mx-auto p-6 gap-6"):

                ui.label("Rapport d'organisation").classes(
                    "text-2xl font-bold text-green-400"
                )

                @ui.refreshable
                def report_ui():
                    if not state.archive_roots_text:
                        defaults = infer_default_archive_roots(state.destination, state.sources)
                        state.archive_roots_text = "\n".join(defaults)

                    async def open_pending_ai_modal():
                        if not db_archive._available:
                            ui.notify("PostgreSQL non disponible.", type="warning")
                            return

                        with ui.dialog() as dialog, ui.card().classes("w-[92vw] max-w-6xl bg-gray-900 border border-gray-700"):
                            ui.label("Fichiers en attente de validation IA").classes("text-lg font-bold text-yellow-300")
                            ui.label("Requête PostgreSQL en direct (ai_traite_le IS NULL)").classes("text-xs text-gray-400")

                            body = ui.column().classes("w-full gap-2")

                            async def load_pending_rows():
                                body.clear()
                                with body:
                                    ui.label("Chargement...").classes("text-sm text-gray-400")

                                rows = await db_archive.get_pending_ai_files(limit=1000)

                                body.clear()
                                with body:
                                    ui.label(f"Total en attente: {len(rows):,}").classes("text-sm font-semibold text-yellow-200")
                                    if not rows:
                                        ui.label("Aucun fichier en attente.").classes("text-sm text-gray-400")
                                        return

                                    with ui.scroll_area().classes("w-full h-[60vh] bg-black/30 rounded p-2"):
                                        with ui.column().classes("w-full gap-2"):
                                            for row in rows:
                                                original = row.get("chemin_original", "")
                                                destination = row.get("chemin_destination", "")
                                                cat = row.get("categorie") or "-"
                                                sub = row.get("sous_categorie") or "-"
                                                with ui.card().classes("w-full bg-gray-800/80 border border-gray-700 p-2"):
                                                    ui.label(row.get("nom_fichier", "(sans nom)")).classes("text-sm font-semibold text-cyan-200")
                                                    ui.label(f"{cat} / {sub}").classes("text-xs text-gray-300")
                                                    ui.label(original).classes("text-xs font-mono text-yellow-100 break-all")
                                                    if destination:
                                                        ui.label(destination).classes("text-xs font-mono text-green-100 break-all")

                            with ui.row().classes("w-full justify-between items-center mt-2"):
                                ui.button(
                                    "Rafraîchir",
                                    icon="refresh",
                                    on_click=load_pending_rows,
                                ).props("outline color=yellow")
                                ui.button("Fermer", icon="close", on_click=dialog.close).props("flat color=gray")

                        await load_pending_rows()
                        dialog.open()

                    async def enqueue_pending_ai_from_db():
                        if not db_archive._available:
                            ui.notify("PostgreSQL non disponible.", type="warning")
                            return
                        rows = await db_archive.get_pending_ai_files(limit=2000)
                        jobs = []
                        deferred_non_image = 0
                        for row in rows:
                            destination = (row.get("chemin_destination") or "").strip()
                            original = (row.get("chemin_original") or "").strip()
                            analyze = destination if (destination and os.path.isfile(destination)) else original
                            if not (analyze and os.path.isfile(analyze)):
                                continue
                            if not is_image_path(analyze):
                                deferred_non_image += 1
                                continue
                            jobs.append({
                                "analyze_path": analyze,
                                "original_path": original or analyze,
                                "destination_path": destination or analyze,
                            })
                        if not jobs:
                            ui.notify("Aucune image valide à envoyer à la file IA.", type="warning")
                            return
                        llm_queue.configure_validation(
                            state.llm_validate_aesthetic,
                            state.llm_validate_nsfw,
                            state.llm_validate_tags,
                            split_archive_roots(state.archive_roots_text),
                        )
                        llm_queue.enqueue_jobs(jobs)
                        msg = f"{len(jobs):,} image(s) ajoutée(s) à la file IA (queue: {llm_queue.pending_count():,})"
                        if deferred_non_image:
                            msg += f" | {deferred_non_image:,} vidéo/audio laissés en attente"
                        ui.notify(msg, type="positive")

                    with ui.card().classes("w-full bg-purple-900/20 border border-purple-700 p-4"):
                        ui.label("Mode LLM / MediaMind AI").classes("text-sm font-bold text-purple-300")
                        ui.label(
                            "Utilise d'abord l'API live media_mind_ai, puis fallback cache local si l'API est indisponible."
                        ).classes("text-xs text-gray-300")

                        with ui.card().classes("w-full bg-black/20 border border-purple-800 p-3 mt-2"):
                            ui.label("Validations requises avant archivage final").classes("text-xs font-bold text-purple-200")
                            chk_val_aes = ui.checkbox("Esthétique", value=state.llm_validate_aesthetic)
                            chk_val_nsfw = ui.checkbox("NSFW", value=state.llm_validate_nsfw)
                            chk_val_tags = ui.checkbox("Tags", value=state.llm_validate_tags)
                            archive_roots_input = ui.textarea(
                                "Racines archive (une ligne par dossier, ex: D:\\00-Archives)",
                                value=state.archive_roots_text,
                            ).classes("w-full")

                            def apply_validation_settings():
                                state.llm_validate_aesthetic = bool(chk_val_aes.value)
                                state.llm_validate_nsfw = bool(chk_val_nsfw.value)
                                state.llm_validate_tags = bool(chk_val_tags.value)
                                state.archive_roots_text = str(archive_roots_input.value or "")
                                llm_queue.configure_validation(
                                    state.llm_validate_aesthetic,
                                    state.llm_validate_nsfw,
                                    state.llm_validate_tags,
                                    split_archive_roots(state.archive_roots_text),
                                )
                                ui.notify("Paramètres de validation IA appliqués.", type="positive")

                            ui.button("Appliquer la politique de validation", on_click=apply_validation_settings).props("outline color=purple")

                        api_status_label = ui.label("API live: vérification...").classes("text-xs text-cyan-200")

                        async def refresh_llm_api_status():
                            enrich_url = os.getenv("MEDIAMIND_API_URL", "http://127.0.0.1:8190/api/llm/enrich").strip()
                            ping_url = enrich_url.replace("/api/llm/enrich", "/api/llm/ping") if "/api/llm/enrich" in enrich_url else "http://127.0.0.1:8190/api/llm/ping"

                            def _ping():
                                req = urllib_request.Request(ping_url, method="GET")
                                with urllib_request.urlopen(req, timeout=4) as resp:
                                    raw = resp.read().decode("utf-8", errors="ignore")
                                parsed = json.loads(raw) if raw else {}
                                return bool(parsed.get("ok")), parsed

                            try:
                                ok, parsed = await run.io_bound(_ping)
                                if ok:
                                    api_status_label.set_text(f"API live: connectée ({ping_url})")
                                    api_status_label.classes("text-xs text-green-300", remove="text-red-300 text-amber-300 text-cyan-200")
                                else:
                                    api_status_label.set_text(f"API live: réponse invalide, fallback cache local ({ping_url})")
                                    api_status_label.classes("text-xs text-amber-300", remove="text-green-300 text-red-300 text-cyan-200")
                            except Exception:
                                api_status_label.set_text(f"API live: hors ligne, fallback cache local ({ping_url})")
                                api_status_label.classes("text-xs text-red-300", remove="text-green-300 text-amber-300 text-cyan-200")

                        ui.timer(0.2, refresh_llm_api_status, once=True)

                        llm_queue.configure_validation(
                            state.llm_validate_aesthetic,
                            state.llm_validate_nsfw,
                            state.llm_validate_tags,
                            split_archive_roots(state.archive_roots_text),
                        )

                        with ui.row().classes("w-full gap-2 mt-2"):
                            ui.button(
                                "Tester API live",
                                icon="network_check",
                                on_click=refresh_llm_api_status,
                            ).props("outline color=cyan")
                            ui.button(
                                "Traiter la file IA en attente (BD)",
                                icon="auto_awesome",
                                on_click=enqueue_pending_ai_from_db,
                            ).props("outline color=purple")
                            ui.button(
                                "Voir fichiers en attente IA",
                                icon="pending_actions",
                                on_click=open_pending_ai_modal,
                            ).props("outline color=yellow")

                    report = state.report
                    if not report:
                        ui.label("Aucun rapport disponible.").classes("text-gray-500")
                        return

                    def open_in_explorer(path: str, select_file: bool = False):
                        if not path:
                            return
                        norm = os.path.normpath(path)
                        if select_file and os.path.exists(norm):
                            subprocess.Popen(["explorer", "/select,", norm])
                            return
                        folder = norm if os.path.isdir(norm) else os.path.dirname(norm)
                        if folder and os.path.exists(folder):
                            subprocess.Popen(["explorer", folder])

                    with ui.grid(columns=3).classes("w-full gap-4"):
                        with ui.card().classes("bg-green-900/30 border border-green-700 p-4 text-center"):
                            ui.icon("check_circle", size="2rem").classes("text-green-400 mx-auto")
                            ui.label(f"{report.get('success', 0):,}").classes(
                                "text-3xl font-bold text-green-400"
                            )
                            ui.label("Fichiers organisés").classes("text-sm text-gray-300")

                        with ui.card().classes("bg-red-900/30 border border-red-700 p-4 text-center"):
                            ui.icon("error", size="2rem").classes("text-red-400 mx-auto")
                            ui.label(f"{report.get('errors', 0):,}").classes(
                                "text-3xl font-bold text-red-400"
                            )
                            ui.label("Erreurs").classes("text-sm text-gray-300")

                        with ui.card().classes("bg-amber-900/30 border border-amber-700 p-4 text-center"):
                            ui.icon("delete_forever", size="2rem").classes("text-amber-400 mx-auto")
                            ui.label(f"{report.get('deleted', 0):,}").classes(
                                "text-3xl font-bold text-amber-400"
                            )
                            ui.label("Supprimés").classes("text-sm text-gray-300")

                        with ui.card().classes("bg-yellow-900/30 border border-yellow-700 p-4 text-center"):
                            ui.icon("delete_sweep", size="2rem").classes("text-yellow-300 mx-auto")
                            ui.label(f"{report.get('trashed', 0):,}").classes(
                                "text-3xl font-bold text-yellow-300"
                            )
                            ui.label("Corbeille").classes("text-sm text-gray-300")

                        with ui.card().classes("bg-gray-800 border border-gray-600 p-4 text-center"):
                            ui.icon("skip_next", size="2rem").classes("text-gray-400 mx-auto")
                            ui.label(f"{report.get('skipped', 0):,}").classes(
                                "text-3xl font-bold text-gray-400"
                            )
                            ui.label("Ignorés").classes("text-sm text-gray-300")

                    # Stats PostgreSQL
                    with ui.card().classes("w-full bg-gray-800 border border-gray-700 p-4 mt-2"):
                        ui.label("Base de données PostgreSQL").classes("font-bold mb-2")

                        async def load_db_stats():
                            stats = await db_archive.get_stats()
                            if not stats:
                                ui.label("PostgreSQL non disponible.").classes("text-gray-500 text-sm")
                                return
                            ui.label(
                                f"Total archivé : {stats.get('total', 0):,} fichiers"
                            ).classes("text-blue-300 text-sm")
                            ui.label(
                                f"En attente enrichissement IA : {stats.get('en_attente_ia', 0):,}"
                            ).classes("text-yellow-300 text-sm")

                        async def migrate_legacy_statuses():
                            if not db_archive._available:
                                ui.notify("PostgreSQL non disponible.", type="warning")
                                return
                            updated = await db_archive.migrate_existing_records_to_validation()
                            if updated > 0:
                                ui.notify(f"Migration faite: {updated:,} ligne(s) -> en_attente_validation", type="positive")
                            else:
                                ui.notify("Aucune ligne à migrer.", type="info")

                        ui.timer(0.1, load_db_stats, once=True)
                        with ui.row().classes("w-full gap-2 mt-2"):
                            ui.button(
                                "Migrer anciens statuts",
                                icon="rule",
                                on_click=migrate_legacy_statuses,
                            ).props("outline color=orange")

                    # Détail complet des opérations
                    with ui.expansion("Détail complet", icon="list_alt").classes("w-full bg-gray-800 border border-gray-700 mt-3"):
                        ui.label("Fichiers organisés").classes("text-sm font-bold text-green-300 mt-2")
                        organized_rows = report.get("organized", [])
                        if organized_rows:
                            with ui.scroll_area().classes("w-full h-40 bg-black/30 rounded p-2"):
                                with ui.column().classes("gap-2"):
                                    for row in organized_rows:
                                        with ui.row().classes("w-full items-start justify-between gap-2"):
                                            ui.label(
                                                f"{row.get('original')} -> {row.get('destination')}"
                                            ).classes("text-xs font-mono text-green-100 break-all")
                                            ui.button(
                                                "Ouvrir cible",
                                                icon="folder_open",
                                                on_click=lambda dest=row.get("destination", ""): open_in_explorer(dest, select_file=True),
                                            ).props("dense flat color=green")
                        else:
                            ui.label("Aucun fichier organisé.").classes("text-xs text-gray-400")

                        ui.separator().classes("my-2")
                        ui.label("Fichiers envoyés à la corbeille").classes("text-sm font-bold text-yellow-300")
                        trashed_rows = report.get("trashed_files", [])
                        if trashed_rows:
                            with ui.scroll_area().classes("w-full h-28 bg-black/30 rounded p-2"):
                                with ui.column().classes("gap-1"):
                                    for p in trashed_rows:
                                        ui.label(p).classes("text-xs font-mono text-yellow-100 break-all")
                        else:
                            ui.label("Aucun fichier en corbeille.").classes("text-xs text-gray-400")

                        ui.separator().classes("my-2")
                        ui.label("Fichiers supprimés définitivement").classes("text-sm font-bold text-red-300")
                        deleted_rows = report.get("deleted_files", [])
                        if deleted_rows:
                            with ui.scroll_area().classes("w-full h-28 bg-black/30 rounded p-2"):
                                with ui.column().classes("gap-1"):
                                    for p in deleted_rows:
                                        ui.label(p).classes("text-xs font-mono text-red-100 break-all")
                        else:
                            ui.label("Aucun fichier supprimé définitivement.").classes("text-xs text-gray-400")

                    json_path = report.get("json_export_path", "")
                    if json_path:
                        ui.separator().classes("my-2")
                        ui.label("Rapport JSON exporté").classes("text-sm font-bold text-cyan-300")
                        with ui.row().classes("w-full items-center justify-between gap-2"):
                            ui.label(json_path).classes("text-xs font-mono text-cyan-100 break-all")
                            ui.button(
                                "Ouvrir dossier JSON",
                                icon="description",
                                on_click=lambda p=json_path: open_in_explorer(p),
                            ).props("dense flat color=cyan")

                    json_compact_path = report.get("json_compact_export_path", "")
                    if json_compact_path:
                        ui.label("Rapport JSON compact exporté").classes("text-sm font-bold text-cyan-300")
                        with ui.row().classes("w-full items-center justify-between gap-2"):
                            ui.label(json_compact_path).classes("text-xs font-mono text-cyan-100 break-all")
                            ui.button(
                                "Ouvrir dossier JSON compact",
                                icon="description",
                                on_click=lambda p=json_compact_path: open_in_explorer(p),
                            ).props("dense flat color=cyan")

                    txt_detailed_path = report.get("txt_detailed_export_path", "")
                    if txt_detailed_path:
                        ui.label("Rapport TXT détaillé exporté").classes("text-sm font-bold text-cyan-300")
                        with ui.row().classes("w-full items-center justify-between gap-2"):
                            ui.label(txt_detailed_path).classes("text-xs font-mono text-cyan-100 break-all")
                            ui.button(
                                "Ouvrir dossier TXT",
                                icon="description",
                                on_click=lambda p=txt_detailed_path: open_in_explorer(p),
                            ).props("dense flat color=cyan")

                    async def enqueue_report_for_ai():
                        rows = report.get("organized", [])
                        jobs = []
                        deferred_non_image = 0
                        for item in rows:
                            destination = str(item.get("destination") or "").strip()
                            original = str(item.get("original") or "").strip()
                            analyze = destination if (destination and os.path.isfile(destination)) else original
                            if not (analyze and os.path.isfile(analyze)):
                                continue
                            if not is_image_path(analyze):
                                deferred_non_image += 1
                                continue
                            jobs.append({
                                "analyze_path": analyze,
                                "original_path": original or analyze,
                                "destination_path": destination or analyze,
                            })

                        if not jobs:
                            ui.notify("Aucune image organisée valide pour la file IA.", type="warning")
                            return

                        llm_queue.configure_validation(
                            state.llm_validate_aesthetic,
                            state.llm_validate_nsfw,
                            state.llm_validate_tags,
                            split_archive_roots(state.archive_roots_text),
                        )
                        llm_queue.enqueue_jobs(jobs)
                        msg = f"{len(jobs):,} image(s) en file IA (queue: {llm_queue.pending_count():,})"
                        if deferred_non_image:
                            msg += f" | {deferred_non_image:,} vidéo/audio laissés en attente"
                        ui.notify(msg, type="positive")

                    # Actions post-organisation
                    with ui.row().classes("gap-3 mt-4"):
                        ui.button(
                            "🤖 Lancer enrichissement LLM (Passe 2)",
                            icon="smart_toy",
                            on_click=enqueue_report_for_ai,
                        ).props("outline color=purple")

                        ui.button(
                            "Voir fichiers en attente IA",
                            icon="pending_actions",
                            on_click=open_pending_ai_modal,
                        ).props("outline color=yellow")

                        ui.button(
                            "Nouvelle organisation",
                            icon="refresh",
                            on_click=lambda: (stepper.set_value("Étape 1"), state.reset_all())
                        ).props("outline color=blue")

                        async def open_destination():
                            if state.destination and os.path.exists(state.destination):
                                subprocess.Popen(["explorer", os.path.normpath(state.destination)])
                        ui.button(
                            "Ouvrir destination",
                            icon="folder_open",
                            on_click=open_destination
                        ).props("outline color=green")

                report_ui()

                with ui.stepper_navigation():
                    ui.button("← Retour", on_click=stepper.previous).props("flat color=gray")

    # ---- Connexion PostgreSQL au démarrage ----
    async def init_db():
        db_archive.set_main_loop(asyncio.get_running_loop())
        ok = await db_archive.connect()
        if ok:
            db_status_label.set_text("BD: ✅ PostgreSQL")
            db_status_label.classes("text-xs text-green-400 font-mono", remove="text-gray-500")
            state.add_log("✅ PostgreSQL connecté.")
            await llm_queue.start()
            state.add_log("🤖 Worker LLM démarré.")
        else:
            db_status_label.set_text("BD: ⚠️ Hors ligne")

    ui.timer(0.5, init_db, once=True)


# ============================================================
# PAGE BD — EXPLORATEUR POSTGRESQL
# ============================================================

@ui.page("/bd")
async def bd_page():
    """Page de visualisation de la base de données PostgreSQL."""

    _db = PostgreSQLArchive()
    await _db.connect()

    with ui.header().classes("bg-gray-900 border-b border-gray-800 px-4 py-2"):
        with ui.row().classes("w-full items-center justify-between"):
            with ui.row().classes("items-center gap-3"):
                ui.button(icon="arrow_back", on_click=lambda: ui.navigate.to("/")).props(
                    "flat text-color=white size=sm dense"
                ).tooltip("Retour à l'app")
                ui.label("Base de données — Fichiers archivés").classes("text-sm font-bold text-cyan-300")
            with ui.row().classes("items-center gap-2"):
                total_lbl = ui.label("…").classes("text-xs text-gray-400 font-mono")
                ui.button(icon="refresh", on_click=lambda: ui.navigate.reload()).props(
                    "flat text-color=white size=sm dense"
                ).tooltip("Actualiser")

    # ---- Filtres ----
    with ui.row().classes("w-full px-4 pt-3 gap-3 flex-wrap items-end"):
        f_nom    = ui.input("Nom fichier", placeholder="G001…").classes("w-48")
        f_cat    = ui.select(
            ["(toutes)", "PHOTO-VIDEO", "MUSIQUE-AUDIO", "DOCUMENTS", "PROJETS", "DIVERS"],
            value="(toutes)", label="Catégorie"
        ).classes("w-44")
        f_statut = ui.select(
            ["(tous)", "organisé", "en_attente_validation", "archivé"],
            value="(tous)", label="Statut"
        ).classes("w-48")
        f_ai     = ui.select(
            ["(tous)", "Avec IA", "Sans IA"],
            value="(tous)", label="Métadonnées IA"
        ).classes("w-36")
        f_limit  = ui.number("Lignes max", value=200, min=10, max=2000, step=50).classes("w-32")
        search_btn = ui.button("Rechercher", icon="search").props("color=cyan outline")

    # ---- Résumé stats ----
    with ui.row().classes("w-full px-4 pt-2 gap-6 flex-wrap"):
        stats_lbl = ui.label("").classes("text-xs text-gray-400 font-mono")

    # ---- Tableau ----
    selected_ai_row = {"row": None}
    ai_provider_state = {"value": "local"}
    ollama_model_state = {"value": ""}
    ollama_models_state = {"items": []}

    columns = [
        {"name": "id",      "label": "ID",          "field": "id",      "sortable": True,  "align": "right",  "style": "width:60px"},
        {"name": "nom",     "label": "Fichier",      "field": "nom",     "sortable": True,  "align": "left",   "style": "max-width:260px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap"},
        {"name": "cat",     "label": "Catégorie",    "field": "cat",     "sortable": True,  "align": "left"},
        {"name": "ext",     "label": "Ext",          "field": "ext",     "sortable": True,  "align": "left",   "style": "width:60px"},
        {"name": "taille",  "label": "Taille",       "field": "taille",  "sortable": True,  "align": "right",  "style": "width:90px"},
        {"name": "statut",  "label": "Statut",       "field": "statut",  "sortable": True,  "align": "left"},
        {"name": "ai",      "label": "Signaux IA",   "field": "ai",      "sortable": False, "align": "left"},
        {"name": "ai_date", "label": "IA traitée le","field": "ai_date", "sortable": True,  "align": "left",   "style": "width:150px"},
        {"name": "dest",    "label": "Destination",  "field": "dest",    "sortable": False, "align": "left",   "style": "max-width:320px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap"},
    ]

    def _on_ai_row_select(e):
        selected_ai_row["row"] = e.selection[0] if e.selection else None
        ai_detail_panel.refresh()

    table = ui.table(
        columns=columns, rows=[], row_key="id", pagination={"rowsPerPage": 50}, selection="single", on_select=_on_ai_row_select
    ).classes("w-full text-xs").props("dense flat dark")

    table.add_slot("body-cell-statut", """
        <q-td :props="props">
            <q-badge
                :color="props.value === 'archivé' ? 'positive' : props.value === 'en_attente_validation' ? 'warning' : 'grey'"
                :label="props.value"
            />
        </q-td>
    """)
    table.add_slot("body-cell-ai", """
        <q-td :props="props">
            <span :class="props.value && props.value !== '—' ? 'text-green-400' : 'text-gray-500'">
                {{ props.value || '—' }}
            </span>
        </q-td>
    """)

    def _fmt_size(n):
        if n is None:
            return "—"
        if n >= 1_073_741_824:
            return f"{n/1_073_741_824:.1f} Go"
        if n >= 1_048_576:
            return f"{n/1_048_576:.1f} Mo"
        if n >= 1024:
            return f"{n/1024:.1f} Ko"
        return f"{n} o"

    def _fmt_ai(row_ai_json):
        if not row_ai_json:
            return "—"
        signals = []
        if "aesthetic" in row_ai_json:
            sc = row_ai_json["aesthetic"].get("avg_score")
            signals.append(f"aes:{sc:.2f}" if sc is not None else "aes")
        if "nsfw" in row_ai_json:
            lbl = row_ai_json["nsfw"].get("top_label", "")
            signals.append(f"nsfw:{lbl[:6]}" if lbl else "nsfw")
        if "tags" in row_ai_json:
            vals = row_ai_json["tags"].get("values", {})
            top = sorted(vals.items(), key=lambda x: -x[1])[:2] if isinstance(vals, dict) else []
            signals.append("tags:" + ",".join(t[0] for t in top) if top else "tags")
        if "prompt" in row_ai_json:
            prompt_text = str(row_ai_json["prompt"].get("text", "")).strip()
            signals.append(f"prompt:{min(len(prompt_text), 999)}c" if prompt_text else "prompt")
        if "detailed_prompt" in row_ai_json:
            detailed_text = str(row_ai_json["detailed_prompt"].get("text", "")).strip()
            signals.append(f"detail:{min(len(detailed_text), 999)}c" if detailed_text else "detail")
        if "faces" in row_ai_json:
            signals.append(f"faces:{row_ai_json['faces'].get('count', '?')}")
        return " | ".join(signals) if signals else "—"

    async def _media_api_post(endpoint: str, payload: dict) -> dict:
        def _send():
            url = f"{LLMMetadataQueue._server_base_url()}{endpoint}"
            body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
            req = urllib_request.Request(
                url,
                data=body,
                headers={"Content-Type": "application/json"},
                method="POST",
            )
            with urllib_request.urlopen(req, timeout=30) as resp:
                raw = resp.read().decode("utf-8", errors="ignore")
            return json.loads(raw) if raw else {}

        return await asyncio.get_running_loop().run_in_executor(None, _send)

    async def _media_api_get(endpoint: str) -> dict:
        def _send():
            url = f"{LLMMetadataQueue._server_base_url()}{endpoint}"
            req = urllib_request.Request(url, method="GET")
            with urllib_request.urlopen(req, timeout=20) as resp:
                raw = resp.read().decode("utf-8", errors="ignore")
            return json.loads(raw) if raw else {}

        return await asyncio.get_running_loop().run_in_executor(None, _send)

    async def _refresh_ollama_models(model_select=None):
        try:
            result = await _media_api_get("/api/llm/ollama_models")
            models = result.get("models") or []
            ollama_models_state["items"] = models
            if models and not ollama_model_state["value"]:
                ollama_model_state["value"] = models[0]
            if model_select is not None:
                model_select.options = models
                if ollama_model_state["value"]:
                    model_select.value = ollama_model_state["value"]
                model_select.update()
            if result.get("ok"):
                ui.notify(f"{len(models)} modèle(s) Ollama détecté(s).", type="positive")
            else:
                ui.notify(f"Ollama indisponible: {result.get('error', 'erreur inconnue')}", type="warning")
        except Exception as exc:
            ui.notify(f"Erreur lecture modèles Ollama: {exc}", type="negative")

    async def _persist_ai_payload(row: dict, payload_update: dict):
        if not row:
            return
        merged = dict(row.get("ai_meta_json") or {})
        merged.update(payload_update or {})
        row["ai_meta_json"] = merged
        await _db.update_metadata(row.get("orig") or row.get("dest") or "", merged, is_ai=True, set_archived=False)
        ai_detail_panel.refresh()

    async def _action_prompt_to_tags():
        row = selected_ai_row.get("row")
        if not row:
            ui.notify("Sélectionnez un fichier.", type="warning")
            return
        ai_meta = row.get("ai_meta_json") or {}
        prompt_text = str(((ai_meta.get("prompt") or {}).get("text") or "")).strip()
        if not prompt_text:
            ui.notify("Aucun prompt disponible pour ce fichier.", type="warning")
            return
        try:
            result = await _media_api_post("/api/llm/prompt_to_tags", {
                "path": row.get("orig") or row.get("dest") or "",
                "prompt": prompt_text,
                "provider": ai_provider_state["value"],
                "model": ollama_model_state["value"],
            })
            if not result.get("ok"):
                ui.notify(f"Conversion impossible: {result.get('error', 'erreur inconnue')}", type="negative")
                return
            await _persist_ai_payload(row, result.get("payload") or {})
            ui.notify(f"Tags générés ({result.get('count', 0)}). Source: {result.get('engine', 'unknown')}", type="positive")
            await do_search()
        except Exception as exc:
            ui.notify(f"Erreur conversion prompt -> tags: {exc}", type="negative")

    async def _action_generate_detailed_prompt():
        row = selected_ai_row.get("row")
        if not row:
            ui.notify("Sélectionnez un fichier.", type="warning")
            return
        ai_meta = row.get("ai_meta_json") or {}
        try:
            result = await _media_api_post("/api/llm/generate_detailed_prompt", {
                "path": row.get("orig") or row.get("dest") or "",
                "prompt": ((ai_meta.get("prompt") or {}).get("text") or ""),
                "tags": ((ai_meta.get("tags") or {}).get("values") or {}),
                "provider": ai_provider_state["value"],
                "model": ollama_model_state["value"],
            })
            if not result.get("ok"):
                ui.notify(f"Génération impossible: {result.get('error', 'erreur inconnue')}", type="negative")
                return
            await _persist_ai_payload(row, result.get("payload") or {})
            ui.notify(f"Prompt détaillé généré. Source: {result.get('engine', 'unknown')}", type="positive")
            await do_search()
        except Exception as exc:
            ui.notify(f"Erreur génération prompt: {exc}", type="negative")

    @ui.refreshable
    def ai_detail_panel():
        row = selected_ai_row.get("row")
        with ui.card().classes("w-full bg-gray-900/80 border border-gray-700 p-4 mt-3"):
            ui.label("Détails IA").classes("text-sm font-bold text-cyan-300")
            if not row:
                ui.label("Sélectionnez une ligne dans le tableau pour voir Prompt / Tags / Actions.").classes("text-sm text-gray-400")
                return

            ai_meta = row.get("ai_meta_json") or {}
            prompt_text = str(((ai_meta.get("prompt") or {}).get("text") or "")).strip()
            detailed_prompt_text = str(((ai_meta.get("detailed_prompt") or {}).get("text") or "")).strip()
            tags_values = ((ai_meta.get("tags") or {}).get("values") or {})
            sorted_tags = sorted(tags_values.items(), key=lambda x: -float(x[1])) if isinstance(tags_values, dict) else []
            tags_preview = "\n".join(f"{name}: {score:.3f}" for name, score in sorted_tags[:80]) if sorted_tags else ""

            with ui.row().classes("w-full items-start gap-4 flex-wrap"):
                with ui.column().classes("flex-1 min-w-[320px] gap-2"):
                    ui.label(f"Fichier: {row.get('nom', '—')}").classes("text-sm text-white font-medium")
                    ui.label(f"Source: {row.get('orig') or '—'}").classes("text-xs text-gray-400 break-all")
                    ui.label(f"Destination: {row.get('dest') or '—'}").classes("text-xs text-gray-400 break-all")
                    with ui.row().classes("gap-2 mt-2 items-end flex-wrap"):
                        provider_select = ui.select(
                            {"local": "Local actuel", "ollama": "Ollama"},
                            value=ai_provider_state["value"],
                            label="Moteur"
                        ).classes("w-40")
                        provider_select.on_value_change(lambda e: ai_provider_state.__setitem__("value", e.value or "local"))
                        ollama_select = ui.select(
                            ollama_models_state["items"] or [],
                            value=ollama_model_state["value"] or None,
                            label="Modèle Ollama"
                        ).classes("w-64")
                        ollama_select.bind_visibility_from(provider_select, "value", backward=lambda v: v == "ollama")
                        ollama_select.on_value_change(lambda e: ollama_model_state.__setitem__("value", e.value or ""))
                        ui.button(icon="refresh", on_click=lambda: _refresh_ollama_models(ollama_select)).props("flat color=cyan")
                    with ui.row().classes("gap-2 mt-2"):
                        ui.button("Prompt -> Tags", icon="label", on_click=_action_prompt_to_tags).props("outline color=amber")
                        ui.button("Générer prompt détaillé", icon="auto_awesome", on_click=_action_generate_detailed_prompt).props("outline color=cyan")
                    ui.label("Prompt brut").classes("text-xs uppercase tracking-wide text-gray-400 mt-2")
                    ui.textarea(value=prompt_text or "Aucun prompt disponible.").props("readonly autogrow filled").classes("w-full")
                    ui.label("Prompt détaillé").classes("text-xs uppercase tracking-wide text-gray-400 mt-2")
                    ui.textarea(value=detailed_prompt_text or "Aucun prompt détaillé généré.").props("readonly autogrow filled").classes("w-full")
                with ui.column().classes("flex-1 min-w-[320px] gap-2"):
                    ui.label("Tags modèle").classes("text-xs uppercase tracking-wide text-gray-400")
                    ui.textarea(value=tags_preview or "Aucun tag disponible.").props("readonly autogrow filled").classes("w-full")

    ai_detail_panel()

    async def do_search():
        if not _db._available:
            ui.notify("PostgreSQL non disponible", type="negative")
            return
        try:
            limit = int(f_limit.value or 200)
            
            nom_val = (f_nom.value or "").strip()
            cat_val = f_cat.value or "(toutes)"
            statut_val = f_statut.value or "(tous)"
            ai_val = f_ai.value or "(tous)"
            
            # Construire la requête SELECT de base
            select_sql = """SELECT id, nom_fichier, categorie, sous_categorie, extension, taille_octets, statut, 
                           ai_metadata_json, ai_traite_le, chemin_destination, chemin_original FROM fichiers"""
            
            where_parts = []
            if nom_val:
                where_parts.append(f"nom_fichier ILIKE '{nom_val.replace(chr(39), chr(39)+chr(39))}%'")
            if cat_val != "(toutes)":
                where_parts.append(f"categorie = '{cat_val}'")
            if statut_val != "(tous)":
                where_parts.append(f"statut = '{statut_val}'")
            if ai_val == "Avec IA":
                where_parts.append("ai_traite_le IS NOT NULL")
            elif ai_val == "Sans IA":
                where_parts.append("ai_traite_le IS NULL")
            
            where_sql = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""
            
            query = select_sql + where_sql + f" ORDER BY id DESC LIMIT {limit}"
            stats_query = (f"SELECT COUNT(*) AS total, COUNT(*) FILTER (WHERE ai_traite_le IS NOT NULL) AS avec_ia, "
                          f"COUNT(*) FILTER (WHERE statut = 'archivé') AS archives, "
                          f"COUNT(*) FILTER (WHERE statut = 'en_attente_validation') AS en_attente "
                          f"FROM fichiers{where_sql}")
            
            async with _db._pool.acquire() as conn:
                rows = await conn.fetch(query)
                stats_row = await conn.fetchrow(stats_query)

            total_count = stats_row["total"] if stats_row else 0
            total_lbl.set_text(f"{total_count} enregistrement(s)")
            if stats_row:
                stats_lbl.set_text(
                    f"Total: {stats_row['total']}  |  Avec IA: {stats_row['avec_ia']}"
                    f"  |  Archivés: {stats_row['archives']}"
                    f"  |  En attente validation: {stats_row['en_attente']}"
                    + (f"  (filtrés — {limit} max affichés)" if len(rows) == limit else "")
                )

            table_rows = []
            selected_id = (selected_ai_row.get("row") or {}).get("id")
            for r in rows:
                ai_meta = r["ai_metadata_json"] if isinstance(r["ai_metadata_json"], dict) else {}
                ai_date = r["ai_traite_le"]
                row_data = {
                    "id":      r["id"],
                    "nom":     r["nom_fichier"],
                    "cat":     f"{r['categorie'] or '—'}/{r['sous_categorie'] or '—'}",
                    "ext":     r["extension"] or "—",
                    "taille":  _fmt_size(r["taille_octets"]),
                    "statut":  r["statut"] or "—",
                    "ai":      _fmt_ai(ai_meta),
                    "ai_date": ai_date.strftime("%Y-%m-%d %H:%M") if ai_date else "—",
                    "dest":    r["chemin_destination"] or "—",
                    "orig":    r["chemin_original"] or "",
                    "ai_meta_json": ai_meta,
                }
                table_rows.append(row_data)
                if selected_id and row_data["id"] == selected_id:
                    selected_ai_row["row"] = row_data
            table.rows = table_rows
            table.update()
            ai_detail_panel.refresh()
        except Exception as exc:
            import traceback
            ui.notify(f"Erreur requête BD: {exc}", type="negative")
            state.add_log(f"⚠️ BD erreur: {traceback.format_exc()}")

    search_btn.on("click", do_search)
    # Chargement initial automatique
    ui.timer(0.3, do_search, once=True)
    ui.timer(0.6, lambda: _refresh_ollama_models(), once=True)


# ============================================================
# POINT D'ENTRÉE
# ============================================================

if __name__ in {"__main__", "__mp_main__"}:
    parser = argparse.ArgumentParser(description="Organizator — Organisateur de disques durs")
    parser.add_argument("--server-only", action="store_true", help="Mode serveur web (sans fenêtre native)")
    parser.add_argument("--host", type=str, default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8191)
    parser.add_argument(
        "--llm-mode",
        action="store_true",
        help="Ouvre directement l'Étape 7 (mode LLM / file d'attente IA).",
    )
    args, _ = parser.parse_known_args()

    START_IN_LLM_MODE = bool(args.llm_mode)

    if args.server_only:
        requested_port = args.port
        try:
            selected_port = _find_available_port(args.host, requested_port)
        except RuntimeError as e:
            print(f"❌ {e}")
            sys.exit(1)

        if selected_port != requested_port:
            print(
                f"⚠️ Port {requested_port} occupé, bascule automatique sur le port {selected_port}."
            )
        print(f"🌐 Mode serveur — Ouvrir dans le navigateur : http://{args.host}:{selected_port}")
        ui.run(
            title="Organizator",
            host=args.host,
            port=selected_port,
            native=False,
            show=False,
            dark=True,
            reload=False
        )
    else:
        ui.run(
            title="Organizator",
            port=args.port,
            native=True,
            dark=True,
            window_size=(1400, 900),
            reload=False
        )
